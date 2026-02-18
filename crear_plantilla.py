"""
Script para crear plantilla Excel de importación de clientes
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Clientes"

# Definir columnas
columnas = [
    ("nombre", "Nombre del Cliente", 30),
    ("ip_address", "IP Address", 15),
    ("plan", "Plan", 20),
    ("velocidad_download", "Velocidad Download", 18),
    ("velocidad_upload", "Velocidad Upload", 18),
    ("telefono", "Teléfono", 15),
    ("email", "Email", 25),
    ("direccion", "Dirección", 35),
    ("cedula", "Cédula/DPI", 15),
    ("dia_corte", "Día de Corte", 12),
    ("precio_mensual", "Precio Mensual (Q)", 18),
    ("estado", "Estado", 12),
]

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Escribir encabezados
for col_num, (field, header, width) in enumerate(columnas, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_num)].width = width

# Agregar datos de ejemplo
ejemplos = [
    ("Juan Pérez García", "192.168.1.100", "Básico 5Mbps", "5M", "2M", "5555-1234", "juan@email.com", "Zona 1, Ciudad", "1234567890101", 15, 150.00, "activo"),
    ("María López Hernández", "192.168.1.101", "Estándar 10Mbps", "10M", "5M", "5555-5678", "maria@email.com", "Zona 2, Ciudad", "9876543210102", 1, 200.00, "activo"),
    ("Carlos Ramírez", "192.168.1.102", "Premium 20Mbps", "20M", "10M", "5555-9012", "", "Zona 3, Ciudad", "", 10, 300.00, "activo"),
]

# Estilos para datos
data_alignment = Alignment(horizontal="left", vertical="center")
example_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

for row_num, ejemplo in enumerate(ejemplos, 2):
    for col_num, valor in enumerate(ejemplo, 1):
        cell = ws.cell(row=row_num, column=col_num, value=valor)
        cell.alignment = data_alignment
        cell.border = thin_border
        cell.fill = example_fill

# Agregar hoja de instrucciones
ws_inst = wb.create_sheet("Instrucciones")
instrucciones = [
    ("INSTRUCCIONES PARA IMPORTAR CLIENTES", ""),
    ("", ""),
    ("Columnas Obligatorias:", ""),
    ("  • nombre", "Nombre completo del cliente"),
    ("  • ip_address", "Dirección IP única (ej: 192.168.1.100)"),
    ("  • plan", "Nombre del plan de internet"),
    ("  • velocidad_download", "Velocidad de bajada (ej: 10M, 20M, 50M)"),
    ("  • velocidad_upload", "Velocidad de subida (ej: 5M, 10M, 25M)"),
    ("", ""),
    ("Columnas Opcionales:", ""),
    ("  • telefono", "Número de teléfono"),
    ("  • email", "Correo electrónico"),
    ("  • direccion", "Dirección física"),
    ("  • cedula", "Número de cédula o DPI"),
    ("  • dia_corte", "Día del mes para corte (1-31)"),
    ("  • precio_mensual", "Precio mensual en Quetzales"),
    ("  • estado", "activo, suspendido, o cortado"),
    ("", ""),
    ("Notas Importantes:", ""),
    ("  1. No modificar los nombres de las columnas en la fila 1", ""),
    ("  2. El IP debe ser único para cada cliente", ""),
    ("  3. Los ejemplos en la hoja 'Clientes' son solo de referencia", ""),
    ("  4. Puede eliminar los ejemplos antes de importar", ""),
    ("  5. Si no conoce la velocidad, use valores como: 5M, 10M, 20M", ""),
]

for row_num, (col1, col2) in enumerate(instrucciones, 1):
    ws_inst.cell(row=row_num, column=1, value=col1)
    ws_inst.cell(row=row_num, column=2, value=col2)
    if row_num == 1:
        ws_inst.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="667EEA")

ws_inst.column_dimensions['A'].width = 40
ws_inst.column_dimensions['B'].width = 50

# Guardar archivo
wb.save("plantilla_clientes.xlsx")
print("✅ Plantilla creada: plantilla_clientes.xlsx")
