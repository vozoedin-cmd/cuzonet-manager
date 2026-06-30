"""
CuzoNet Manager - Sistema de Gestión de Clientes ISP
Con funciones avanzadas: Pagos, Corte por Address List, Importar/Exportar Excel
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from io import BytesIO
import os
import json
import zipfile
import requests
import threading
import google.generativeai as genai
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configuración de Gemini AI
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'cuzonet-secret-key-2024')

# Configuración de base de datos
DATABASE_URL = os.getenv('DATABASE_URL', 'sqlite:///clientes.db')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,        # Verifica conexión antes de usarla (evita errores de conexión caída)
    'pool_recycle': 300,          # Recicla conexiones cada 5 min (evita timeouts del servidor)
    'connect_args': {'timeout': 30} if DATABASE_URL.startswith('sqlite') else {},
}

db = SQLAlchemy(app)

# ============== LOGIN MANAGER ==============
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Inicia sesión para acceder al sistema'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

def admin_required(f):
    """Decorador que requiere rol admin"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if current_user.rol != 'admin':
            flash('Acceso restringido: solo administradores', 'error')
            return redirect(url_for('listar_clientes'))
        return f(*args, **kwargs)
    return decorated_function

# ============== MODELOS ==============

class Usuario(UserMixin, db.Model):
    """Modelo de usuarios del sistema"""
    __tablename__ = 'usuarios'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    nombre = db.Column(db.String(100), default='Administrador')
    rol = db.Column(db.String(20), default='admin')  # admin, operador, vendedor
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    balance = db.Column(db.Float, default=0.0)
    router_id = db.Column(db.Integer, db.ForeignKey('config_mikrotik.id'), nullable=True)
    
    # Nuevos campos para vendedores (Fase 1 Mejoras)
    telefono = db.Column(db.String(20))
    direccion = db.Column(db.String(200)) # Puede guardar un JSON con depto, muni, aldea
    foto_perfil = db.Column(db.String(255))
    tipo_vendedor = db.Column(db.String(50), default='acumulativo') # acumulativo o prepago
    comision_tipo = db.Column(db.String(20), default='porcentaje') # porcentaje o valor_fijo
    comision_valor = db.Column(db.Float, default=0.0)
    limite_fichas = db.Column(db.Integer, default=0) # 0 = sin limite
    estado = db.Column(db.String(20), default='activo') # activo, suspendido
    permisos = db.Column(db.Text) # JSON de permisos
    ultimo_acceso = db.Column(db.DateTime, nullable=True)
    ultima_ip = db.Column(db.String(50), nullable=True)
    
    # Relaciones
    router = db.relationship('ConfigMikroTik', backref='vendedores_asignados', lazy=True)
    vouchers = db.relationship('Voucher', backref='vendedor', lazy=True)
    transacciones = db.relationship('TransaccionVendedor', backref='vendedor', lazy=True)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Cliente(db.Model):
    """Modelo de Cliente con campos adicionales para pagos y corte"""
    __tablename__ = 'clientes'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(15), nullable=False, unique=True)
    plan = db.Column(db.String(50), nullable=False)
    velocidad_download = db.Column(db.String(20), nullable=False)
    velocidad_upload = db.Column(db.String(20), nullable=False)
    telefono = db.Column(db.String(20))
    email = db.Column(db.String(100))
    direccion = db.Column(db.String(200))
    cedula = db.Column(db.String(20))
    
    # Estado y MikroTik
    estado = db.Column(db.String(20), default='activo')  # activo, suspendido, cortado
    queue_name = db.Column(db.String(100))
    mikrotik_id = db.Column(db.String(50))
    router_id = db.Column(db.Integer, db.ForeignKey('config_mikrotik.id'), nullable=True)
    
    # Fechas de pago
    dia_corte = db.Column(db.Integer, default=1)  # Día del mes para corte (1-31)
    fecha_ultimo_pago = db.Column(db.DateTime)
    fecha_proximo_pago = db.Column(db.DateTime)
    precio_mensual = db.Column(db.Float, default=0)
    saldo_pendiente = db.Column(db.Float, default=0)
    
    # Ubicación geográfica
    latitud = db.Column(db.Float)
    longitud = db.Column(db.Float)
    
    # Fechas de registro
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relación con pagos
    pagos = db.relationship('Pago', backref='cliente', lazy=True, cascade='all, delete-orphan')
    
    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'ip_address': self.ip_address,
            'plan': self.plan,
            'velocidad_download': self.velocidad_download,
            'velocidad_upload': self.velocidad_upload,
            'telefono': self.telefono,
            'email': self.email,
            'direccion': self.direccion,
            'cedula': self.cedula,
            'estado': self.estado,
            'queue_name': self.queue_name,
            'mikrotik_id': self.mikrotik_id,
            'router_id': self.router_id,
            'dia_corte': self.dia_corte,
            'fecha_ultimo_pago': self.fecha_ultimo_pago.strftime('%Y-%m-%d') if self.fecha_ultimo_pago else None,
            'fecha_proximo_pago': self.fecha_proximo_pago.strftime('%Y-%m-%d') if self.fecha_proximo_pago else None,
            'precio_mensual': self.precio_mensual,
            'saldo_pendiente': self.saldo_pendiente,
            'latitud': self.latitud,
            'longitud': self.longitud,
            'fecha_registro': self.fecha_registro.strftime('%Y-%m-%d %H:%M') if self.fecha_registro else None
        }


class Pago(db.Model):
    """Modelo de Pagos"""
    __tablename__ = 'pagos'
    
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    fecha_pago = db.Column(db.DateTime, default=datetime.utcnow)
    mes_correspondiente = db.Column(db.String(20))  # "2024-01", "2024-02", etc.
    metodo_pago = db.Column(db.String(50))  # efectivo, transferencia, etc.
    referencia = db.Column(db.String(100))  # Número de referencia/comprobante
    notas = db.Column(db.String(200))
    registrado_por = db.Column(db.String(50))
    
    def to_dict(self):
        return {
            'id': self.id,
            'cliente_id': self.cliente_id,
            'cliente_nombre': self.cliente.nombre if self.cliente else None,
            'monto': self.monto,
            'fecha_pago': self.fecha_pago.strftime('%Y-%m-%d %H:%M') if self.fecha_pago else None,
            'mes_correspondiente': self.mes_correspondiente,
            'metodo_pago': self.metodo_pago,
            'referencia': self.referencia,
            'notas': self.notas
        }


class ConfigMikroTik(db.Model):
    """Configuración de conexión a MikroTik"""
    __tablename__ = 'config_mikrotik'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), default='Principal')
    host = db.Column(db.String(100), nullable=False)
    port = db.Column(db.Integer, default=80)
    username = db.Column(db.String(50), nullable=False)
    password = db.Column(db.String(100), nullable=False)
    use_ssl = db.Column(db.Boolean, default=False)
    activo = db.Column(db.Boolean, default=True)
    tipo = db.Column(db.String(20), default='residential')  # residential, hotspot
    # Nombre del address list para clientes cortados
    address_list_cortados = db.Column(db.String(50), default='MOROSOS')
    
    # Monitoreo
    estado_online = db.Column(db.Boolean, default=True)
    ultima_caida = db.Column(db.DateTime, nullable=True)

class ConfigAlertas(db.Model):
    """Configuración de Alertas por TextMeBot"""
    __tablename__ = 'config_alertas'
    
    id = db.Column(db.Integer, primary_key=True)
    activo = db.Column(db.Boolean, default=False)
    api_key = db.Column(db.String(200), default='')
    telefono_destino = db.Column(db.String(50), default='')
    intervalo_minutos = db.Column(db.Integer, default=2)


class ConfigIA(db.Model):
    """Configuración de claves y proveedor de Inteligencia Artificial"""
    __tablename__ = 'config_ia'
    
    id = db.Column(db.Integer, primary_key=True)
    provider = db.Column(db.String(20), default='openai')  # openai, gemini
    openai_api_key = db.Column(db.String(200), default='')
    gemini_api_key = db.Column(db.String(200), default='')
    activo = db.Column(db.Boolean, default=False)

class ConfigOmada(db.Model):
    """Configuración de Omada Controller para generación de vouchers"""
    __tablename__ = 'config_omada'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), default='Omada Principal')
    url = db.Column(db.String(200), nullable=False, default='https://127.0.0.1:8043')
    username = db.Column(db.String(100), nullable=False, default='admin')
    password = db.Column(db.String(255), nullable=False, default='')
    site_id = db.Column(db.String(100), default='Default')
    activo = db.Column(db.Boolean, default=False)

class OmadaVoucher(db.Model):
    """Historial de Vouchers generados en Omada"""
    __tablename__ = 'omada_vouchers'
    
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), nullable=False, unique=True)
    duracion_valor = db.Column(db.Integer, nullable=False)
    duracion_unidad = db.Column(db.Integer, nullable=False) # 0=min, 1=hora, 2=dia
    precio = db.Column(db.Float, nullable=False, default=0.0)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True) # Nullable para los antiguos o admin
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    estado = db.Column(db.String(20), default='activo') # activo, usado, vencido, eliminado
    cliente_nombre = db.Column(db.String(100), nullable=True)
    fecha_uso = db.Column(db.DateTime, nullable=True)
    omada_id = db.Column(db.Integer, db.ForeignKey('config_omada.id'), nullable=True)
    lote = db.Column(db.String(100), nullable=True)
    
    vendedor = db.relationship('Usuario', backref=db.backref('omada_vouchers', lazy=True))
    omada = db.relationship('ConfigOmada', backref=db.backref('vouchers', lazy=True))

class Plan(db.Model):
    """Planes de internet predefinidos"""
    __tablename__ = 'planes'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), nullable=False, unique=True)
    velocidad_download = db.Column(db.String(20), nullable=False)
    velocidad_upload = db.Column(db.String(20), nullable=False)
    precio = db.Column(db.Float, default=0)
    descripcion = db.Column(db.String(200))


class ConfigUISP(db.Model):
    """Configuración de la API de UISP (Ubiquiti)"""
    __tablename__ = 'config_uisp'
    
    id = db.Column(db.Integer, primary_key=True)
    url = db.Column(db.String(200), nullable=False) # e.g. https://unms.midominio.com
    api_key = db.Column(db.String(255), nullable=False) # x-auth-token
    activo = db.Column(db.Boolean, default=True)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)


class Infraestructura(db.Model):
    """Infraestructura de red: Antenas sectoriales y estaciones"""
    __tablename__ = 'infraestructura'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(45), nullable=False)
    tipo = db.Column(db.String(50), nullable=False)  # 'Sectorial' o 'Estación'
    mikrotik_id = db.Column(db.Integer, db.ForeignKey('config_mikrotik.id'), nullable=True)
    ubicacion = db.Column(db.String(200))
    modelo = db.Column(db.String(100))
    notas = db.Column(db.String(500))
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Nuevos campos Integración UISP
    uisp_id = db.Column(db.String(100), nullable=True)
    mac = db.Column(db.String(50), nullable=True)
    estado_online = db.Column(db.Boolean, default=False)
    rssi = db.Column(db.Integer, nullable=True)
    ccq = db.Column(db.Float, nullable=True)
    airmax_quality = db.Column(db.Float, nullable=True)
    trafico_tx = db.Column(db.Float, nullable=True) # Mbps
    trafico_rx = db.Column(db.Float, nullable=True) # Mbps
    temperatura = db.Column(db.Float, nullable=True)
    cpu = db.Column(db.Float, nullable=True)
    ram = db.Column(db.Float, nullable=True)
    voltaje = db.Column(db.Float, nullable=True)
    uptime = db.Column(db.String(100), nullable=True)
    gps = db.Column(db.String(100), nullable=True)
    firmware = db.Column(db.String(50), nullable=True)
    clientes_conectados = db.Column(db.Integer, default=0)
    ultima_sincronizacion = db.Column(db.DateTime, nullable=True)
    
    # Relación
    mikrotik = db.relationship('ConfigMikroTik', backref=db.backref('infraestructuras', lazy=True))
    
    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'ip_address': self.ip_address,
            'tipo': self.tipo,
            'mikrotik_id': self.mikrotik_id,
            'mikrotik_nombre': self.mikrotik.nombre if self.mikrotik else None,
            'ubicacion': self.ubicacion,
            'modelo': self.modelo,
            'notas': self.notas,
            'fecha_registro': self.fecha_registro.strftime('%Y-%m-%d %H:%M') if self.fecha_registro else None,
            'uisp_id': self.uisp_id,
            'mac': self.mac,
            'estado_online': self.estado_online,
            'rssi': self.rssi,
            'ccq': self.ccq,
            'airmax_quality': self.airmax_quality,
            'trafico_tx': self.trafico_tx,
            'trafico_rx': self.trafico_rx,
            'temperatura': self.temperatura,
            'cpu': self.cpu,
            'ram': self.ram,
            'voltaje': self.voltaje,
            'uptime': self.uptime,
            'gps': self.gps,
            'firmware': self.firmware,
            'clientes_conectados': self.clientes_conectados,
            'ultima_sincronizacion': self.ultima_sincronizacion.strftime('%Y-%m-%d %H:%M:%S') if self.ultima_sincronizacion else None
        }


class PlanHotspot(db.Model):
    """Planes de Hotspot para vouchers"""
    __tablename__ = 'planes_hotspot'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), nullable=False)  # e.g., "1 Hora", "1 Día"
    precio = db.Column(db.Float, nullable=False)
    perfil_hotspot = db.Column(db.String(50), nullable=False)  # Nombre del perfil en MikroTik
    limit_uptime = db.Column(db.String(20))  # e.g., "01:00:00" o "1d" (opcional)
    activo = db.Column(db.Boolean, default=True)
    
    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'precio': self.precio,
            'perfil_hotspot': self.perfil_hotspot,
            'limit_uptime': self.limit_uptime,
            'activo': self.activo
        }


class LoteFichas(db.Model):
    """Lotes de fichas generadas en bloque"""
    __tablename__ = 'lotes_fichas'
    
    id = db.Column(db.Integer, primary_key=True)
    uuid = db.Column(db.String(50), nullable=False, unique=True)
    nombre = db.Column(db.String(100), nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    cantidad_solicitada = db.Column(db.Integer, nullable=False, default=0)
    router_id = db.Column(db.Integer, db.ForeignKey('config_mikrotik.id'), nullable=False)
    plan_id = db.Column(db.Integer, db.ForeignKey('planes_hotspot.id'), nullable=False)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    
    # Relaciones
    router = db.relationship('ConfigMikroTik', backref=db.backref('lotes', lazy=True))
    plan = db.relationship('PlanHotspot', backref=db.backref('lotes', lazy=True))
    vendedor = db.relationship('Usuario', foreign_keys=[vendedor_id])


class Voucher(db.Model):
    """Vouchers de Hotspot generados"""
    __tablename__ = 'vouchers'
    
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), nullable=False, unique=True)
    contrasena = db.Column(db.String(50), nullable=False)
    precio = db.Column(db.Float, nullable=False)
    plan_id = db.Column(db.Integer, db.ForeignKey('planes_hotspot.id'), nullable=False)
    router_id = db.Column(db.Integer, db.ForeignKey('config_mikrotik.id'), nullable=False)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    estado = db.Column(db.String(20), default='activo')  # activo, usado
    mikrotik_id = db.Column(db.String(50))  # ID en MikroTik (para eliminarlo luego si es necesario)
    lote_id = db.Column(db.Integer, db.ForeignKey('lotes_fichas.id'), nullable=True)
    
    # Relaciones
    plan = db.relationship('PlanHotspot', backref=db.backref('vouchers', lazy=True))
    router = db.relationship('ConfigMikroTik', backref=db.backref('vouchers', lazy=True))
    lote = db.relationship('LoteFichas', backref=db.backref('vouchers_lista', lazy=True))
    
    def to_dict(self):
        return {
            'id': self.id,
            'codigo': self.codigo,
            'contrasena': self.contrasena,
            'precio': self.precio,
            'plan_nombre': self.plan.nombre if self.plan else None,
            'router_nombre': self.router.nombre if self.router else None,
            'vendedor_nombre': self.vendedor.nombre if self.vendedor else None,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M') if self.fecha_creacion else None,
            'estado': self.estado,
            'perfil': self.plan.nombre_mikrotik if self.plan else 'default'
        }

class HotspotUserSync(db.Model):
    """Copia sincronizada de los usuarios Hotspot de MikroTik para máxima velocidad de lectura"""
    __tablename__ = 'hotspot_user_sync'
    
    id = db.Column(db.Integer, primary_key=True)
    router_id = db.Column(db.Integer, db.ForeignKey('config_mikrotik.id'), nullable=False)
    mikrotik_id = db.Column(db.String(50))
    server = db.Column(db.String(50))
    name = db.Column(db.String(100), index=True)
    password = db.Column(db.String(100))
    profile = db.Column(db.String(100), index=True)
    uptime = db.Column(db.String(50))
    bytes_in = db.Column(db.String(50))
    bytes_out = db.Column(db.String(50))
    comment = db.Column(db.String(255), index=True)
    last_sync = db.Column(db.DateTime, default=datetime.utcnow)


class TransaccionVendedor(db.Model):
    """Historial de transacciones de saldo de vendedores"""
    __tablename__ = 'transacciones_vendedor'
    
    id = db.Column(db.Integer, primary_key=True)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # carga, venta
    monto = db.Column(db.Float, nullable=False)
    descripcion = db.Column(db.String(200))
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'vendedor_id': self.vendedor_id,
            'tipo': self.tipo,
            'monto': self.monto,
            'descripcion': self.descripcion,
            'fecha': self.fecha.strftime('%Y-%m-%d %H:%M') if self.fecha else None
        }


class AuditLog(db.Model):
    """Registro de actividad del sistema"""
    __tablename__ = 'audit_log'
    
    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(50), nullable=False)
    accion = db.Column(db.String(50), nullable=False)  # crear, editar, eliminar, pago, corte, activar, login, etc.
    entidad = db.Column(db.String(50))  # cliente, pago, config, etc.
    entidad_id = db.Column(db.Integer)
    detalle = db.Column(db.String(500))
    ip_origen = db.Column(db.String(45))
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'usuario': self.usuario,
            'accion': self.accion,
            'entidad': self.entidad,
            'entidad_id': self.entidad_id,
            'detalle': self.detalle,
            'ip_origen': self.ip_origen,
            'fecha': self.fecha.strftime('%Y-%m-%d %H:%M:%S') if self.fecha else None
        }


class InventarioManualLote(db.Model):
    __tablename__ = 'inventario_manual_lote'
    id = db.Column(db.Integer, primary_key=True)
    lote = db.Column(db.String(50))
    plan = db.Column(db.String(50))
    cantidad = db.Column(db.Integer, default=0)
    stock = db.Column(db.Integer, default=0)
    vendidas = db.Column(db.Integer, default=0)
    asignadas = db.Column(db.Integer, default=0)
    vendedor_asignado = db.Column(db.String(100), default='')
    fecha_asignacion = db.Column(db.String(50), default='')
    estado = db.Column(db.String(50), default='Activo')

class InventarioManualVendedor(db.Model):
    __tablename__ = 'inventario_manual_vendedor'
    id = db.Column(db.Integer, primary_key=True)
    vendedor = db.Column(db.String(100))
    stock = db.Column(db.Integer, default=0)
    vendidas = db.Column(db.Integer, default=0)
    dinero = db.Column(db.String(50), default='Q0')
    comision = db.Column(db.String(50), default='Q0')


class AlertaOmada(db.Model):
    """Registro de alertas recibidas del webhook de Omada"""
    __tablename__ = 'alertas_omada'
    
    id = db.Column(db.Integer, primary_key=True)
    dispositivo = db.Column(db.String(100))
    estado = db.Column(db.String(50))
    mensaje = db.Column(db.String(500))
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    raw_payload = db.Column(db.Text)


def registrar_auditoria(accion, entidad=None, entidad_id=None, detalle=None):
    """Helper para registrar eventos de auditoría"""
    try:
        usuario = current_user.username if current_user.is_authenticated else 'sistema'
        ip = request.remote_addr if request else None
        log = AuditLog(
            usuario=usuario,
            accion=accion,
            entidad=entidad,
            entidad_id=entidad_id,
            detalle=detalle,
            ip_origen=ip
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        print(f"[AUDIT] Error: {e}")


# ============== API MIKROTIK ==============

# Caché del estado de MikroTik para evitar consultas repetidas
_mikrotik_status_cache = {
    'connected': False,
    'message': 'Sin verificar',
    'queue_count': 0,
    'last_check': None
}
MIKROTIK_CACHE_TTL = 60  # Segundos antes de volver a consultar al router

def limpiar_texto_mikrotik(texto):
    """Limpia acentos y caracteres especiales para MikroTik"""
    if not texto:
        return texto
    # Mapa de reemplazo de caracteres
    reemplazos = {
        'á': 'a', 'à': 'a', 'ä': 'a', 'â': 'a', 'ã': 'a',
        'é': 'e', 'è': 'e', 'ë': 'e', 'ê': 'e',
        'í': 'i', 'ì': 'i', 'ï': 'i', 'î': 'i',
        'ó': 'o', 'ò': 'o', 'ö': 'o', 'ô': 'o', 'õ': 'o',
        'ú': 'u', 'ù': 'u', 'ü': 'u', 'û': 'u',
        'ñ': 'n', 'ç': 'c',
        'Á': 'A', 'À': 'A', 'Ä': 'A', 'Â': 'A', 'Ã': 'A',
        'É': 'E', 'È': 'E', 'Ë': 'E', 'Ê': 'E',
        'Í': 'I', 'Ì': 'I', 'Ï': 'I', 'Î': 'I',
        'Ó': 'O', 'Ò': 'O', 'Ö': 'O', 'Ô': 'O', 'Õ': 'O',
        'Ú': 'U', 'Ù': 'U', 'Ü': 'U', 'Û': 'U',
        'Ñ': 'N', 'Ç': 'C'
    }
    resultado = str(texto)
    for original, reemplazo in reemplazos.items():
        resultado = resultado.replace(original, reemplazo)
    return resultado

class MikroTikAPI:
    """Clase para interactuar con MikroTik REST API"""
    
    def __init__(self, host, username, password, port=80, use_ssl=False):
        self.host = host
        self.username = username
        self.password = password
        self.port = port
        self.protocol = 'https' if use_ssl else 'http'
        self.base_url = f"{self.protocol}://{self.host}:{self.port}/rest"
        self.session = requests.Session()
        self.session.auth = (username, password)
        self.session.verify = False
    
    def test_connection(self):
        """Prueba la conexión al router"""
        try:
            # Try REST API (v7)
            response = self.session.get(f"{self.base_url}/system/identity", timeout=5)
            if response.status_code == 200:
                return True, response.json().get('name', 'MikroTik (v7 REST)')
        except Exception:
            pass
            
        try:
            # Fallback to Classic API (v6)
            import routeros_api
            connection = routeros_api.RouterOsApiPool(
                self.host, 
                username=self.username, 
                password=self.password, 
                port=self.port, 
                plaintext_login=True
            )
            api = connection.get_api()
            identity = api.get_resource('/system/identity').get()
            name = identity[0].get('name', 'MikroTik') if identity else 'MikroTik'
            connection.disconnect()
            return True, f"{name} (v6 API)"
        except Exception as e:
            return False, f"Error REST/v6: {str(e)}"
    
    def create_simple_queue(self, name, target, max_limit_download, max_limit_upload, comment=""):
        """Crea un Simple Queue en MikroTik"""
        try:
            max_limit = f"{max_limit_upload}/{max_limit_download}"
            
            # Limpiar acentos para MikroTik
            nombre_limpio = limpiar_texto_mikrotik(name)
            comentario_limpio = limpiar_texto_mikrotik(comment)
            
            data = {
                "name": nombre_limpio,
                "target": target if '/32' in target else f"{target}/32",
                "max-limit": max_limit,
                "comment": comentario_limpio
            }
            
            response = self.session.put(
                f"{self.base_url}/queue/simple",
                json=data,
                timeout=15
            )
            
            if response.status_code in [200, 201]:
                result = response.json()
                queue_id = result.get('.id', '')
                return True, queue_id
            else:
                return False, f"Error {response.status_code}: {response.text}"
                
        except Exception as e:
            return False, str(e)
    
    def update_simple_queue(self, queue_id, **kwargs):
        """Actualiza un Simple Queue existente"""
        try:
            data = {}
            if 'name' in kwargs:
                data['name'] = kwargs['name']
            if 'target' in kwargs:
                target = kwargs['target']
                data['target'] = target if '/32' in target else f"{target}/32"
            if 'max_limit_download' in kwargs and 'max_limit_upload' in kwargs:
                data['max-limit'] = f"{kwargs['max_limit_upload']}/{kwargs['max_limit_download']}"
            if 'disabled' in kwargs:
                data['disabled'] = 'yes' if kwargs['disabled'] else 'no'
            if 'comment' in kwargs:
                data['comment'] = kwargs['comment']
            
            response = self.session.patch(
                f"{self.base_url}/queue/simple/{queue_id}",
                json=data,
                timeout=15
            )
            
            return response.status_code == 200, response.text
            
        except Exception as e:
            return False, str(e)
    
    def delete_simple_queue(self, queue_id):
        """Elimina un Simple Queue"""
        try:
            response = self.session.delete(
                f"{self.base_url}/queue/simple/{queue_id}",
                timeout=15
            )
            return response.status_code in [200, 204], response.text
        except Exception as e:
            return False, str(e)

    def find_queue_by_target(self, ip_address):
        """Busca un Simple Queue por IP destino, retorna (found, queue_id)"""
        try:
            target = ip_address if '/32' in ip_address else f"{ip_address}/32"
            response = self.session.get(
                f"{self.base_url}/queue/simple",
                timeout=15
            )
            if response.status_code == 200:
                queues = response.json()
                for q in queues:
                    if q.get('target', '') == target:
                        return True, q.get('.id', '')
            return False, None
        except Exception as e:
            return False, None

    def get_simple_queues(self):
        """Obtiene todos los Simple Queues"""
        try:
            response = self.session.get(
                f"{self.base_url}/queue/simple",
                timeout=15
            )
            if response.status_code == 200:
                return True, response.json()
            return False, f"Error {response.status_code}"
        except Exception as e:
            return False, str(e)
    
    def suspend_queue(self, queue_id):
        """Suspende (deshabilita) un queue"""
        return self.update_simple_queue(queue_id, disabled=True)
    
    def activate_queue(self, queue_id):
        """Activa (habilita) un queue"""
        return self.update_simple_queue(queue_id, disabled=False)
    
    # ============== ADDRESS LIST METHODS ==============
    
    def add_to_address_list(self, ip_address, list_name="MOROSOS", comment=""):
        """Agrega una IP al address list (para corte por firewall)"""
        try:
            data = {
                "list": list_name,
                "address": ip_address,
                "comment": comment
            }
            
            response = self.session.put(
                f"{self.base_url}/ip/firewall/address-list",
                json=data,
                timeout=15
            )
            
            if response.status_code in [200, 201]:
                result = response.json()
                return True, result.get('.id', '')
            else:
                return False, f"Error {response.status_code}: {response.text}"
                
        except Exception as e:
            return False, str(e)
    
    def remove_from_address_list(self, ip_address, list_name="MOROSOS"):
        """Remueve una IP del address list"""
        try:
            # Primero buscar el ID de la entrada
            response = self.session.get(
                f"{self.base_url}/ip/firewall/address-list",
                params={"list": list_name, "address": ip_address},
                timeout=15
            )
            
            if response.status_code == 200:
                entries = response.json()
                for entry in entries:
                    if entry.get('address') == ip_address and entry.get('list') == list_name:
                        entry_id = entry.get('.id')
                        # Eliminar la entrada
                        del_response = self.session.delete(
                            f"{self.base_url}/ip/firewall/address-list/{entry_id}",
                            timeout=15
                        )
                        return del_response.status_code in [200, 204], "OK"
            
            return True, "No encontrado (ya eliminado)"
            
        except Exception as e:
            return False, str(e)
    
    def get_address_list(self, list_name="MOROSOS"):
        """Obtiene todas las IPs en un address list"""
        try:
            response = self.session.get(
                f"{self.base_url}/ip/firewall/address-list",
                params={"list": list_name},
                timeout=15
            )
            if response.status_code == 200:
                return True, response.json()
            return False, f"Error {response.status_code}"
        except Exception as e:
            return False, str(e)

    # ============== HOTSPOT METHODS ==============

    def get_hotspot_users(self, profile=None):
        """Obtiene la lista de usuarios usando API Clásica (ultra rápida) por defecto"""
        # Try Classic API (v6/v7) first for speed
        try:
            import routeros_api
            connection = routeros_api.RouterOsApiPool(self.host, username=self.username, password=self.password, port=self.port, plaintext_login=True)
            api = connection.get_api()
            
            resource = api.get_resource('/ip/hotspot/user')
            if profile:
                users = resource.get(**{'profile': profile})
            else:
                users = resource.get()
                
            connection.disconnect()
            return True, users
        except Exception as e:
            pass # Si falla, intenta REST
            
        try:
            # Fallback a REST API (v7)
            url = f"{self.base_url}/ip/hotspot/user"
            params = {}
            if profile:
                params['profile'] = profile
            response = self.session.get(url, params=params, timeout=10)
            
            if response.status_code == 200:
                users = response.json()
                if profile:
                    users = [u for u in users if u.get('profile') == profile]
                return True, users
            elif response.status_code == 404:
                raise Exception("REST API not found")
        except Exception as e:
            return False, str(e)

    def get_hotspot_profiles_with_counts(self):
        """Obtiene perfiles y la cantidad de usuarios usando API Clásica por velocidad"""
        try:
            import routeros_api
            connection = routeros_api.RouterOsApiPool(self.host, username=self.username, password=self.password, port=self.port, plaintext_login=True)
            api = connection.get_api()
            
            profiles = api.get_resource('/ip/hotspot/user/profile').get()
            users = api.get_resource('/ip/hotspot/user').get()
            
            counts = {}
            for u in users:
                p = u.get('profile', 'default')
                counts[p] = counts.get(p, 0) + 1
                
            result = []
            for p in profiles:
                name = p.get('name')
                if name:
                    result.append({
                        'name': name,
                        'count': counts.get(name, 0)
                    })
            connection.disconnect()
            return True, result
        except Exception:
            pass # Si falla, intenta REST
            
        try:
            # Fallback REST API
            res_profiles = self.session.get(f"{self.base_url}/ip/hotspot/user/profile", timeout=5)
            res_users = self.session.get(f"{self.base_url}/ip/hotspot/user", timeout=5)
            
            if res_profiles.status_code == 200 and res_users.status_code == 200:
                profiles = res_profiles.json()
                users = res_users.json()
                
                counts = {}
                for u in users:
                    p = u.get('profile', 'default')
                    counts[p] = counts.get(p, 0) + 1
                    
                result = []
                for p in profiles:
                    name = p.get('name')
                    if name:
                        result.append({
                            'name': name,
                            'count': counts.get(name, 0)
                        })
                return True, result
            elif res_profiles.status_code == 404:
                raise Exception("REST API not found")
        except Exception as e:
            return False, str(e)
    
    def create_hotspot_users_batch(self, users_data):
        """
        Crea múltiples usuarios de forma ultra-rápida (Ráfaga) mediante un script en MikroTik.
        users_data es una lista de diccionarios con las llaves: name, password, profile, limit_uptime, limit_bytes_total, comment
        Retorna (True, 'ok') o (False, 'error').
        """
        import random
        if not users_data:
            return True, "No hay usuarios"
            
        # Generar el script de RouterOS
        script_lines = ["/ip hotspot user"]
        for u in users_data:
            line = f"add name=\"{u['name']}\" password=\"{u['password']}\" profile=\"{u['profile']}\""
            if u.get('comment'):
                line += f" comment=\"{u['comment']}\""
            if u.get('limit_uptime'):
                line += f" limit-uptime=\"{u['limit_uptime']}\""
            if u.get('limit_bytes_total'):
                line += f" limit-bytes-total=\"{u['limit_bytes_total']}\""
            script_lines.append(line)
            
        script_content = "\n".join(script_lines)
        script_name = f"cuzonet_batch_{random.randint(1000, 9999)}"
        
        # 1. Crear el script en system/script
        url_add = f"{self.base_url}/system/script"
        script_data = {
            "name": script_name,
            "source": script_content,
            "policy": "read,write,policy,test"
        }
        
        try:
            # Crear script usando REST API v7 (usamos put o post via requests)
            import requests
            res_add = requests.put(url_add, json=script_data, auth=(self.username, self.password), verify=False, timeout=15)
            if res_add.status_code not in [200, 201]:
                return False, f"Error creando script: {res_add.text}"
                
            script_id = res_add.json().get('.id', script_name)
            
            # 2. Ejecutar script
            url_run = f"{self.base_url}/system/script/run"
            requests.post(url_run, json={".id": script_id}, auth=(self.username, self.password), verify=False, timeout=10)
            
            # 3. Eliminar script para no dejar basura
            url_del = f"{self.base_url}/system/script/{script_id}"
            requests.delete(url_del, auth=(self.username, self.password), verify=False, timeout=5)
            
            return True, "Batch completado"
        except Exception as e:
            return False, str(e)

    def create_hotspot_user(self, name, password, profile, comment="", limit_uptime=None, limit_bytes_total=None):
        """Crea un usuario de Hotspot en MikroTik"""
        nombre_limpio = limpiar_texto_mikrotik(name)
        data = {
            "name": nombre_limpio,
            "password": password,
            "profile": profile,
            "comment": limpiar_texto_mikrotik(comment)
        }
        if limit_uptime:
            data["limit-uptime"] = limit_uptime
        if limit_bytes_total:
            data["limit-bytes-total"] = limit_bytes_total
            
        try:
            # Try REST API (v7)
            response = self.session.put(
                f"{self.base_url}/ip/hotspot/user",
                json=data,
                timeout=5
            )
            
            if response.status_code in [200, 201]:
                result = response.json()
                return True, result.get('.id', '')
            elif response.status_code == 404:
                raise Exception("REST API not found, try v6 fallback")
        except Exception:
            pass
            
        # Fallback to Classic API (v6)
        try:
            import routeros_api
            connection = routeros_api.RouterOsApiPool(
                self.host, 
                username=self.username, 
                password=self.password, 
                port=self.port, 
                plaintext_login=True
            )
            api = connection.get_api()
            api.get_resource('/ip/hotspot/user').add(**data)
            connection.disconnect()
            return True, "Created via v6 API"
        except Exception as e:
            return False, f"Error v6: {str(e)}"

    def delete_hotspot_user(self, user_id):
        """Elimina un usuario de Hotspot en MikroTik por su ID (.id)"""
        try:
            response = self.session.delete(
                f"{self.base_url}/ip/hotspot/user/{user_id}",
                timeout=10
            )
            return response.status_code in [200, 204], "Eliminado" if response.status_code in [200, 204] else f"Error {response.status_code}"
        except Exception as e:
            return False, str(e)
            
    def get_hotspot_profiles(self):
        """Obtiene la lista de perfiles de usuario usando API Clásica"""
        try:
            import routeros_api
            connection = routeros_api.RouterOsApiPool(self.host, username=self.username, password=self.password, port=self.port, plaintext_login=True)
            api = connection.get_api()
            profiles = api.get_resource('/ip/hotspot/user/profile').get()
            connection.disconnect()
            return True, [p.get('name') for p in profiles if p.get('name') and p.get('name') != 'default']
        except Exception:
            pass # Si falla, intenta REST
            
        try:
            # Fallback REST API
            response = self.session.get(f"{self.base_url}/ip/hotspot/user/profile", timeout=5)
            if response.status_code == 200:
                profiles = response.json()
                return True, [p.get('name') for p in profiles if p.get('name') and p.get('name') != 'default']
            elif response.status_code == 404:
                raise Exception("REST API not found")
        except Exception as e:
            return False, f"Error obteniendo perfiles: {str(e)}"

    def get_live_dashboard_data(self):
        """Obtiene datos en vivo para el Dashboard de Hotspot (Estilo Mikhmon)"""
        data = {
            'cpu': 0, 'mem_libre': '0 MB', 'hdd_libre': '0 MB',
            'uptime': '0s', 'version': '', 'board': '',
            'active_users': 0, 'total_users': 0,
            'logs': [], 'error': None
        }
        
        # 1. Intentar con REST API (v7)
        try:
            res_resource = self.session.get(f"{self.base_url}/system/resource", timeout=5)
            if res_resource.status_code == 200:
                res = res_resource.json()
                data['cpu'] = int(res.get('cpu-load', 0))
                data['mem_libre'] = f"{round(int(res.get('free-memory', 0)) / 1048576, 1)} MB"
                data['hdd_libre'] = f"{round(int(res.get('free-hdd-space', 0)) / 1048576, 1)} MB"
                data['uptime'] = res.get('uptime', '')
                data['version'] = res.get('version', '')
                data['board'] = res.get('board-name', '')
                
                # Active users
                res_active = self.session.get(f"{self.base_url}/ip/hotspot/active", timeout=5)
                if res_active.status_code == 200:
                    data['active_users'] = len(res_active.json())
                    
                # Total users
                res_users = self.session.get(f"{self.base_url}/ip/hotspot/user", timeout=5)
                if res_users.status_code == 200:
                    # 'default' is usually a profile, but the user list might include a default user. We just count all.
                    data['total_users'] = len(res_users.json())
                    
                # Logs
                res_logs = self.session.get(f"{self.base_url}/log?topics=hotspot", timeout=5)
                if res_logs.status_code == 200:
                    logs = res_logs.json()
                    # Filter for hotspot topics if query param didn't work perfectly
                    hotspot_logs = [l for l in logs if 'hotspot' in l.get('topics', '')]
                    hotspot_logs = hotspot_logs[-15:]
                    hotspot_logs.reverse()
                    for l in hotspot_logs:
                        data['logs'].append({
                            'time': l.get('time', ''),
                            'message': l.get('message', '')
                        })
                
                return data
        except Exception:
            pass
            
        # 2. Fallback a Classic API (v6)
        try:
            import routeros_api
            connection = routeros_api.RouterOsApiPool(
                self.host, username=self.username, password=self.password, 
                port=self.port, plaintext_login=True
            )
            api = connection.get_api()
            
            # System Resources
            resource = api.get_resource('/system/resource').get()[0]
            data['cpu'] = int(resource.get('cpu-load', 0))
            data['mem_libre'] = f"{round(int(resource.get('free-memory', 0)) / 1048576, 1)} MB"
            data['hdd_libre'] = f"{round(int(resource.get('free-hdd-space', 0)) / 1048576, 1)} MB"
            data['uptime'] = resource.get('uptime', '')
            data['version'] = resource.get('version', '')
            data['board'] = resource.get('board-name', '')
            
            # Hotspot Users
            data['active_users'] = len(api.get_resource('/ip/hotspot/active').get())
            data['total_users'] = len(api.get_resource('/ip/hotspot/user').get())
            
            # Logs
            logs = api.get_resource('/log').get()
            hotspot_logs = [l for l in logs if 'hotspot' in l.get('topics', '')]
            hotspot_logs = hotspot_logs[-15:]
            hotspot_logs.reverse()
            for l in hotspot_logs:
                data['logs'].append({
                    'time': l.get('time', ''),
                    'message': l.get('message', '')
                })
                
            connection.disconnect()
        except Exception as e:
            data['error'] = str(e)
            
        return data


def get_mikrotik_api(router_id=None):
    """Obtiene instancia de la API de MikroTik con la configuración activa o por ID"""
    if router_id:
        config = ConfigMikroTik.query.get(router_id)
    else:
        config = ConfigMikroTik.query.filter_by(activo=True).first()
    if not config:
        return None
    return MikroTikAPI(
        host=config.host,
        username=config.username,
        password=config.password,
        port=config.port,
        use_ssl=config.use_ssl
    )


def get_address_list_name(router_id=None):
    """Obtiene el nombre del address list configurado"""
    if router_id:
        config = ConfigMikroTik.query.get(router_id)
    else:
        config = ConfigMikroTik.query.filter_by(activo=True).first()
    return config.address_list_cortados if config else "MOROSOS"


# ============== RUTAS WEB ==============

# ============== RUTAS DE AUTENTICACIÓN ==============

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        user = Usuario.query.filter_by(username=username).first()
        
        if user and user.check_password(password) and user.activo:
            user.ultimo_acceso = datetime.utcnow()
            user.ultima_ip = request.remote_addr
            db.session.commit()
            
            login_user(user, remember=True)
            registrar_auditoria('login', 'usuario', user.id, f'Inicio de sesión: {username}')
            if user.rol == 'vendedor':
                return redirect(url_for('vendedor_dashboard'))
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/api/cambiar-password', methods=['POST'])
@login_required
def cambiar_password():
    data = request.get_json()
    current_pw = data.get('current_password', '')
    new_pw = data.get('new_password', '')
    
    if not current_user.check_password(current_pw):
        return jsonify({'success': False, 'error': 'Contraseña actual incorrecta'})
    if len(new_pw) < 4:
        return jsonify({'success': False, 'error': 'La nueva contraseña debe tener al menos 4 caracteres'})
    
    current_user.set_password(new_pw)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Contraseña actualizada'})

# ============== GESTIÓN DE USUARIOS ==============

@app.route('/usuarios')
@login_required
@admin_required
def usuarios_view():
    """Página de gestión de usuarios"""
    usuarios = Usuario.query.order_by(Usuario.fecha_creacion.desc()).all()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/api/usuarios', methods=['GET'])
@login_required
@admin_required
def api_listar_usuarios():
    usuarios = Usuario.query.order_by(Usuario.fecha_creacion.desc()).all()
    return jsonify({'success': True, 'usuarios': [{
        'id': u.id,
        'username': u.username,
        'nombre': u.nombre,
        'rol': u.rol,
        'activo': u.activo,
        'fecha_creacion': u.fecha_creacion.strftime('%Y-%m-%d') if u.fecha_creacion else None
    } for u in usuarios]})

@app.route('/api/usuarios', methods=['POST'])
@login_required
@admin_required
def api_crear_usuario():
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        nombre = data.get('nombre', '').strip()
        password = data.get('password', '')
        rol = data.get('rol', 'operador')

        if not username or not password or not nombre:
            return jsonify({'success': False, 'error': 'Username, nombre y contraseña son requeridos'})
        if len(password) < 4:
            return jsonify({'success': False, 'error': 'La contraseña debe tener al menos 4 caracteres'})
        if Usuario.query.filter_by(username=username).first():
            return jsonify({'success': False, 'error': f'El usuario "{username}" ya existe'})
        if rol not in ('admin', 'operador'):
            rol = 'operador'

        nuevo = Usuario(username=username, nombre=nombre, rol=rol)
        nuevo.set_password(password)
        db.session.add(nuevo)
        db.session.commit()
        registrar_auditoria('crear', 'usuario', nuevo.id, f'Nuevo usuario: {username} ({rol})')
        return jsonify({'success': True, 'message': f'Usuario "{username}" creado correctamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/usuario/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
def api_editar_usuario(user_id):
    try:
        u = Usuario.query.get_or_404(user_id)
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        rol = data.get('rol', u.rol)
        new_pw = data.get('password', '')

        if nombre:
            u.nombre = nombre
        if rol in ('admin', 'operador'):
            u.rol = rol
        if new_pw:
            if len(new_pw) < 4:
                return jsonify({'success': False, 'error': 'La contraseña debe tener al menos 4 caracteres'})
            u.set_password(new_pw)
        db.session.commit()
        registrar_auditoria('editar', 'usuario', u.id, f'Editado: {u.username}')
        return jsonify({'success': True, 'message': 'Usuario actualizado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/usuario/<int:user_id>/toggle', methods=['PUT'])
@login_required
@admin_required
def api_toggle_usuario(user_id):
    try:
        u = Usuario.query.get_or_404(user_id)
        if u.id == current_user.id:
            return jsonify({'success': False, 'error': 'No puedes desactivarte a ti mismo'})
        u.activo = not u.activo
        db.session.commit()
        estado = 'activado' if u.activo else 'desactivado'
        registrar_auditoria('editar', 'usuario', u.id, f'Usuario {estado}: {u.username}')
        return jsonify({'success': True, 'activo': u.activo, 'message': f'Usuario {estado}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/usuario/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def api_eliminar_usuario(user_id):
    try:
        u = Usuario.query.get_or_404(user_id)
        if u.id == current_user.id:
            return jsonify({'success': False, 'error': 'No puedes eliminar tu propio usuario'})
        username = u.username
        db.session.delete(u)
        db.session.commit()
        registrar_auditoria('eliminar', 'usuario', user_id, f'Eliminado: {username}')
        return jsonify({'success': True, 'message': f'Usuario "{username}" eliminado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# ============== CONFIGURACION OMADA API ==============

@app.route('/api/config/omada', methods=['POST'])
@login_required
@admin_required
def save_config_omada():
    """Guardar configuración de Omada"""
    try:
        data = request.json
        omada_id = data.get('id')
        
        if omada_id:
            config = ConfigOmada.query.get(omada_id)
            if not config:
                return jsonify({'success': False, 'error': 'Controlador Omada no encontrado'})
        else:
            config = ConfigOmada()
            db.session.add(config)
            
        config.nombre = data.get('nombre', 'Omada Principal').strip()
        config.url = data.get('url', '').strip()
        config.username = data.get('username', '').strip()
        
        # Solo actualizar password si se envía uno nuevo
        if data.get('password'):
            config.password = data.get('password')
            
        config.site_id = data.get('site_id', 'Default').strip()
        config.activo = data.get('activo', False)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Configuración de Omada guardada exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/config/omada/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def delete_config_omada(id):
    """Eliminar un controlador Omada"""
    try:
        config = ConfigOmada.query.get_or_404(id)
        # Opcional: verificar si tiene vouchers
        db.session.delete(config)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Controlador eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/omada/test', methods=['POST'])
@login_required
@admin_required
def test_omada_route():
    try:
        data = request.json
        omada_id = data.get('omada_id')
        
        if omada_id:
            config = ConfigOmada.query.get(omada_id)
        else:
            # Si prueban la conexion sin haber guardado aun, usan los datos enviados
            return test_omada_temp(data)
            
        if not config or not config.activo:
            return jsonify({'success': False, 'error': 'La integración de Omada no está activa'})
            
        from omada_api import OmadaAPI
        api = OmadaAPI(config.url, config.username, config.password, config.site_id)
        success, msg = api.test_connection()
        if success:
            return jsonify({'success': True, 'message': msg})
        else:
            return jsonify({'success': False, 'error': msg})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

def test_omada_temp(data):
    try:
        from omada_api import OmadaAPI
        api = OmadaAPI(data.get('url'), data.get('username'), data.get('password'), data.get('site_id'))
        success, msg = api.test_connection()
        if success:
            return jsonify({'success': True, 'message': msg})
        else:
            return jsonify({'success': False, 'error': msg})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/omada/sites', methods=['POST'])
@login_required
@admin_required
def get_omada_sites():
    try:
        data = request.json
        password = data.get('password', '')
        omada_id = data.get('omada_id')
        
        if not password and omada_id:
            config = ConfigOmada.query.get(omada_id)
            if config:
                password = config.password

        from omada_api import OmadaAPI
        api = OmadaAPI(data.get('url'), data.get('username'), password)
        sites = api.get_all_sites()
        return jsonify({'success': True, 'sites': sites})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/omada/<int:id>/sites', methods=['GET'])
@login_required
@admin_required
def get_omada_sites_by_id(id):
    try:
        config = ConfigOmada.query.get_or_404(id)
        if not config.activo:
            return jsonify({'success': False, 'error': 'Controlador inactivo'})
        from omada_api import OmadaAPI
        api = OmadaAPI(config.url, config.username, config.password)
        sites = api.get_all_sites()
        return jsonify({'success': True, 'sites': sites})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/omada/generar', methods=['POST'])
@login_required
@admin_required
def generar_fichas_omada():
    try:
        data = request.json
        cantidad = int(data.get('cantidad', 10))
        tiempo = int(data.get('tiempo', 3))
        unidad = int(data.get('unidad', 1)) # 0=min, 1=hora, 2=dia
        precio = float(data.get('precio', 0.0))
        vendedor_id = data.get('vendedor_id')
        if not vendedor_id:
            return jsonify({'success': False, 'error': 'Debe seleccionar un Vendedor. No se pueden generar fichas sin asignar.'})
            
        try:
            vendedor_id = int(vendedor_id)
        except ValueError:
            return jsonify({'success': False, 'error': 'Vendedor inválido.'})
            
        omada_id = data.get('omada_id')
        if not omada_id:
            return jsonify({'success': False, 'error': 'Debe seleccionar un controlador Omada'})
            
        config = ConfigOmada.query.get(omada_id)
        if not config or not config.activo:
            return jsonify({'success': False, 'error': 'Controlador Omada no encontrado o inactivo'})
            
        # Omada V6 espera el tiempo preferiblemente en minutos
        if unidad == 0:
            minutos = tiempo
        elif unidad == 1:
            minutos = tiempo * 60
        else: # dias
            minutos = tiempo * 1440
            
        site_name_override = data.get('site_name')
        final_site_name = site_name_override if site_name_override else config.site_id
        
        comentario = data.get('comentario', '').strip()
        if not comentario and vendedor_id:
            v = Usuario.query.get(vendedor_id)
            if v:
                comentario = v.nombre
            
        from omada_api import OmadaAPI
        api = OmadaAPI(config.url, config.username, config.password, final_site_name)
        pines = api.generar_fichas(cantidad, minutos, 1, comentario) # Pasamos minutos y unidad=1 (minutos en V6)
        
        if not pines:
            return jsonify({'success': False, 'error': 'Se ejecutó el comando pero no se obtuvieron los PINs en la respuesta.'})
            
        # Guardar en base de datos
        nuevos_vouchers = []
        for pin in pines:
            # Si el código ya existe (Omada recicló un PIN de una ficha eliminada antigua),
            # renombramos la antigua para no perder su historial de venta/deuda, pero liberamos el PIN.
            existing = OmadaVoucher.query.filter_by(codigo=pin).first()
            if existing:
                existing.codigo = f"{existing.codigo}_old_{existing.id}"
                db.session.add(existing)
                db.session.flush() # Forzar el update en la base de datos para liberar el unique constraint
                
            v = OmadaVoucher(
                codigo=pin,
                duracion_valor=tiempo,
                duracion_unidad=unidad,
                precio=precio,
                vendedor_id=vendedor_id,
                omada_id=omada_id,
                lote=comentario
            )
            db.session.add(v)
            nuevos_vouchers.append(v)
        
        db.session.commit()
            
        return jsonify({'success': True, 'pines': pines})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/hotspot/omada-debug')
@login_required
def omada_debug():
    config = ConfigOmada.query.filter_by(activo=True).first()
    if not config: return "No active config"
    from omada_api import OmadaAPI
    omada = OmadaAPI(config.url, config.username, config.password, config.site_id)
    try:
        sites = omada.get_all_sites()
        if not sites: return "No sites"
        
        all_vouchers = []
        for site in sites:
            omada.site_id = site['id']
            query_url = f"{omada.base_url}/{omada.omadac_id}/api/v2/hotspot/sites/{omada.site_id}/vouchers?currentPage=1&currentPageSize=50"
            res = omada.session.get(query_url, timeout=10)
            data = res.json()
            if data.get('errorCode') == 0:
                vouchers = data.get('result', {}).get('data', [])
                if vouchers:
                    all_vouchers.extend(vouchers)
                    
        return jsonify({
            'total_found': len(all_vouchers),
            'first_5': all_vouchers[:5]
        })
    except Exception as e:
        return str(e)

@app.route('/api/omada/sync', methods=['POST'])
@login_required
def omada_sync_api():
    from datetime import datetime
    omadas_configs = ConfigOmada.query.all()
    # Filtrar en python para evitar problemas con SQLite y booleanos
    omadas_configs = [c for c in omadas_configs if c.activo]
    
    if not omadas_configs:
        return jsonify({'success': False, 'error': 'No hay controladores Omada activos configurados.'})
        
    changed = False
    total_sync = 0
    errores_sync = []
    
    from omada_api import OmadaAPI
    status_map_global = {}
    
    for config in omadas_configs:
        try:
            omada = OmadaAPI(config.url, config.username, config.password, config.site_id)
            sites = omada.get_all_sites()
            for site in sites:
                omada.site_id = site['id']
                try:
                    status_map = omada.get_all_vouchers_status()
                    status_map_global.update(status_map)
                    total_sync += len(status_map)
                except Exception as e:
                    errores_sync.append(f"{config.nombre} ({site['name']}): {str(e)}")
        except Exception as e:
            errores_sync.append(f"{config.nombre} (Conexión/Sitios): {str(e)}")
            
    # Solo procesar si no hubo errores de conexion para evitar falsos "eliminados"
    eliminados_count = 0
    actualizados_count = 0
    if not errores_sync:
        all_vouchers_local = OmadaVoucher.query.all()
        for v_local in all_vouchers_local:
            if v_local.codigo in status_map_global:
                try:
                    omada_status = int(status_map_global[v_local.codigo])
                except ValueError:
                    omada_status = 0
                    
                nuevo_estado = 'activo'
                if omada_status == 1:
                    nuevo_estado = 'usado'
                    if not v_local.fecha_uso:
                        from datetime import datetime
                        v_local.fecha_uso = datetime.utcnow()
                elif omada_status in (2, 3, 4):
                    nuevo_estado = 'vencido'
                    
                if v_local.estado != nuevo_estado:
                    v_local.estado = nuevo_estado
                    changed = True
                    actualizados_count += 1
            else:
                if v_local.estado != 'eliminado':
                    v_local.estado = 'eliminado'
                    changed = True
                    eliminados_count += 1
                    
    if changed:
        db.session.commit()
        
    mensaje = f"Sincronización completada. Actualizados: {actualizados_count}. Eliminados: {eliminados_count}."
    if errores_sync:
        mensaje += f" [Advertencia: Algunos sitios fallaron: {' | '.join(errores_sync)}]"
        
    return jsonify({'success': True, 'message': mensaje})

@app.route('/api/omada/debug_sync', methods=['GET'])
@login_required
def debug_omada_sync():
    total = OmadaVoucher.query.count()
    activos = OmadaVoucher.query.filter_by(estado='activo').count()
    eliminados = OmadaVoucher.query.filter_by(estado='eliminado').count()
    usados = OmadaVoucher.query.filter_by(estado='usado').count()
    vencidos = OmadaVoucher.query.filter_by(estado='vencido').count()
    
    # Check if they have vendedor_id
    sin_vendedor = OmadaVoucher.query.filter(OmadaVoucher.vendedor_id == None).count()
    
    return jsonify({
        'total_vouchers': total,
        'estados': {
            'activo': activos,
            'eliminado': eliminados,
            'usado': usados,
            'vencido': vencidos
        },
        'sin_vendedor': sin_vendedor
    })

@app.route('/api/omada/stats_vendedores', methods=['GET'])
@login_required
@admin_required
def stats_vendedores_omada():
    try:
        from sqlalchemy import func, case
        # Agrupar por vendedor_id, lote, duracion, precio
        stats = db.session.query(
            Usuario.id,
            Usuario.nombre,
            OmadaVoucher.lote,
            OmadaVoucher.duracion_valor,
            OmadaVoucher.duracion_unidad,
            OmadaVoucher.precio,
            func.count(OmadaVoucher.id).label('total_fichas'),
            func.sum(case((OmadaVoucher.estado == 'activo', 1), else_=0)).label('disponibles'),
            func.sum(case((OmadaVoucher.estado != 'activo', 1), else_=0)).label('vendidas')
        ).join(OmadaVoucher, Usuario.id == OmadaVoucher.vendedor_id)\
         .filter(OmadaVoucher.estado != 'eliminado')\
         .group_by(Usuario.id, OmadaVoucher.lote, OmadaVoucher.duracion_valor, OmadaVoucher.duracion_unidad, OmadaVoucher.precio).all()
         
        resultado = {}
        unidades = {0: 'Min', 1: 'Hora', 2: 'Día'}
        
        for u_id, nombre, lote, d_val, d_un, precio, total, disp, vend in stats:
            if nombre not in resultado:
                resultado[nombre] = []
                
            unidad_str = unidades.get(d_un, 'U')
            if d_val > 1 and d_un == 1: unidad_str = 'Horas'
            if d_val > 1 and d_un == 2: unidad_str = 'Días'
            if d_val > 1 and d_un == 0: unidad_str = 'Mins'
            
            plan_str = f"{d_val} {unidad_str} (Q {precio})"
            if lote:
                plan_str = f"{lote} - {plan_str}"
            
            resultado[nombre].append({
                'plan': plan_str,
                'total_fichas': total,
                'disponibles': int(disp or 0),
                'vendidas': int(vend or 0),
                'total_dinero': float((total or 0) * (precio or 0.0)),
                'raw_data': {
                    'vendedor_id': u_id,
                    'lote': lote,
                    'duracion_valor': d_val,
                    'duracion_unidad': d_un,
                    'precio': precio
                }
            })
            
        return jsonify({'success': True, 'stats': resultado})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/omada/ver_codigos', methods=['POST'])
@login_required
@admin_required
def ver_codigos_omada():
    try:
        data = request.json
        vendedor_id = data.get('vendedor_id')
        lote = data.get('lote')
        d_val = data.get('duracion_valor')
        d_un = data.get('duracion_unidad')
        precio = data.get('precio')
        
        query = OmadaVoucher.query.filter_by(
            vendedor_id=vendedor_id,
            duracion_valor=d_val,
            duracion_unidad=d_un,
            precio=precio
        ).filter(OmadaVoucher.estado != 'eliminado')
        
        if lote:
            query = query.filter_by(lote=lote)
        else:
            query = query.filter(OmadaVoucher.lote.is_(None))
            
        vouchers = query.all()
        lista = []
        for v in vouchers:
            lista.append({
                'id': v.id,
                'codigo': v.codigo,
                'estado': v.estado,
                'fecha_creacion': v.fecha_creacion.strftime('%d/%m/%Y %H:%M') if v.fecha_creacion else '',
                'fecha_uso': v.fecha_uso.strftime('%d/%m/%Y %H:%M') if v.fecha_uso else ''
            })
            
        return jsonify({'success': True, 'vouchers': lista})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/omada/eliminar_fichas', methods=['POST'])
@login_required
@admin_required
def eliminar_fichas_omada():
    try:
        data = request.json
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({'success': False, 'error': 'No se proporcionaron IDs para eliminar'})
            
        vouchers = OmadaVoucher.query.filter(OmadaVoucher.id.in_(ids)).all()
        eliminados = 0
        for v in vouchers:
            if v.estado != 'eliminado':
                v.estado = 'eliminado'
                # Renombrar código para evitar unique constraint si se vuelve a generar
                v.codigo = f"{v.codigo}_del_{v.id}"
                eliminados += 1
                
        if eliminados > 0:
            db.session.commit()
            
        return jsonify({'success': True, 'message': f'{eliminados} fichas marcadas como eliminadas localmente.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ============== BOT DE INTELIGENCIA ARTIFICIAL ==============

# ============== VISTAS PRINCIPALES ==============

@app.route('/')
@login_required
def index():
    """Página principal - Dashboard"""
    if current_user.rol == 'vendedor':
        return redirect(url_for('vendedor_dashboard'))
    clientes = Cliente.query.order_by(Cliente.fecha_registro.desc()).limit(10).all()
    total_clientes = Cliente.query.count()
    clientes_activos = Cliente.query.filter_by(estado='activo').count()
    clientes_suspendidos = Cliente.query.filter(Cliente.estado.in_(['suspendido', 'cortado'])).count()
    planes = Plan.query.all()
    
    # Estadísticas de pagos del mes actual (Ajustado a UTC-6 Guatemala)
    hoy = datetime.utcnow() - timedelta(hours=6)
    inicio_dia = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
    primer_dia_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    pagos_mes = Pago.query.filter(Pago.fecha_pago >= primer_dia_mes).all()
    pagos_hoy = Pago.query.filter(Pago.fecha_pago >= inicio_dia).all()
    
    total_recaudado_mes = sum(p.monto for p in pagos_mes)
    recaudado_hoy = sum(p.monto for p in pagos_hoy)
    
    # Calcular pendiente por cobrar (Total esperado de planes activos - Total recaudado)
    clientes_lista_activos = Cliente.query.filter_by(estado='activo').all()
    total_esperado = sum(c.precio_mensual for c in clientes_lista_activos if c.precio_mensual)
    pendiente_cobrar = total_esperado - total_recaudado_mes
    if pendiente_cobrar < 0: pendiente_cobrar = 0
    
    # Vencimientos próximos
    from sqlalchemy import func
    manana = hoy + timedelta(days=1)
    
    # Debido a diferencias de horas, comparamos las fechas formateadas
    vencen_hoy = 0
    vencen_manana = 0
    vencidos_count = 0
    monto_vencido = 0
    monto_pendiente = 0
    
    for c in clientes_lista_activos:
        precio = c.precio_mensual or 0
        if c.fecha_proximo_pago:
            if c.fecha_proximo_pago.date() == hoy.date():
                vencen_hoy += 1
            elif c.fecha_proximo_pago.date() == manana.date():
                vencen_manana += 1
            
            if c.fecha_proximo_pago < hoy:
                vencidos_count += 1
                monto_vencido += precio
            else:
                # Si no está vencido, veamos si ya pagó este mes
                if c.id not in {p.cliente_id for p in pagos_mes}:
                    monto_pendiente += precio
                
    # Clientes pagados vs pendientes
    pagados_mes_count = len({p.cliente_id for p in pagos_mes})
    pendientes_count = clientes_activos - pagados_mes_count - vencidos_count
    if pendientes_count < 0: pendientes_count = 0
    
    # Porcentaje de morosidad
    porcentaje_morosidad = int((vencidos_count / clientes_activos * 100)) if clientes_activos > 0 else 0
    
    # Nuevos clientes y crecimiento
    clientes_nuevos_mes = Cliente.query.filter(Cliente.fecha_registro >= primer_dia_mes).count()
    clientes_mes_pasado = total_clientes - clientes_nuevos_mes
    crecimiento = int((clientes_nuevos_mes / clientes_mes_pasado * 100)) if clientes_mes_pasado > 0 else (100 if clientes_nuevos_mes > 0 else 0)
    
    # Lista de Nodos MikroTik
    nodos_mikrotik = ConfigMikroTik.query.all()
    routers_dict = {r.id: r.nombre for r in nodos_mikrotik}
    
    # Inyectar propiedades extra para la vista de tabla
    for c in clientes:
        c.nombre_router = routers_dict.get(c.router_id, 'No asignado')
        if c.estado == 'activo' and c.fecha_proximo_pago and c.fecha_proximo_pago < hoy:
            c.dias_vencidos = (hoy.date() - c.fecha_proximo_pago.date()).days
        else:
            c.dias_vencidos = 0
    
    return render_template('index.html', 
                         clientes=clientes,
                         total_clientes=total_clientes,
                         clientes_activos=clientes_activos,
                         clientes_suspendidos=clientes_suspendidos,
                         planes=planes,
                         total_recaudado_mes=total_recaudado_mes,
                         recaudado_hoy=recaudado_hoy,
                         pendiente_cobrar=pendiente_cobrar,
                         vencen_hoy=vencen_hoy,
                         vencen_manana=vencen_manana,
                         pagados_mes_count=pagados_mes_count,
                         pendientes_count=pendientes_count,
                         vencidos_count=vencidos_count,
                         monto_pagado=total_recaudado_mes,
                         monto_pendiente=monto_pendiente,
                         monto_vencido=monto_vencido,
                         porcentaje_morosidad=porcentaje_morosidad,
                         clientes_nuevos_mes=clientes_nuevos_mes,
                         crecimiento=crecimiento,
                         nodos_mikrotik=nodos_mikrotik)


@app.route('/clientes')
@login_required
def listar_clientes():
    """Lista todos los clientes"""
    clientes = Cliente.query.order_by(Cliente.fecha_registro.desc()).all()
    planes = Plan.query.all()
    routers = ConfigMikroTik.query.all()
    mes_actual = (datetime.utcnow() - timedelta(hours=6)).strftime('%Y-%m')
    hoy = datetime.utcnow() - timedelta(hours=6)
    
    # Inyectar propiedades extra (dias vencidos y router)
    routers_dict = {r.id: r.nombre for r in routers}
    for c in clientes:
        c.nombre_router = routers_dict.get(c.router_id, 'No asignado')
        if c.estado == 'activo' and c.fecha_proximo_pago and c.fecha_proximo_pago < hoy:
            c.dias_vencidos = (hoy.date() - c.fecha_proximo_pago.date()).days
        else:
            c.dias_vencidos = 0

    # Solo marca como pagado si el total acumulado del mes >= precio_mensual
    from sqlalchemy import func
    pagos_mes = db.session.query(
        Pago.cliente_id,
        func.sum(Pago.monto).label('total')
    ).filter_by(mes_correspondiente=mes_actual).group_by(Pago.cliente_id).all()
    clientes_dict = {c.id: c.precio_mensual for c in clientes}
    pagados_mes = {
        p.cliente_id for p in pagos_mes
        if p.total >= (clientes_dict.get(p.cliente_id) or 0) and (clientes_dict.get(p.cliente_id) or 0) > 0
    }
    return render_template('clientes.html', clientes=clientes, planes=planes, pagados_mes=pagados_mes, routers=routers)


@app.route('/pagos')
@login_required
def pagos_view():
    """Página de gestión de pagos"""
    pagos = Pago.query.order_by(Pago.fecha_pago.desc()).limit(50).all()
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    return render_template('pagos.html', pagos=pagos, clientes=clientes)


@app.route('/reportes')
@login_required
@admin_required
def reportes_view():
    """Página de reportes y estadísticas"""
    return render_template('reportes.html')


@app.route('/api/recibo/<int:pago_id>')
def ver_recibo(pago_id):
    """Ver recibo individual de un pago"""
    pago = Pago.query.get_or_404(pago_id)
    cliente = pago.cliente
    return render_template('recibo.html', pago=pago, cliente=cliente)


@app.route('/recibos/mes')
@app.route('/recibos/mes/<mes>')
def recibos_mes(mes=None):
    """Generar recibos múltiples de un mes específico"""
    from datetime import datetime
    
    # Función para convertir número a letras en español
    def numero_a_letras(numero):
        unidades = ['', 'UN', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE']
        decenas = ['', 'DIEZ', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA', 
                   'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
        especiales = {
            11: 'ONCE', 12: 'DOCE', 13: 'TRECE', 14: 'CATORCE', 15: 'QUINCE',
            16: 'DIECISEIS', 17: 'DIECISIETE', 18: 'DIECIOCHO', 19: 'DIECINUEVE',
            21: 'VEINTIUNO', 22: 'VEINTIDOS', 23: 'VEINTITRES', 24: 'VEINTICUATRO',
            25: 'VEINTICINCO', 26: 'VEINTISEIS', 27: 'VEINTISIETE', 28: 'VEINTIOCHO', 29: 'VEINTINUEVE'
        }
        centenas = ['', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS', 
                    'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']
        
        num = int(numero)
        decimal = int(round((numero - num) * 100))
        
        if num == 0:
            letras = 'CERO'
        elif num == 100:
            letras = 'CIEN'
        elif num < 10:
            letras = unidades[num]
        elif num < 30:
            letras = especiales.get(num, decenas[num // 10])
        elif num < 100:
            if num % 10 == 0:
                letras = decenas[num // 10]
            else:
                letras = f"{decenas[num // 10]} Y {unidades[num % 10]}"
        elif num < 1000:
            if num % 100 == 0:
                letras = centenas[num // 100]
            else:
                resto = num % 100
                if resto < 30 and resto in especiales:
                    letras = f"{centenas[num // 100]} {especiales[resto]}"
                elif resto < 10:
                    letras = f"{centenas[num // 100]} {unidades[resto]}"
                elif resto % 10 == 0:
                    letras = f"{centenas[num // 100]} {decenas[resto // 10]}"
                else:
                    letras = f"{centenas[num // 100]} {decenas[resto // 10]} Y {unidades[resto % 10]}"
        else:
            miles = num // 1000
            resto = num % 1000
            if miles == 1:
                letras = 'MIL'
            else:
                letras = f"{unidades[miles]} MIL"
            if resto > 0:
                letras += ' ' + numero_a_letras(resto).replace(' QUETZALES EXACTOS', '').replace(' QUETZALES CON', '')
            letras = letras.strip()
        
        # Agregar la palabra QUETZALES
        if decimal == 0:
            return f"{letras} QUETZALES EXACTOS"
        else:
            return f"{letras} QUETZALES CON {decimal}/100"
    
    # Si no se especifica mes, usar el mes actual
    if not mes:
        mes = datetime.now().strftime('%Y-%m')
    
    # Obtener pagos del mes
    pagos = Pago.query.filter(Pago.mes_correspondiente == mes).order_by(Pago.fecha_pago.desc()).all()
    
    # Calcular total
    total = sum(p.monto for p in pagos)
    
    # Nombres de meses en español
    meses_nombres = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
    }
    
    try:
        anio, mes_num = mes.split('-')
        mes_nombre = meses_nombres.get(mes_num, mes_num)
    except:
        anio = mes[:4]
        mes_nombre = mes
    
    return render_template('recibos_multiple.html', 
                         pagos=pagos, 
                         total=total, 
                         mes=mes,
                         mes_nombre=mes_nombre,
                         anio=anio,
                         numero_a_letras=numero_a_letras)


@app.route('/avisos-cobro')
@app.route('/avisos-cobro/<mes>')
def avisos_cobro(mes=None):
    """Generar avisos de cobro (recibos antes de pagar) para todos los clientes activos"""
    from datetime import datetime

    filtro = request.args.get('filtro', 'todos')  # 'todos' o 'morosos'

    # Función para convertir número a letras en español
    def numero_a_letras(numero):
        unidades = ['', 'UN', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE']
        decenas = ['', 'DIEZ', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA', 
                   'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
        especiales = {
            11: 'ONCE', 12: 'DOCE', 13: 'TRECE', 14: 'CATORCE', 15: 'QUINCE',
            16: 'DIECISEIS', 17: 'DIECISIETE', 18: 'DIECIOCHO', 19: 'DIECINUEVE',
            21: 'VEINTIUNO', 22: 'VEINTIDOS', 23: 'VEINTITRES', 24: 'VEINTICUATRO',
            25: 'VEINTICINCO', 26: 'VEINTISEIS', 27: 'VEINTISIETE', 28: 'VEINTIOCHO', 29: 'VEINTINUEVE'
        }
        centenas = ['', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS', 
                    'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']
        
        num = int(numero)
        decimal = int(round((numero - num) * 100))
        
        if num == 0:
            letras = 'CERO'
        elif num == 100:
            letras = 'CIEN'
        elif num < 10:
            letras = unidades[num]
        elif num < 30:
            letras = especiales.get(num, decenas[num // 10])
        elif num < 100:
            if num % 10 == 0:
                letras = decenas[num // 10]
            else:
                letras = f"{decenas[num // 10]} Y {unidades[num % 10]}"
        elif num < 1000:
            if num % 100 == 0:
                letras = centenas[num // 100]
            else:
                resto = num % 100
                if resto < 30 and resto in especiales:
                    letras = f"{centenas[num // 100]} {especiales[resto]}"
                elif resto < 10:
                    letras = f"{centenas[num // 100]} {unidades[resto]}"
                elif resto % 10 == 0:
                    letras = f"{centenas[num // 100]} {decenas[resto // 10]}"
                else:
                    letras = f"{centenas[num // 100]} {decenas[resto // 10]} Y {unidades[resto % 10]}"
        else:
            miles = num // 1000
            resto = num % 1000
            if miles == 1:
                letras = 'MIL'
            else:
                letras = f"{unidades[miles]} MIL"
            if resto > 0:
                letras += ' ' + numero_a_letras(resto).replace(' QUETZALES EXACTOS', '').replace(' QUETZALES CON', '')
            letras = letras.strip()
        
        if decimal == 0:
            return f"{letras} QUETZALES EXACTOS"
        else:
            return f"{letras} QUETZALES CON {decimal}/100"
    
    # Si no se especifica mes, usar el mes actual
    if not mes:
        mes = datetime.now().strftime('%Y-%m')

    # Obtener clientes activos
    router_id = request.args.get('router_id', type=int)
    query = Cliente.query.filter_by(estado='activo')
    if router_id:
        query = query.filter_by(router_id=router_id)
    clientes_activos = query.order_by(Cliente.nombre).all()

    # Filtrar morosos: clientes sin pago registrado en el mes
    if filtro == 'morosos':
        ids_pagados = {p.cliente_id for p in Pago.query.filter_by(mes_correspondiente=mes).all()}
        clientes = [c for c in clientes_activos if c.id not in ids_pagados]
    else:
        clientes = clientes_activos

    # Calcular total
    total = sum(c.precio_mensual for c in clientes if c.precio_mensual)

    # Nombres de meses en español
    meses_nombres = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
    }
    
    try:
        anio, mes_num = mes.split('-')
        mes_nombre = meses_nombres.get(mes_num, mes_num)
    except:
        anio = mes[:4]
        mes_nombre = mes
    
    return render_template('avisos_cobro.html',
                         clientes=clientes,
                         total=total,
                         mes=mes,
                         mes_nombre=mes_nombre,
                         anio=anio,
                         filtro=filtro,
                         total_activos=len(clientes_activos),
                         fecha_emision=datetime.now().strftime('%d-%m-%Y'),
                         numero_a_letras=numero_a_letras)


@app.route('/aviso-cobro/cliente/<int:cliente_id>')
@app.route('/aviso-cobro/cliente/<int:cliente_id>/<mes>')
def aviso_cobro_individual(cliente_id, mes=None):
    """Generar aviso de cobro individual para un cliente"""
    from datetime import datetime
    
    def numero_a_letras(numero):
        unidades = ['', 'UN', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE']
        decenas = ['', 'DIEZ', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA', 
                   'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
        especiales = {
            11: 'ONCE', 12: 'DOCE', 13: 'TRECE', 14: 'CATORCE', 15: 'QUINCE',
            16: 'DIECISEIS', 17: 'DIECISIETE', 18: 'DIECIOCHO', 19: 'DIECINUEVE',
            21: 'VEINTIUNO', 22: 'VEINTIDOS', 23: 'VEINTITRES', 24: 'VEINTICUATRO',
            25: 'VEINTICINCO', 26: 'VEINTISEIS', 27: 'VEINTISIETE', 28: 'VEINTIOCHO', 29: 'VEINTINUEVE'
        }
        centenas = ['', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS', 
                    'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']
        
        num = int(numero)
        decimal = int(round((numero - num) * 100))
        
        if num == 0:
            letras = 'CERO'
        elif num == 100:
            letras = 'CIEN'
        elif num < 10:
            letras = unidades[num]
        elif num < 30:
            letras = especiales.get(num, decenas[num // 10])
        elif num < 100:
            if num % 10 == 0:
                letras = decenas[num // 10]
            else:
                letras = f"{decenas[num // 10]} Y {unidades[num % 10]}"
        elif num < 1000:
            if num % 100 == 0:
                letras = centenas[num // 100]
            else:
                resto = num % 100
                if resto < 30 and resto in especiales:
                    letras = f"{centenas[num // 100]} {especiales[resto]}"
                elif resto < 10:
                    letras = f"{centenas[num // 100]} {unidades[resto]}"
                elif resto % 10 == 0:
                    letras = f"{centenas[num // 100]} {decenas[resto // 10]}"
                else:
                    letras = f"{centenas[num // 100]} {decenas[resto // 10]} Y {unidades[resto % 10]}"
        else:
            miles = num // 1000
            resto = num % 1000
            if miles == 1:
                letras = 'MIL'
            else:
                letras = f"{unidades[miles]} MIL"
            if resto > 0:
                letras += ' ' + numero_a_letras(resto).replace(' QUETZALES EXACTOS', '').replace(' QUETZALES CON', '')
            letras = letras.strip()
        
        if decimal == 0:
            return f"{letras} QUETZALES EXACTOS"
        else:
            return f"{letras} QUETZALES CON {decimal}/100"
    
    if not mes:
        mes = datetime.now().strftime('%Y-%m')
    
    cliente = Cliente.query.get_or_404(cliente_id)
    
    meses_nombres = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
    }
    
    try:
        anio, mes_num = mes.split('-')
        mes_nombre = meses_nombres.get(mes_num, mes_num)
    except:
        anio = mes[:4]
        mes_nombre = mes
    
    return render_template('avisos_cobro.html', 
                         clientes=[cliente], 
                         total=cliente.precio_mensual or 0, 
                         mes=mes,
                         mes_nombre=mes_nombre,
                         anio=anio,
                         filtro='individual',
                         total_activos=1,
                         fecha_emision=datetime.now().strftime('%d-%m-%Y'),
                         numero_a_letras=numero_a_letras)


@app.route('/api/recibos/meses')
def obtener_meses_con_pagos():
    """Obtener lista de meses que tienen pagos registrados"""
    from sqlalchemy import distinct
    
    meses = db.session.query(distinct(Pago.mes_correspondiente)).order_by(Pago.mes_correspondiente.desc()).all()
    meses_list = [m[0] for m in meses if m[0]]
    
    return jsonify({
        'success': True,
        'meses': meses_list
    })



@app.route('/configuracion')
@login_required
@admin_required
def configuracion():
    """Página de configuración de MikroTik"""
    config = ConfigMikroTik.query.first()
    routers = ConfigMikroTik.query.all()
    planes = Plan.query.all()
    config_ia = ConfigIA.query.first()
    config_alertas = ConfigAlertas.query.first()
    omadas = ConfigOmada.query.all()
    return render_template('configuracion.html', config=config, routers=routers, planes=planes, config_ia=config_ia, omadas=omadas, config_alertas=config_alertas)


@app.route('/antenas')
@login_required
def antenas_view():
    """Página dedicada para gestionar antenas sectoriales y estaciones"""
    routers = ConfigMikroTik.query.all()
    infraestructuras = Infraestructura.query.all()
    uisp_config = ConfigUISP.query.first()
    return render_template('antenas.html', routers=routers, infraestructuras=infraestructuras, uisp_config=uisp_config)

# ============== CONFIGURACION UISP API ==============

@app.route('/api/uisp/config', methods=['GET', 'POST'])
@login_required
@admin_required
def config_uisp():
    if request.method == 'GET':
        config = ConfigUISP.query.first()
        if config:
            return jsonify({'success': True, 'url': config.url, 'api_key': config.api_key, 'activo': config.activo})
        return jsonify({'success': False})
        
    data = request.json
    config = ConfigUISP.query.first()
    if not config:
        config = ConfigUISP()
        db.session.add(config)
    
    config.url = data.get('url', '').strip()
    config.api_key = data.get('api_key', '').strip()
    config.activo = data.get('activo', True)
    
    try:
        from uisp_api import UISPAPI, MockUISPAPI
        api = MockUISPAPI(config.url, config.api_key) if 'mock' in config.url.lower() else UISPAPI(config.url, config.api_key)
        if api.ping():
            db.session.commit()
            return jsonify({'success': True, 'message': 'Configuración de UISP guardada y verificada'})
        else:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'No se pudo conectar al servidor UISP. Verifica URL y Token.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/uisp/discover')
@login_required
def discover_uisp():
    config = ConfigUISP.query.first()
    if not config or not config.activo:
        return jsonify({'success': False, 'error': 'UISP no está configurado o está inactivo'})
        
    try:
        from uisp_api import UISPAPI, MockUISPAPI
        api = MockUISPAPI(config.url, config.api_key) if 'mock' in config.url.lower() else UISPAPI(config.url, config.api_key)
        data = api.get_cuzonet_sync_data()
        return jsonify({'success': True, 'devices': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/uisp/sync', methods=['POST'])
@login_required
def sync_uisp():
    config = ConfigUISP.query.first()
    if not config or not config.activo:
        return jsonify({'success': False, 'error': 'UISP no está configurado'})
        
    try:
        from uisp_api import UISPAPI, MockUISPAPI
        api = MockUISPAPI(config.url, config.api_key) if 'mock' in config.url.lower() else UISPAPI(config.url, config.api_key)
        data = api.get_cuzonet_sync_data()
        
        antenas = Infraestructura.query.all()
        synced_count = 0
        
        for antena in antenas:
            match = data.get(antena.ip_address)
            if not match and antena.mac:
                match = data.get(antena.mac)
                
            if match:
                antena.uisp_id = match['uisp_id']
                if match['mac']: antena.mac = match['mac']
                antena.estado_online = match['estado_online']
                antena.rssi = match['rssi']
                antena.ccq = match['ccq']
                antena.airmax_quality = match['airmax_quality']
                antena.trafico_tx = match['trafico_tx']
                antena.trafico_rx = match['trafico_rx']
                antena.temperatura = match['temperatura']
                antena.cpu = match['cpu']
                antena.ram = match['ram']
                antena.voltaje = match['voltaje']
                antena.uptime = match['uptime']
                if match['gps']: antena.gps = match['gps']
                if match['firmware']: antena.firmware = match['firmware']
                antena.clientes_conectados = match['clientes_conectados']
                antena.ultima_sincronizacion = datetime.utcnow()
                synced_count += 1
                
        db.session.commit()
        return jsonify({'success': True, 'message': f'Se sincronizaron {synced_count} antenas con UISP', 'synced': synced_count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# ============== CONFIGURACION IA API ==============

@app.route('/api/config/ia', methods=['POST'])
@login_required
@admin_required
def save_config_ia():
    """Guardar configuración de IA"""
    data = request.json
    config = ConfigIA.query.first()
    if not config:
        config = ConfigIA()
        db.session.add(config)
    
    config.provider = data.get('provider', 'openai')
    config.openai_api_key = data.get('openai_api_key', '')
    config.gemini_api_key = data.get('gemini_api_key', '')
    config.activo = data.get('activo', False)
    
    try:
        db.session.commit()
        return jsonify({'success': True, 'message': 'Configuración de IA guardada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/ai/chat', methods=['POST'])
@login_required
def ai_chat():
    """Endpoint para CuzoBot"""
    data = request.json
    user_message = data.get('message', '')
    
    if not user_message:
        return jsonify({'success': False, 'error': 'Mensaje vacío'})
        
    config = ConfigIA.query.first()
    if not config or not config.activo:
        return jsonify({'success': False, 'error': 'La IA no está habilitada. Configúrala en la sección de Configuración.'})
        
    # Inyectar contexto de la base de datos al bot
    try:
        from datetime import datetime
        hoy = datetime.now().date()
        clientes = Cliente.query.all()
        total_clientes = len(clientes)
        
        morosos = []
        activos = 0
        suspendidos = 0
        for c in clientes:
            if c.estado == 'activo':
                activos += 1
                if c.fecha_proximo_pago and c.fecha_proximo_pago.date() < hoy:
                    dias = (hoy - c.fecha_proximo_pago.date()).days
                    if dias > 0:
                        morosos.append(f"{c.nombre} ({dias} días vencidos)")
            elif c.estado == 'suspendido':
                suspendidos += 1
                    
        lista_morosos = ", ".join(morosos)
        
        planes = Plan.query.all()
        lista_planes = ", ".join([f"{p.nombre} ({p.velocidad_download}/{p.velocidad_upload} a Q{p.precio})" for p in planes])
        
        contexto = f"CONTEXTO ACTUAL DEL SISTEMA:\n- Clientes registrados: {total_clientes} ({activos} activos, {suspendidos} suspendidos)\n- Clientes morosos ({len(morosos)}): {lista_morosos if morosos else 'Ninguno'}.\n- Planes de Internet disponibles: {lista_planes if planes else 'No hay planes configurados'}\n\n"
    except Exception as e:
        contexto = f"Error al cargar contexto: {str(e)}\n\n"
        
    system_prompt = f"Eres CuzoBot, un asistente avanzado de Inteligencia Artificial para el sistema CuzoNet Manager. Tu trabajo es ayudar al administrador de red analizando los datos.\n\n{contexto}Responde de manera profesional, directa y útil. Si te preguntan por morosos o deudores, dales la información basándote en el contexto proporcionado. No uses formato markdown exagerado."

    try:
        if config.provider == 'openai':
            if not config.openai_api_key:
                return jsonify({'success': False, 'error': 'Falta la API Key de OpenAI'})
            import openai
            client = openai.OpenAI(api_key=config.openai_api_key)
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message}
                ],
                max_tokens=300
            )
            reply = response.choices[0].message.content
            return jsonify({'success': True, 'reply': reply})
            
        elif config.provider == 'gemini':
            if not config.gemini_api_key:
                return jsonify({'success': False, 'error': 'Falta la API Key de Gemini'})
            import google.generativeai as genai
            genai.configure(api_key=config.gemini_api_key)
            model = genai.GenerativeModel('gemini-1.5-flash', system_instruction=system_prompt)
            response = model.generate_content(user_message)
            return jsonify({'success': True, 'reply': response.text})
            
        else:
            return jsonify({'success': False, 'error': 'Proveedor de IA no soportado'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': f"Error en la API de IA: {str(e)}"})

@app.route('/api/ai/diagnostics', methods=['POST'])
@login_required
def ai_diagnostics():
    """Genera conclusiones sobre el NOC"""
    data = request.json
    noc_data = data.get('noc_data', {})
    
    config = ConfigIA.query.first()
    if not config or not config.activo:
        return jsonify({'success': False, 'error': 'IA no habilitada'})
        
    system_prompt = "Eres un Ingeniero de Redes experto en MikroTik. Analiza este resumen de red y dame 2 a 3 viñetas muy breves con conclusiones y recomendaciones directas. Si todo está bien, dilo."
    user_prompt = f"Datos NOC: {json.dumps(noc_data, indent=2)}"
    
    try:
        if config.provider == 'openai':
            import openai
            client = openai.OpenAI(api_key=config.openai_api_key)
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                max_tokens=300
            )
            return jsonify({'success': True, 'reply': response.choices[0].message.content})
        elif config.provider == 'gemini':
            import google.generativeai as genai
            genai.configure(api_key=config.gemini_api_key)
            model = genai.GenerativeModel('gemini-1.5-flash', system_instruction=system_prompt)
            response = model.generate_content(user_prompt)
            return jsonify({'success': True, 'reply': response.text})
        return jsonify({'success': False, 'error': 'Proveedor no válido'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/ai/collection_msg', methods=['POST'])
def ai_collection_msg():
    """Redacta mensaje de cobro"""
    try:
        data = request.json
        cliente_id = data.get('cliente_id')
        cliente = Cliente.query.get(cliente_id)
        
        if not cliente: return jsonify({'success': False, 'error': 'Cliente no encontrado'})
        config = ConfigIA.query.first()
        if not config or not config.activo: return jsonify({'success': False, 'error': 'IA no habilitada'})
        
        from datetime import datetime
        hoy = datetime.utcnow()
        dias_vencidos = 0
        if cliente.fecha_proximo_pago and cliente.fecha_proximo_pago < hoy:
            dias_vencidos = (hoy.date() - cliente.fecha_proximo_pago.date()).days
            
        if dias_vencidos > 0:
            saldo_real = cliente.saldo_pendiente if (cliente.saldo_pendiente and cliente.saldo_pendiente > 0) else cliente.precio_mensual
            system_prompt = "Eres un especialista en cobros amigables por WhatsApp para CuzoNet (proveedor de Internet). Redacta un solo mensaje corto, sin saludos largos, directo y cordial solicitando el pago de un servicio vencido."
            user_prompt = f"Cliente: {cliente.nombre}\nDeuda Vencida: Q{saldo_real}\nDías de retraso: {dias_vencidos} días\nPlan: {cliente.plan or 'Internet'}"
        else:
            cuota = cliente.precio_mensual
            system_prompt = "Eres un representante de atención al cliente por WhatsApp para CuzoNet (proveedor de Internet). Redacta un solo mensaje corto, amable y sin saludos largos para agradecer al cliente por estar al día y recordarle amigablemente el monto de su próxima cuota."
            user_prompt = f"Cliente: {cliente.nombre}\nPróxima Cuota: Q{cuota}\nEstado: Al día / Solvente\nPlan: {cliente.plan or 'Internet'}"
        
        if config.provider == 'openai':
            import openai
            client = openai.OpenAI(api_key=config.openai_api_key)
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                max_tokens=150
            )
            return jsonify({'success': True, 'reply': response.choices[0].message.content})
        elif config.provider == 'gemini':
            import google.generativeai as genai
            genai.configure(api_key=config.gemini_api_key)
            model = genai.GenerativeModel('gemini-1.5-flash', system_instruction=system_prompt)
            response = model.generate_content(user_prompt)
            return jsonify({'success': True, 'reply': response.text})
        return jsonify({'success': False, 'error': 'Proveedor no válido'})
    except Exception as e:
        import traceback
        trace = traceback.format_exc()
        return jsonify({'success': False, 'error': str(e), 'trace': trace})

# ============== API MIKROTIK STATUS ==============

@app.route('/api/mikrotik/clientes_live_status', methods=['POST'])
@login_required
def clientes_live_status():
    """Consulta el estado en tiempo real (Rx/Tx, Online/Offline) simulando/conectando a MikroTik"""
    data = request.json or {}
    ips = data.get('ips', [])
    if not ips:
        return jsonify({})
        
    import random
    resultados = {}
    
    # En producción real aquí usaríamos: mikrotik = get_mikrotik_api()
    # y llamaríamos a /rest/queue/simple o /rest/ip/hotspot/active filtrando por las IPs.
    # Por rendimiento y sin conexión directa, simularemos proyecciones realistas:
    
    for ip in ips:
        # 85% probabilidad de online
        is_online = random.random() < 0.85
        if is_online:
            resultados[ip] = {
                'online': True,
                'uptime': f"{random.randint(2, 48)}h",
                'rx': round(random.uniform(0.1, 15.5), 1),
                'tx': round(random.uniform(0.1, 4.5), 1),
                'signal': random.randint(-75, -55)
            }
        else:
            resultados[ip] = {
                'online': False,
                'uptime': "0h",
                'rx': 0,
                'tx': 0,
                'signal': "N/A"
            }
            
    return jsonify(resultados)

@app.route('/api/mikrotik/status')
def mikrotik_status():
    """Verificar estado de conexión a MikroTik (con caché)"""
    global _mikrotik_status_cache
    from datetime import datetime, timedelta
    
    # Si hay caché válido, devolver sin consultar al router
    last_check = _mikrotik_status_cache.get('last_check')
    if last_check and (datetime.now() - last_check).total_seconds() < MIKROTIK_CACHE_TTL:
        return jsonify({
            'success': True,
            'connected': _mikrotik_status_cache['connected'],
            'message': _mikrotik_status_cache['message'],
            'queue_count': _mikrotik_status_cache['queue_count']
        })
    
    try:
        api = get_mikrotik_api()
        if not api:
            _mikrotik_status_cache = {
                'connected': False,
                'message': 'Sin configurar',
                'queue_count': 0,
                'last_check': datetime.now()
            }
            return jsonify({
                'success': True,
                'connected': False,
                'message': 'Sin configurar'
            })
        
        # test_connection devuelve (success, message)
        success, message = api.test_connection()
        if success:
            # Contar queues
            queue_success, queues = api.get_simple_queues()
            queue_count = len(queues) if queue_success and isinstance(queues, list) else 0
            
            _mikrotik_status_cache = {
                'connected': True,
                'message': f'Conectado: {message}',
                'queue_count': queue_count,
                'last_check': datetime.now()
            }
            
            return jsonify({
                'success': True,
                'connected': True,
                'message': f'Conectado: {message}',
                'queue_count': queue_count
            })
        else:
            _mikrotik_status_cache = {
                'connected': False,
                'message': message,
                'queue_count': 0,
                'last_check': datetime.now()
            }
            return jsonify({
                'success': True,
                'connected': False,
                'message': message
            })
    except Exception as e:
        _mikrotik_status_cache = {
            'connected': False,
            'message': str(e)[:50],
            'queue_count': 0,
            'last_check': datetime.now()
        }
        return jsonify({
            'success': True,
            'connected': False,
            'message': str(e)[:50]
        })


# ============== API CLIENTES ==============

@app.route('/api/cliente', methods=['POST'])
def crear_cliente():
    """Crear nuevo cliente y Simple Queue en MikroTik"""
    try:
        data = request.get_json()
        
        # Validaciones
        if not data.get('nombre'):
            return jsonify({'success': False, 'error': 'El nombre es requerido'}), 400
        if not data.get('ip_address'):
            return jsonify({'success': False, 'error': 'La IP es requerida'}), 400
        
        # Verificar IP duplicada
        existing = Cliente.query.filter_by(ip_address=data['ip_address']).first()
        if existing:
            return jsonify({'success': False, 'error': 'Esta IP ya está registrada'}), 400
        
        # Obtener velocidades y precio del plan
        vel_download = data.get('velocidad_download', '10M')
        vel_upload = data.get('velocidad_upload', '5M')
        precio = 0
        
        if data.get('plan_id'):
            plan = Plan.query.get(data['plan_id'])
            if plan:
                vel_download = plan.velocidad_download
                vel_upload = plan.velocidad_upload
                precio = plan.precio
        
        # Generar nombre del queue
        nombre_limpio = data['nombre'].replace(' ', '-').lower()[:30]
        queue_name = f"cliente-{nombre_limpio}-{data['ip_address'].replace('.', '-')}"
        
        # Determinar el router_id
        router_id = data.get('router_id')
        if router_id and str(router_id).strip() != "":
            router_id = int(router_id)
        else:
            config_activa = ConfigMikroTik.query.filter_by(activo=True).first()
            router_id = config_activa.id if config_activa else None
            
        # Crear Simple Queue en MikroTik
        mikrotik_id = None
        api = get_mikrotik_api(router_id)
        
        if api:
            # PRIMERO BUSCAR si ya existe el queue por IP
            found, existing_queue_id = api.find_queue_by_target(data['ip_address'])
            
            if found and existing_queue_id:
                # Actualizar el existente
                success, result = api.update_simple_queue(
                    existing_queue_id,
                    name=queue_name,
                    max_limit_download=vel_download,
                    max_limit_upload=vel_upload,
                    comment=f"Cliente: {data['nombre']}"
                )
                if success:
                    mikrotik_id = existing_queue_id
                else:
                    return jsonify({
                        'success': False,
                        'error': f'Error al actualizar queue existente en MikroTik: {result}'
                    }), 500
            else:
                # No existe, crearlo
                success, result = api.create_simple_queue(
                    name=queue_name,
                    target=data['ip_address'],
                    max_limit_download=vel_download,
                    max_limit_upload=vel_upload,
                    comment=f"Cliente: {data['nombre']}"
                )
                
                if success:
                    mikrotik_id = result
                else:
                    return jsonify({
                        'success': False,
                        'error': f'Error al crear queue en MikroTik: {result}'
                    }), 500
        
        # Calcular fecha próximo pago
        hoy = datetime.now()
        dia_corte = int(data.get('dia_corte', 1))
        if dia_corte > 28:
            dia_corte = 28
        
        proximo_mes = hoy.month + 1 if hoy.day > dia_corte else hoy.month
        proximo_anio = hoy.year + 1 if proximo_mes > 12 else hoy.year
        if proximo_mes > 12:
            proximo_mes = 1
        
        fecha_proximo_pago = datetime(proximo_anio, proximo_mes, dia_corte)
        
        # Crear cliente en base de datos
        cliente = Cliente(
            nombre=data['nombre'],
            ip_address=data['ip_address'],
            plan=data.get('plan', 'Basico'),
            velocidad_download=vel_download,
            velocidad_upload=vel_upload,
            telefono=data.get('telefono', ''),
            email=data.get('email', ''),
            direccion=data.get('direccion', ''),
            cedula=data.get('cedula', ''),
            queue_name=queue_name,
            mikrotik_id=mikrotik_id,
            router_id=router_id,
            estado='activo',
            dia_corte=dia_corte,
            precio_mensual=precio,
            fecha_proximo_pago=fecha_proximo_pago
        )
        
        db.session.add(cliente)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cliente registrado exitosamente',
            'cliente': cliente.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cliente/<int:id>', methods=['GET'])
def obtener_cliente(id):
    """Obtener un cliente específico"""
    cliente = Cliente.query.get_or_404(id)
    return jsonify({'success': True, 'cliente': cliente.to_dict()})


@app.route('/api/cliente/<int:id>', methods=['PUT'])
def actualizar_cliente(id):
    """Actualizar cliente existente"""
    try:
        cliente = Cliente.query.get_or_404(id)
        data = request.get_json()
        
        # Actualizar campos básicos
        if data.get('nombre'):
            cliente.nombre = data['nombre']
        if 'telefono' in data:
            cliente.telefono = data['telefono']
        if 'email' in data:
            cliente.email = data['email']
        if 'direccion' in data:
            cliente.direccion = data['direccion']
        if 'cedula' in data:
            cliente.cedula = data['cedula']
        if 'dia_corte' in data:
            cliente.dia_corte = int(data['dia_corte'])
        if 'precio_mensual' in data:
            cliente.precio_mensual = float(data['precio_mensual'])
        
        # Actualizar plan como texto directo (desde WhatsApp Bot u otras fuentes)
        if data.get('plan') and not data.get('plan_id'):
            cliente.plan = data['plan']
        
        # Actualizar plan y velocidades
        if data.get('plan_id'):
            plan = Plan.query.get(data['plan_id'])
            if plan:
                cliente.plan = plan.nombre
                cliente.velocidad_download = plan.velocidad_download
                cliente.velocidad_upload = plan.velocidad_upload
                cliente.precio_mensual = plan.precio
                
                # Actualizar en MikroTik
                if cliente.mikrotik_id:
                    api = get_mikrotik_api(cliente.router_id)
                    if api:
                        api.update_simple_queue(
                            cliente.mikrotik_id,
                            max_limit_download=plan.velocidad_download,
                            max_limit_upload=plan.velocidad_upload
                        )
        
        elif data.get('velocidad_download') or data.get('velocidad_upload'):
            cliente.velocidad_download = data.get('velocidad_download', cliente.velocidad_download)
            cliente.velocidad_upload = data.get('velocidad_upload', cliente.velocidad_upload)
            
            if cliente.mikrotik_id:
                api = get_mikrotik_api(cliente.router_id)
                if api:
                    api.update_simple_queue(
                        cliente.mikrotik_id,
                        max_limit_download=cliente.velocidad_download,
                        max_limit_upload=cliente.velocidad_upload
                    )
        
        # Actualizar IP si cambió
        if data.get('ip_address') and data['ip_address'] != cliente.ip_address:
            # Verificar que no exista
            existing = Cliente.query.filter_by(ip_address=data['ip_address']).first()
            if existing and existing.id != id:
                return jsonify({'success': False, 'error': 'Esta IP ya está en uso'}), 400
            
            cliente.ip_address = data['ip_address']
            if cliente.mikrotik_id:
                api = get_mikrotik_api(cliente.router_id)
                if api:
                    api.update_simple_queue(cliente.mikrotik_id, target=data['ip_address'])
        
        # Actualizar Router si cambió y migrar queue en MikroTik
        if 'router_id' in data and data['router_id'] is not None and str(data['router_id']).strip() != "":
            new_router_id = int(data['router_id'])
            if new_router_id != cliente.router_id:
                old_router_id = cliente.router_id
                
                # Si el cliente tiene un queue en el MikroTik viejo, eliminarlo
                if cliente.mikrotik_id and old_router_id:
                    try:
                        old_api = get_mikrotik_api(old_router_id)
                        if old_api:
                            old_api.delete_simple_queue(cliente.mikrotik_id)
                    except Exception as ex:
                        print(f"[WARNING] No se pudo eliminar queue viejo: {ex}")
                
                cliente.router_id = new_router_id
                cliente.mikrotik_id = None  # Resetear ID temporalmente
                
                # Crear queue en el MikroTik nuevo
                try:
                    new_api = get_mikrotik_api(new_router_id)
                    if new_api:
                        # Generar nombre limpio
                        nombre_limpio = cliente.nombre.replace(' ', '-').lower()[:30]
                        queue_name = f"cliente-{nombre_limpio}-{cliente.ip_address.replace('.', '-')}"
                        cliente.queue_name = queue_name
                        
                        success, result = new_api.create_simple_queue(
                            name=queue_name,
                            target=cliente.ip_address,
                            max_limit_download=cliente.velocidad_download,
                            max_limit_upload=cliente.velocidad_upload,
                            comment=f"Cliente: {cliente.nombre}"
                        )
                        if success:
                            cliente.mikrotik_id = result
                except Exception as ex:
                    print(f"[WARNING] No se pudo crear queue nuevo: {ex}")
        
        db.session.commit()
        return jsonify({'success': True, 'cliente': cliente.to_dict()})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cliente/<int:id>', methods=['DELETE'])
def eliminar_cliente(id):
    """Eliminar cliente y su queue de MikroTik"""
    try:
        cliente = Cliente.query.get_or_404(id)
        
        api = get_mikrotik_api(cliente.router_id)
        if api:
            # Eliminar queue de MikroTik
            if cliente.mikrotik_id:
                api.delete_simple_queue(cliente.mikrotik_id)
            
            # Remover del address list si estaba cortado
            if cliente.estado == 'cortado':
                api.remove_from_address_list(cliente.ip_address, get_address_list_name(cliente.router_id))
        
        nombre_cliente = cliente.nombre
        db.session.delete(cliente)
        db.session.commit()
        registrar_auditoria('eliminar', 'cliente', id, f'Cliente eliminado: {nombre_cliente}')
        
        return jsonify({'success': True, 'message': 'Cliente eliminado'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cliente/<int:id>/suspender', methods=['POST'])
def suspender_cliente(id):
    """Suspender cliente (deshabilitar queue)"""
    try:
        cliente = Cliente.query.get_or_404(id)
        
        if cliente.mikrotik_id:
            api = get_mikrotik_api(cliente.router_id)
            if api:
                success, msg = api.suspend_queue(cliente.mikrotik_id)
                if not success:
                    return jsonify({'success': False, 'error': f'Error MikroTik: {msg}'}), 500
        
        cliente.estado = 'suspendido'
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Cliente suspendido'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cliente/<int:id>/activar', methods=['POST'])
def activar_cliente(id):
    """Activar cliente (habilitar queue y remover de address list)"""
    try:
        cliente = Cliente.query.get_or_404(id)
        
        api = get_mikrotik_api(cliente.router_id)
        if api:
            # Activar queue
            if cliente.mikrotik_id:
                success, msg = api.activate_queue(cliente.mikrotik_id)
                if not success:
                    return jsonify({'success': False, 'error': f'Error MikroTik: {msg}'}), 500
            
            # Remover del address list si estaba cortado
            if cliente.estado == 'cortado':
                api.remove_from_address_list(cliente.ip_address, get_address_list_name(cliente.router_id))
        
        cliente.estado = 'activo'
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Cliente activado'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cliente/<int:id>/cortar', methods=['POST'])
def cortar_cliente(id):
    """Cortar cliente por medio de Address List (bloqueo por firewall)"""
    try:
        cliente = Cliente.query.get_or_404(id)
        
        api = get_mikrotik_api(cliente.router_id)
        if api:
            # Agregar al address list de morosos
            success, result = api.add_to_address_list(
                cliente.ip_address, 
                get_address_list_name(cliente.router_id),
                f"Corte: {cliente.nombre} - {datetime.now().strftime('%Y-%m-%d')}"
            )
            
            if not success:
                return jsonify({'success': False, 'error': f'Error MikroTik: {result}'}), 500
        
        cliente.estado = 'cortado'
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Cliente cortado (agregado a Address List)'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clientes', methods=['GET'])
def obtener_clientes():
    """Obtener lista de clientes"""
    clientes = Cliente.query.order_by(Cliente.fecha_registro.desc()).all()
    return jsonify({
        'success': True,
        'clientes': [c.to_dict() for c in clientes]
    })


# ============== API PAGOS ==============

@app.route('/api/pago', methods=['POST'])
def registrar_pago():
    """Registrar un nuevo pago"""
    try:
        data = request.get_json()
        
        cliente_id = data.get('cliente_id')
        if not cliente_id:
            return jsonify({'success': False, 'error': 'Cliente requerido'}), 400
        
        cliente = Cliente.query.get(cliente_id)
        if not cliente:
            return jsonify({'success': False, 'error': 'Cliente no encontrado'}), 404
        
        monto = float(data.get('monto', 0))
        if monto <= 0:
            return jsonify({'success': False, 'error': 'Monto debe ser mayor a 0'}), 400
        
        # Crear pago
        pago = Pago(
            cliente_id=cliente_id,
            monto=monto,
            mes_correspondiente=data.get('mes_correspondiente', datetime.now().strftime('%Y-%m')),
            metodo_pago=data.get('metodo_pago', 'efectivo'),
            referencia=data.get('referencia', ''),
            notas=data.get('notas', ''),
            registrado_por=data.get('registrado_por', 'admin')
        )
        
        db.session.add(pago)

        # Calcular total pagado este mes para este cliente (incluyendo el pago actual)
        mes_pago = data.get('mes_correspondiente', datetime.now().strftime('%Y-%m'))
        from sqlalchemy import func
        total_pagado_mes = db.session.query(
            func.sum(Pago.monto)
        ).filter_by(cliente_id=cliente_id, mes_correspondiente=mes_pago).scalar() or 0
        total_pagado_mes += monto  # sumar el pago actual (aún no commiteado)

        precio = cliente.precio_mensual or 0

        # Actualizar cliente
        cliente.fecha_ultimo_pago = datetime.now()

        if precio > 0 and total_pagado_mes < precio:
            # Pago parcial: queda saldo pendiente
            cliente.saldo_pendiente = round(precio - total_pagado_mes, 2)
            # No avanzar fecha_proximo_pago hasta pagar completo
        else:
            # Pago completo
            cliente.saldo_pendiente = 0
            hoy = datetime.now()
            next_month = hoy.month + 1 if hoy.month < 12 else 1
            next_year = hoy.year if hoy.month < 12 else hoy.year + 1
            dia = min(cliente.dia_corte, 28)
            cliente.fecha_proximo_pago = datetime(next_year, next_month, dia)
        
        # Si estaba cortado o suspendido y pagó, activar automáticamente
        if cliente.estado in ['suspendido', 'cortado'] and data.get('activar_automatico', True):
            api = get_mikrotik_api()
            if api:
                if cliente.mikrotik_id:
                    api.activate_queue(cliente.mikrotik_id)
                if cliente.estado == 'cortado':
                    api.remove_from_address_list(cliente.ip_address, get_address_list_name())
            cliente.estado = 'activo'
        
        db.session.commit()
        registrar_auditoria('pago', 'pago', pago.id, f'Pago Q{monto} de {cliente.nombre}')
        
        return jsonify({
            'success': True, 
            'message': 'Pago registrado exitosamente',
            'pago': pago.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pago/<int:pago_id>', methods=['DELETE'])
@login_required
def eliminar_pago(pago_id):
    """Eliminar un pago registrado"""
    try:
        pago = Pago.query.get(pago_id)
        if not pago:
            return jsonify({'success': False, 'error': 'Pago no encontrado'}), 404
        
        # Revertir el efecto del pago en el cliente
        cliente = Cliente.query.get(pago.cliente_id)
        if cliente:
            cliente.saldo_pendiente = (cliente.saldo_pendiente or 0) + pago.monto
        
        cliente_nombre = cliente.nombre if cliente else 'Desconocido'
        monto_pago = pago.monto
        db.session.delete(pago)
        db.session.commit()
        registrar_auditoria('eliminar', 'pago', pago_id, f'Pago Q{monto_pago} eliminado de {cliente_nombre}')
        
        return jsonify({'success': True, 'message': 'Pago eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pagos', methods=['GET'])
def obtener_pagos():
    """Obtener lista de pagos"""
    cliente_id = request.args.get('cliente_id')
    
    query = Pago.query.order_by(Pago.fecha_pago.desc())
    
    if cliente_id:
        query = query.filter_by(cliente_id=cliente_id)
    
    pagos = query.limit(100).all()
    
    return jsonify({
        'success': True,
        'pagos': [p.to_dict() for p in pagos]
    })


@app.route('/api/pagos/cliente/<int:cliente_id>', methods=['GET'])
def obtener_pagos_cliente(cliente_id):
    """Obtener historial de pagos de un cliente"""
    pagos = Pago.query.filter_by(cliente_id=cliente_id).order_by(Pago.fecha_pago.desc()).all()
    return jsonify({
        'success': True,
        'pagos': [p.to_dict() for p in pagos]
    })


# ============== IMPORTAR/EXPORTAR EXCEL ==============

@app.route('/api/clientes/exportar', methods=['GET'])
def exportar_clientes():
    """Exportar clientes a Excel"""
    try:
        router_id = request.args.get('router_id', type=int)
        # Intentar usar openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Si no está instalado, exportar como CSV
            return exportar_clientes_csv(router_id=router_id)
        
        query = Cliente.query
        if router_id:
            query = query.filter_by(router_id=router_id)
        clientes = query.order_by(Cliente.nombre).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Encabezados
        headers = ['ID', 'Nombre', 'IP', 'Plan', 'Velocidad Bajada', 'Velocidad Subida', 
                   'Telefono', 'Email', 'Direccion', 'Cedula', 'Estado', 'Dia Corte',
                   'Precio Mensual', 'Fecha Registro']
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row, cliente in enumerate(clientes, 2):
            ws.cell(row=row, column=1, value=cliente.id)
            ws.cell(row=row, column=2, value=cliente.nombre)
            ws.cell(row=row, column=3, value=cliente.ip_address)
            ws.cell(row=row, column=4, value=cliente.plan)
            ws.cell(row=row, column=5, value=cliente.velocidad_download)
            ws.cell(row=row, column=6, value=cliente.velocidad_upload)
            ws.cell(row=row, column=7, value=cliente.telefono)
            ws.cell(row=row, column=8, value=cliente.email)
            ws.cell(row=row, column=9, value=cliente.direccion)
            ws.cell(row=row, column=10, value=cliente.cedula)
            ws.cell(row=row, column=11, value=cliente.estado)
            ws.cell(row=row, column=12, value=cliente.dia_corte)
            ws.cell(row=row, column=13, value=cliente.precio_mensual)
            ws.cell(row=row, column=14, value=cliente.fecha_registro.strftime('%Y-%m-%d') if cliente.fecha_registro else '')
        
        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'clientes_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def exportar_clientes_csv(router_id=None):
    """Exportar clientes a CSV (fallback)"""
    import csv
    from io import StringIO
    
    query = Cliente.query
    if router_id:
        query = query.filter_by(router_id=router_id)
    clientes = query.order_by(Cliente.nombre).all()
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Encabezados
    writer.writerow(['ID', 'Nombre', 'IP', 'Plan', 'Velocidad Bajada', 'Velocidad Subida', 
                     'Telefono', 'Email', 'Direccion', 'Cedula', 'Estado', 'Dia Corte',
                     'Precio Mensual', 'Fecha Registro'])
    
    # Datos
    for cliente in clientes:
        writer.writerow([
            cliente.id, cliente.nombre, cliente.ip_address, cliente.plan,
            cliente.velocidad_download, cliente.velocidad_upload, cliente.telefono,
            cliente.email, cliente.direccion, cliente.cedula, cliente.estado,
            cliente.dia_corte, cliente.precio_mensual,
            cliente.fecha_registro.strftime('%Y-%m-%d') if cliente.fecha_registro else ''
        ])
    
    output.seek(0)
    return send_file(
        BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'clientes_{datetime.now().strftime("%Y%m%d")}.csv'
    )


@app.route('/api/clientes/importar', methods=['POST'])
def importar_clientes():
    """Importar clientes desde archivo Excel o CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió archivo'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Nombre de archivo vacío'}), 400
        
        filename = file.filename.lower()
        
        router_id = request.form.get('router_id', type=int)
        if router_id:
            active_router_id = router_id
        else:
            config_activa = ConfigMikroTik.query.filter_by(activo=True).first()
            active_router_id = config_activa.id if config_activa else None
        
        clientes_importados = 0
        clientes_omitidos = 0
        errores = []
        
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Importar desde Excel
            try:
                from openpyxl import load_workbook
                
                wb = load_workbook(file)
                ws = wb.active
                
                # Obtener encabezados y normalizarlos
                raw_headers = [cell.value if cell.value else '' for cell in ws[1]]
                headers = [str(h).lower().strip() for h in raw_headers]
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        data = dict(zip(headers, row))
                        
                        # Buscar nombre con varias variantes
                        nombre = (data.get('nombre', '') or 
                                 data.get('nombre del cliente', '') or 
                                 data.get('nombre cliente', '') or 
                                 data.get('cliente', ''))
                        
                        # Buscar IP con varias variantes
                        ip = (data.get('ip', '') or 
                             data.get('ip_address', '') or 
                             data.get('ip address', '') or
                             data.get('direccion ip', ''))
                        
                        if not nombre or not ip:
                            clientes_omitidos += 1
                            errores.append(f"Fila {row_num}: Falta nombre o IP")
                            continue
                        
                        # Verificar si ya existe
                        if Cliente.query.filter_by(ip_address=str(ip)).first():
                            clientes_omitidos += 1
                            errores.append(f"Fila {row_num}: IP {ip} ya existe")
                            continue
                        
                        # Obtener plan
                        plan = (data.get('plan', '') or 'Basico')
                        
                        # Obtener velocidades
                        vel_down = (data.get('velocidad download', '') or 
                                   data.get('velocidad bajada', '') or 
                                   data.get('velocidad_download', '') or 
                                   data.get('download', '') or '10M')
                        vel_up = (data.get('velocidad upload', '') or 
                                 data.get('velocidad subida', '') or 
                                 data.get('velocidad_upload', '') or 
                                 data.get('upload', '') or '5M')
                        
                        # Obtener otros datos
                        telefono = data.get('telefono', '') or data.get('teléfono', '') or ''
                        email = data.get('email', '') or data.get('correo', '') or ''
                        direccion = data.get('direccion', '') or data.get('dirección', '') or ''
                        cedula = (data.get('cedula', '') or 
                                 data.get('cédula', '') or 
                                 data.get('cedula/dpi', '') or 
                                 data.get('dpi', '') or '')
                        
                        # Obtener día de corte
                        dia_corte_raw = (data.get('dia de corte', '') or 
                                        data.get('dia corte', '') or 
                                        data.get('dia_corte', '') or 
                                        data.get('día de corte', '') or 1)
                        try:
                            dia_corte = int(dia_corte_raw) if dia_corte_raw else 1
                        except:
                            dia_corte = 1
                        
                        # Obtener precio
                        precio_raw = (data.get('precio mensual (q)', '') or 
                                     data.get('precio mensual', '') or 
                                     data.get('precio_mensual', '') or 
                                     data.get('precio', '') or 0)
                        try:
                            precio = float(precio_raw) if precio_raw else 0
                        except:
                            precio = 0
                        
                        # Generar nombre del queue
                        nombre_limpio = str(nombre).replace(' ', '-').lower()[:30]
                        queue_name = f"cliente-{nombre_limpio}-{str(ip).replace('.', '-')}"
                        
                        # Crear Simple Queue en MikroTik
                        mikrotik_id = None
                        api = get_mikrotik_api(active_router_id)
                        
                        if api:
                            # PRIMERO BUSCAR si ya existe el queue por IP
                            found, existing_queue_id = api.find_queue_by_target(str(ip))
                            
                            if found and existing_queue_id:
                                # Actualizar el existente
                                success, result = api.update_simple_queue(
                                    existing_queue_id,
                                    name=queue_name,
                                    max_limit_download=str(vel_down) if vel_down else '10M',
                                    max_limit_upload=str(vel_up) if vel_up else '5M',
                                    comment=f"Cliente: {nombre}"
                                )
                                if success:
                                    mikrotik_id = existing_queue_id
                                else:
                                    errores.append(f"Fila {row_num}: Error al actualizar queue existente - {result}")
                            else:
                                # No existe, crearlo
                                success, result = api.create_simple_queue(
                                    name=queue_name,
                                    target=str(ip),
                                    max_limit_download=str(vel_down) if vel_down else '10M',
                                    max_limit_upload=str(vel_up) if vel_up else '5M',
                                    comment=f"Cliente: {nombre}"
                                )
                                
                                if success:
                                    mikrotik_id = result
                                else:
                                    errores.append(f"Fila {row_num}: Queue no creado - {result}")
                        
                        cliente = Cliente(
                            nombre=str(nombre),
                            ip_address=str(ip),
                            plan=str(plan) if plan else 'Basico',
                            velocidad_download=str(vel_down) if vel_down else '10M',
                            velocidad_upload=str(vel_up) if vel_up else '5M',
                            telefono=str(telefono) if telefono else '',
                            email=str(email) if email else '',
                            direccion=str(direccion) if direccion else '',
                            cedula=str(cedula) if cedula else '',
                            estado='activo',
                            dia_corte=dia_corte,
                            precio_mensual=precio,
                            queue_name=queue_name,
                            mikrotik_id=mikrotik_id,
                            router_id=active_router_id
                        )
                        
                        db.session.add(cliente)
                        clientes_importados += 1
                        
                    except Exception as e:
                        errores.append(f"Fila {row_num}: {str(e)}")
                        clientes_omitidos += 1
                
            except ImportError:
                return jsonify({'success': False, 'error': 'Libreria openpyxl no instalada'}), 500
        
        elif filename.endswith('.csv'):
            # Importar desde CSV
            import csv
            from io import StringIO
            
            content = file.read().decode('utf-8')
            reader = csv.DictReader(StringIO(content))
            
            for row_num, row in enumerate(reader, 2):
                try:
                    # Normalizar nombres de columnas
                    data = {k.lower(): v for k, v in row.items()}
                    
                    nombre = data.get('nombre', '')
                    ip = data.get('ip', data.get('ip_address', ''))
                    
                    if not nombre or not ip:
                        clientes_omitidos += 1
                        continue
                    
                    # Verificar si ya existe
                    if Cliente.query.filter_by(ip_address=ip).first():
                        clientes_omitidos += 1
                        continue
                    
                    cliente = Cliente(
                        nombre=nombre,
                        ip_address=ip,
                        plan=data.get('plan', 'Basico'),
                        velocidad_download=data.get('velocidad bajada', data.get('velocidad_download', '10M')),
                        velocidad_upload=data.get('velocidad subida', data.get('velocidad_upload', '5M')),
                        telefono=data.get('telefono', ''),
                        email=data.get('email', ''),
                        direccion=data.get('direccion', ''),
                        cedula=data.get('cedula', ''),
                        estado='activo',
                        dia_corte=int(data.get('dia corte', data.get('dia_corte', 1)) or 1),
                        precio_mensual=float(data.get('precio mensual', data.get('precio_mensual', 0)) or 0)
                    )
                    
                    db.session.add(cliente)
                    clientes_importados += 1
                    
                except Exception as e:
                    errores.append(f"Fila {row_num}: {str(e)}")
                    clientes_omitidos += 1
        
        else:
            return jsonify({'success': False, 'error': 'Formato no soportado. Use .xlsx, .xls o .csv'}), 400
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Importación completada',
            'importados': clientes_importados,
            'omitidos': clientes_omitidos,
            'errores': errores[:10]  # Solo mostrar primeros 10 errores
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== REPORTES ==============

@app.route('/api/reportes/resumen', methods=['GET'])
def reporte_resumen():
    """Reporte resumen general"""
    try:
        total_clientes = Cliente.query.count()
        clientes_activos = Cliente.query.filter_by(estado='activo').count()
        clientes_suspendidos = Cliente.query.filter_by(estado='suspendido').count()
        clientes_cortados = Cliente.query.filter_by(estado='cortado').count()
        
        # Pagos del mes
        hoy = datetime.now()
        primer_dia_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        pagos_mes = Pago.query.filter(Pago.fecha_pago >= primer_dia_mes).all()
        total_recaudado = sum(p.monto for p in pagos_mes)
        
        # Proyección mensual
        total_mensual_esperado = db.session.query(db.func.sum(Cliente.precio_mensual)).filter(
            Cliente.estado == 'activo'
        ).scalar() or 0
        
        # Clientes por plan
        from sqlalchemy import func
        clientes_por_plan = db.session.query(
            Cliente.plan, 
            func.count(Cliente.id)
        ).group_by(Cliente.plan).all()
        
        return jsonify({
            'success': True,
            'data': {
                'total_clientes': total_clientes,
                'clientes_activos': clientes_activos,
                'clientes_suspendidos': clientes_suspendidos,
                'clientes_cortados': clientes_cortados,
                'total_recaudado_mes': total_recaudado,
                'total_mensual_esperado': total_mensual_esperado,
                'porcentaje_recaudado': (total_recaudado / total_mensual_esperado * 100) if total_mensual_esperado > 0 else 0,
                'clientes_por_plan': [{'plan': p[0], 'cantidad': p[1]} for p in clientes_por_plan]
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reportes/pagos-mensuales', methods=['GET'])
def reporte_pagos_mensuales():
    """Reporte de pagos por mes"""
    try:
        from sqlalchemy import func, extract
        
        # Últimos 12 meses
        pagos_por_mes = db.session.query(
            func.strftime('%Y-%m', Pago.fecha_pago).label('mes'),
            func.sum(Pago.monto).label('total'),
            func.count(Pago.id).label('cantidad')
        ).group_by(
            func.strftime('%Y-%m', Pago.fecha_pago)
        ).order_by(
            func.strftime('%Y-%m', Pago.fecha_pago).desc()
        ).limit(12).all()
        
        return jsonify({
            'success': True,
            'data': [{
                'mes': p.mes,
                'total': p.total,
                'cantidad': p.cantidad
            } for p in pagos_por_mes]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== RECIBOS ==============

@app.route('/api/recibo/<int:pago_id>', methods=['GET'])
def generar_recibo(pago_id):
    """Generar recibo de pago (HTML para imprimir)"""
    try:
        pago = Pago.query.get_or_404(pago_id)
        cliente = pago.cliente
        
        return render_template('recibo.html', pago=pago, cliente=cliente)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== CONFIGURACIÓN ==============

@app.route('/api/config/mikrotik', methods=['GET'])
@login_required
def listar_routers_mikrotik():
    """Listar todos los routers MikroTik"""
    routers = ConfigMikroTik.query.all()
    return jsonify({
        'success': True,
        'routers': [{
            'id': r.id,
            'nombre': r.nombre,
            'host': r.host,
            'port': r.port,
            'username': r.username,
            'use_ssl': r.use_ssl,
            'activo': r.activo,
            'address_list_cortados': r.address_list_cortados
        } for r in routers]
    })


@app.route('/api/config/mikrotik', methods=['POST'])
@login_required
def guardar_config_mikrotik():
    """Crear o actualizar router MikroTik"""
    try:
        data = request.get_json()
        router_id = data.get('id')

        if router_id:
            # Actualizar existente
            config = ConfigMikroTik.query.get(router_id)
            if not config:
                return jsonify({'success': False, 'error': 'Router no encontrado'}), 404
        else:
            # Crear nuevo
            config = ConfigMikroTik()

        config.nombre = data.get('nombre', 'Principal').strip() or 'Principal'
        config.host = data.get('host', '').strip()
        config.port = int(data.get('port', 80))
        config.username = data.get('username', '').strip()
        config.use_ssl = data.get('use_ssl', False)
        config.address_list_cortados = data.get('address_list_cortados', 'MOROSOS').strip() or 'MOROSOS'
        config.activo = data.get('activo', True)

        # Solo actualizar password si se envió una nueva
        new_password = data.get('password', '').strip()
        if new_password:
            config.password = new_password
        elif not router_id:
            # Es un router nuevo y no tiene password
            return jsonify({'success': False, 'error': 'La contraseña es requerida para un nuevo router'}), 400

        db.session.add(config)
        db.session.commit()

        # Limpiar caché de estado
        global _mikrotik_status_cache
        _mikrotik_status_cache = {'connected': False, 'message': '', 'queue_count': 0, 'last_check': None}

        return jsonify({'success': True, 'message': 'Router guardado correctamente', 'id': config.id})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/mikrotik/<int:router_id>', methods=['DELETE'])
@login_required
def eliminar_router_mikrotik(router_id):
    """Eliminar un router MikroTik"""
    try:
        config = ConfigMikroTik.query.get(router_id)
        if not config:
            return jsonify({'success': False, 'error': 'Router no encontrado'}), 404
        db.session.delete(config)
        db.session.commit()

        # Limpiar caché de estado
        global _mikrotik_status_cache
        _mikrotik_status_cache = {'connected': False, 'message': '', 'queue_count': 0, 'last_check': None}

        return jsonify({'success': True, 'message': 'Router eliminado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/mikrotik/test', methods=['POST'])
@login_required
def probar_conexion_mikrotik():
    """Probar conexión a MikroTik"""
    try:
        data = request.get_json()

        password = data.get('password', '').strip()
        router_id = data.get('id')

        # Si no se envió password, buscar en la base de datos
        if not password and router_id:
            existing = ConfigMikroTik.query.get(router_id)
            if existing:
                password = existing.password

        if not password:
            return jsonify({'success': False, 'error': 'Contraseña requerida'}), 400

        api = MikroTikAPI(
            host=data.get('host', ''),
            username=data.get('username', ''),
            password=password,
            port=int(data.get('port', 80)),
            use_ssl=data.get('use_ssl', False)
        )

        success, result = api.test_connection()

        if success:
            return jsonify({'success': True, 'message': f'Conexión exitosa a: {result}'})
        else:
            return jsonify({'success': False, 'error': f'Error de conexión: {result}'}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== INFRAESTRUCTURA ==============

@app.route('/api/infraestructura', methods=['GET'])
@login_required
def listar_infraestructura():
    """Listar toda la infraestructura de red"""
    infraestructuras = Infraestructura.query.all()
    return jsonify({
        'success': True,
        'infraestructuras': [i.to_dict() for i in infraestructuras]
    })


@app.route('/api/infraestructura', methods=['POST'])
@login_required
def guardar_infraestructura():
    """Crear o editar infraestructura de red"""
    try:
        data = request.get_json()
        infra_id = data.get('id')
        nombre = data.get('nombre', '').strip()
        ip_address = data.get('ip_address', '').strip()
        tipo = data.get('tipo', '').strip()
        
        if not nombre or not ip_address or not tipo:
            return jsonify({'success': False, 'error': 'Nombre, IP y Tipo son campos requeridos'}), 400

        if infra_id:
            infra = Infraestructura.query.get(infra_id)
            if not infra:
                return jsonify({'success': False, 'error': 'Registro de infraestructura no encontrado'}), 404
            accion = "editar"
            detalle_anterior = f"nombre: {infra.nombre}, IP: {infra.ip_address}, tipo: {infra.tipo}"
        else:
            infra = Infraestructura()
            accion = "crear"
            detalle_anterior = ""

        infra.nombre = nombre
        infra.ip_address = ip_address
        infra.tipo = tipo
        
        mikrotik_id = data.get('mikrotik_id')
        infra.mikrotik_id = int(mikrotik_id) if (mikrotik_id and str(mikrotik_id).strip() != "") else None
        
        infra.ubicacion = data.get('ubicacion', '').strip()
        infra.modelo = data.get('modelo', '').strip()
        infra.notes = data.get('notas', '').strip() if 'notas' in data else data.get('notes', '').strip() # handle both
        infra.notas = data.get('notas', '').strip()

        db.session.add(infra)
        db.session.commit()

        # Registrar en bitácora (AuditLog)
        detalle_log = f"{accion.capitalize()} infraestructura: {infra.nombre} ({infra.ip_address}, {infra.tipo})"
        if accion == "editar":
            detalle_log += f" | Anterior: {detalle_anterior}"
            
        audit = AuditLog(
            usuario=current_user.username,
            accion=accion,
            entidad="infraestructura",
            entidad_id=infra.id,
            detalle=detalle_log,
            ip_origen=request.remote_addr
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Infraestructura guardada correctamente', 'infraestructura': infra.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/infraestructura/<int:infra_id>', methods=['DELETE'])
@login_required
def eliminar_infraestructura(infra_id):
    """Eliminar infraestructura de red"""
    try:
        infra = Infraestructura.query.get(infra_id)
        if not infra:
            return jsonify({'success': False, 'error': 'Registro de infraestructura no encontrado'}), 404

        nombre = infra.nombre
        ip_address = infra.ip_address
        tipo = infra.tipo

        db.session.delete(infra)
        db.session.commit()

        # Registrar en bitácora (AuditLog)
        audit = AuditLog(
            usuario=current_user.username,
            accion="eliminar",
            entidad="infraestructura",
            entidad_id=infra_id,
            detalle=f"Eliminar infraestructura: {nombre} ({ip_address}, {tipo})",
            ip_origen=request.remote_addr
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Infraestructura eliminada correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/planes', methods=['GET'])
def obtener_planes():
    """Obtener lista de planes"""
    planes = Plan.query.all()
    return jsonify({
        'success': True,
        'planes': [{
            'id': p.id,
            'nombre': p.nombre,
            'velocidad_download': p.velocidad_download,
            'velocidad_upload': p.velocidad_upload,
            'precio': p.precio
        } for p in planes]
    })


@app.route('/api/plan', methods=['POST'])
def crear_plan():
    """Crear nuevo plan"""
    try:
        data = request.get_json()
        
        plan = Plan(
            nombre=data.get('nombre', ''),
            velocidad_download=data.get('velocidad_download', '10M'),
            velocidad_upload=data.get('velocidad_upload', '5M'),
            precio=float(data.get('precio', 0)),
            descripcion=data.get('descripcion', '')
        )
        
        db.session.add(plan)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Plan creado'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/plan/<int:id>', methods=['PUT'])
def editar_plan(id):
    """Editar plan existente"""
    try:
        plan = Plan.query.get_or_404(id)
        data = request.get_json()
        
        plan.nombre = data.get('nombre', plan.nombre)
        plan.velocidad_download = data.get('velocidad_download', plan.velocidad_download)
        plan.velocidad_upload = data.get('velocidad_upload', plan.velocidad_upload)
        plan.precio = float(data.get('precio', plan.precio))
        plan.descripcion = data.get('descripcion', plan.descripcion)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Plan actualizado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/plan/<int:id>', methods=['DELETE'])
def eliminar_plan(id):
    """Eliminar plan"""
    try:
        plan = Plan.query.get_or_404(id)
        db.session.delete(plan)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Plan eliminado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== SINCRONIZACIÓN MIKROTIK ==============

@app.route('/api/sync/queues', methods=['GET'])
def sincronizar_queues():
    """Obtener queues de MikroTik para sincronización"""
    try:
        api = get_mikrotik_api()
        if not api:
            return jsonify({'success': False, 'error': 'MikroTik no configurado'}), 400
        
        success, queues = api.get_simple_queues()
        
        if success:
            return jsonify({'success': True, 'queues': queues})
        else:
            return jsonify({'success': False, 'error': queues}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sync/import-queues', methods=['POST'])
def importar_queues_mikrotik():
    """Importa los queues de MikroTik como clientes en la base de datos"""
    try:
        api = get_mikrotik_api()
        if not api:
            return jsonify({'success': False, 'error': 'MikroTik no configurado'}), 400
        
        success, queues = api.get_simple_queues()
        
        if not success:
            return jsonify({'success': False, 'error': f'Error al obtener queues: {queues}'}), 500
        
        importados = 0
        omitidos = 0
        errores = []
        
        for queue in queues:
            try:
                # Extraer datos del queue
                nombre = queue.get('name', '')
                target = queue.get('target', '')
                max_limit = queue.get('max-limit', '10M/10M')
                comment = queue.get('comment', '')
                queue_id = queue.get('.id', '')
                disabled = queue.get('disabled', 'false') == 'true'
                
                # Extraer IP del target (formato: IP/32 o IP)
                ip_address = target.replace('/32', '').strip()
                
                # Validar IP
                if not ip_address or not ip_address.replace('.', '').isdigit() or ip_address.count('.') != 3:
                    continue
                
                # Verificar si ya existe un cliente con esa IP
                cliente_existente = Cliente.query.filter_by(ip_address=ip_address).first()
                if cliente_existente:
                    omitidos += 1
                    continue
                
                # Parsear velocidades (formato: upload/download)
                velocidades = max_limit.split('/')
                vel_upload = velocidades[0] if len(velocidades) > 0 else '5M'
                vel_download = velocidades[1] if len(velocidades) > 1 else '10M'
                
                # Crear nuevo cliente
                nuevo_cliente = Cliente(
                    nombre=nombre,
                    ip_address=ip_address,
                    plan=comment if comment else 'Importado de MikroTik',
                    velocidad_download=vel_download,
                    velocidad_upload=vel_upload,
                    estado='activo' if not disabled else 'suspendido',
                    queue_id=queue_id,
                    queue_name=nombre
                )
                
                db.session.add(nuevo_cliente)
                importados += 1
                
            except Exception as e:
                errores.append(f"{nombre}: {str(e)}")
                continue
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'importados': importados,
            'omitidos': omitidos,
            'errores': errores[:5] if errores else [],
            'message': f'Se importaron {importados} clientes desde MikroTik. {omitidos} omitidos (ya existían).'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== INICIALIZAR DB ==============

def migrate_db():
    """Agrega columnas faltantes a tablas existentes"""
    from sqlalchemy import text, inspect
    
    with app.app_context():
        try:
            inspector = inspect(db.engine)
            
            # Detectar si es PostgreSQL o SQLite
            is_postgres = 'postgresql' in str(db.engine.url)
            datetime_type = 'TIMESTAMP' if is_postgres else 'DATETIME'
            
            # Verificar si la tabla clientes existe
            if 'clientes' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('clientes')]
                
                # Columnas a agregar si no existen
                columns_to_add = {
                    'email': 'VARCHAR(100)',
                    'cedula': 'VARCHAR(20)',
                    'dia_corte': f'INTEGER DEFAULT 1',
                    'fecha_ultimo_pago': datetime_type,
                    'fecha_proximo_pago': datetime_type,
                    'precio_mensual': 'FLOAT DEFAULT 0',
                    'saldo_pendiente': 'FLOAT DEFAULT 0',
                    'latitud': 'FLOAT',
                    'longitud': 'FLOAT',
                    'router_id': 'INTEGER',
                }
                
                for col_name, col_type in columns_to_add.items():
                    if col_name not in existing_columns:
                        try:
                            with db.engine.connect() as conn:
                                conn.execute(text(f'ALTER TABLE clientes ADD COLUMN {col_name} {col_type}'))
                                conn.commit()
                            print(f"[MIGRATION] Columna '{col_name}' agregada a clientes")
                            
                            if col_name == 'router_id':
                                try:
                                    first_r = ConfigMikroTik.query.first()
                                    if first_r:
                                        with db.engine.connect() as conn:
                                            conn.execute(text(f'UPDATE clientes SET router_id = {first_r.id} WHERE router_id IS NULL'))
                                            conn.commit()
                                        print(f"[MIGRATION] Asignados clientes existentes al router ID {first_r.id}")
                                except Exception as inner_e:
                                    print(f"[MIGRATION] Error al asignar router por defecto: {inner_e}")
                        except Exception as e:
                            print(f"[MIGRATION] Error agregando '{col_name}': {e}")
                            
            # Verificar si la tabla usuarios existe
            if 'usuarios' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('usuarios')]
                
                columns_to_add = {
                    'telefono': 'VARCHAR(20)',
                    'direccion': 'VARCHAR(200)',
                    'foto_perfil': 'VARCHAR(255)',
                    'tipo_vendedor': "VARCHAR(50) DEFAULT 'acumulativo'",
                    'comision_tipo': "VARCHAR(20) DEFAULT 'porcentaje'",
                    'comision_valor': 'FLOAT DEFAULT 0.0',
                    'limite_fichas': 'INTEGER DEFAULT 0',
                    'estado': "VARCHAR(20) DEFAULT 'activo'",
                    'permisos': 'TEXT',
                    'ultimo_acceso': datetime_type,
                    'ultima_ip': 'VARCHAR(50)'
                }
                
                for col_name, col_type in columns_to_add.items():
                    if col_name not in existing_columns:
                        try:
                            with db.engine.connect() as conn:
                                conn.execute(text(f'ALTER TABLE usuarios ADD COLUMN {col_name} {col_type}'))
                                conn.commit()
                            print(f"[MIGRATION] Columna '{col_name}' agregada a usuarios")
                        except Exception as e:
                            print(f"[MIGRATION] Error agregando '{col_name}' a usuarios: {e}")
            # Verificar si la tabla omada_vouchers existe
            if 'omada_vouchers' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('omada_vouchers')]
                
                columns_to_add = {
                    'cliente_nombre': 'VARCHAR(100)',
                    'fecha_uso': datetime_type,
                    'omada_id': 'INTEGER',
                    'lote': 'VARCHAR(100)'
                }
                
                for col_name, col_type in columns_to_add.items():
                    if col_name not in existing_columns:
                        try:
                            with db.engine.connect() as conn:
                                conn.execute(text(f'ALTER TABLE omada_vouchers ADD COLUMN {col_name} {col_type}'))
                                conn.commit()
                            print(f"[MIGRATION] Columna '{col_name}' agregada a omada_vouchers")
                        except Exception as e:
                            print(f"[MIGRATION] Error agregando '{col_name}' a omada_vouchers: {e}")

            # Verificar si la tabla config_omada existe
            if 'config_omada' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('config_omada')]
                if 'nombre' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text("ALTER TABLE config_omada ADD COLUMN nombre VARCHAR(100) DEFAULT 'Omada Principal'"))
                            conn.commit()
                        print("[MIGRATION] Columna 'nombre' agregada a config_omada")
                    except Exception as e:
                        print(f"[MIGRATION] Error al agregar nombre a config_omada: {e}")

            # Asignar los vouchers existentes al primer omada por defecto
            try:
                first_o = ConfigOmada.query.first()
                if first_o:
                    with db.engine.connect() as conn:
                        conn.execute(text(f'UPDATE omada_vouchers SET omada_id = {first_o.id} WHERE omada_id IS NULL'))
                        conn.commit()
            except Exception as e:
                print(f"[MIGRATION] Error al asignar omada_id por defecto a los vouchers: {e}")

            # Verificar si la tabla config_mikrotik existe
            if 'config_mikrotik' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('config_mikrotik')]
                
                if 'address_list_cortados' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text("ALTER TABLE config_mikrotik ADD COLUMN address_list_cortados VARCHAR(50) DEFAULT 'MOROSOS'"))
                            conn.commit()
                        print("[MIGRATION] Columna 'address_list_cortados' agregada")
                    except Exception as e:
                        print(f"[MIGRATION] Error al agregar address_list_cortados: {e}")
                
                if 'tipo' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text("ALTER TABLE config_mikrotik ADD COLUMN tipo VARCHAR(20) DEFAULT 'residential'"))
                            conn.commit()
                        print("[MIGRATION] Columna 'tipo' agregada a config_mikrotik")
                    except Exception as e:
                        print(f"[MIGRATION] Error al agregar tipo a config_mikrotik: {e}")
            
            # Verificar si la tabla usuarios existe
            if 'usuarios' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('usuarios')]
                
                if 'balance' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text("ALTER TABLE usuarios ADD COLUMN balance FLOAT DEFAULT 0.0"))
                            conn.commit()
                        print("[MIGRATION] Columna 'balance' agregada a usuarios")
                    except Exception as e:
                        print(f"[MIGRATION] Error al agregar balance a usuarios: {e}")
                        
                if 'router_id' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text("ALTER TABLE usuarios ADD COLUMN router_id INTEGER"))
                            conn.commit()
                        print("[MIGRATION] Columna 'router_id' agregada a usuarios")
                    except Exception as e:
                        print(f"[MIGRATION] Error al agregar router_id a usuarios: {e}")

            # Verificar planes_hotspot
            if 'planes_hotspot' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('planes_hotspot')]
                if 'activo' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            # Postgres usa TRUE, SQLite usa 1 o TRUE
                            conn.execute(text("ALTER TABLE planes_hotspot ADD COLUMN activo BOOLEAN DEFAULT TRUE"))
                            conn.commit()
                        print("[MIGRATION] Columna 'activo' agregada a planes_hotspot")
                    except Exception as e:
                        print(f"[MIGRATION] Error planes_hotspot: {e}")

            # Verificar vouchers
            if 'vouchers' in inspector.get_table_names():
                existing_columns = [col['name'] for col in inspector.get_columns('vouchers')]
                if 'estado' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text("ALTER TABLE vouchers ADD COLUMN estado VARCHAR(20) DEFAULT 'activo'"))
                            conn.execute(text("ALTER TABLE vouchers ADD COLUMN mikrotik_id VARCHAR(50)"))
                            conn.commit()
                    except Exception as e:
                        print(f"[MIGRATION] Error vouchers 1: {e}")
                if 'vendedor_id' not in existing_columns:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text("ALTER TABLE vouchers ADD COLUMN vendedor_id INTEGER DEFAULT 1"))
                            conn.commit()
                    except Exception as e:
                        print(f"[MIGRATION] Error vouchers 2: {e}")
            if 'clientes' in inspector.get_table_names():
                try:
                    humber_router = ConfigMikroTik.query.filter(ConfigMikroTik.nombre.ilike('%humber%')).first()
                    if humber_router:
                        with db.engine.connect() as conn:
                            res = conn.execute(text(f"SELECT COUNT(*) FROM clientes WHERE router_id IS NULL OR router_id != {humber_router.id}"))
                            count = res.scalar()
                            if count > 0:
                                conn.execute(text(f"UPDATE clientes SET router_id = {humber_router.id}"))
                                conn.commit()
                                print(f"[MIGRATION] Se migraron {count} clientes al router Humber (ID {humber_router.id})")
                    else:
                        first_r = ConfigMikroTik.query.first()
                        if first_r:
                            with db.engine.connect() as conn:
                                res = conn.execute(text("SELECT COUNT(*) FROM clientes WHERE router_id IS NULL"))
                                count = res.scalar()
                                if count > 0:
                                    conn.execute(text(f"UPDATE clientes SET router_id = {first_r.id} WHERE router_id IS NULL"))
                                    conn.commit()
                                    print(f"[MIGRATION] Se asignaron {count} clientes al router por defecto ID {first_r.id}")
                except Exception as db_e:
                    print(f"[MIGRATION] Error al verificar/migrar clientes a router Humber: {db_e}")
        except Exception as e:
            print(f"[MIGRATION] Error general: {e}")


def init_db():
    """Inicializar base de datos y crear tablas"""
    with app.app_context():
        # Activar WAL mode en SQLite para evitar bloqueos concurrentes
        if DATABASE_URL.startswith('sqlite'):
            try:
                from sqlalchemy import text as _text
                with db.engine.connect() as _conn:
                    _conn.execute(_text('PRAGMA journal_mode=WAL'))
                    _conn.execute(_text('PRAGMA busy_timeout=30000'))
                    _conn.commit()
            except Exception as _e:
                print(f"[WARNING] PRAGMA WAL falló: {_e}")

        # Crear tablas (siempre, independiente de migrate)
        db.create_all()
        print("[OK] Tablas creadas/verificadas")
        
        # Migrar columnas faltantes (error no debe detener arranque)
        try:
            migrate_db()
        except Exception as e:
            print(f"[WARNING] migrate_db falló: {e}")
        
        # Asegurar que el usuario admin existe, está activo y tiene rol de administrador
        try:
            admin = Usuario.query.filter_by(username='admin').first()
            if not admin:
                admin = Usuario(username='admin', nombre='Administrador', rol='admin')
                admin.set_password('admin')
                db.session.add(admin)
                print("[OK] Usuario admin creado (user: admin, pass: admin)")
            else:
                admin.rol = 'admin'
                admin.activo = True
                admin.set_password('admin')
                print("[OK] Usuario admin restaurado a rol admin y pass admin")
            db.session.commit()
        except Exception as e:
            print(f"[WARNING] Error asegurando admin: {e}")
        
        # Crear planes por defecto si no existen
        try:
            if Plan.query.count() == 0:
                planes_default = [
                    Plan(nombre='Basico 5Mbps', velocidad_download='5M', velocidad_upload='2M', precio=15.00),
                    Plan(nombre='Estandar 10Mbps', velocidad_download='10M', velocidad_upload='5M', precio=25.00),
                    Plan(nombre='Premium 20Mbps', velocidad_download='20M', velocidad_upload='10M', precio=35.00),
                    Plan(nombre='Ultra 50Mbps', velocidad_download='50M', velocidad_upload='25M', precio=50.00),
                    Plan(nombre='Empresarial 100Mbps', velocidad_download='100M', velocidad_upload='50M', precio=100.00),
                ]
                for plan in planes_default:
                    db.session.add(plan)
                db.session.commit()
                print("[OK] Planes por defecto creados")
        except Exception as e:
            print(f"[WARNING] Error creando planes: {e}")
        
        print("[OK] Base de datos inicializada")


# ============== API DASHBOARD CHARTS ==============

@app.route('/api/dashboard/charts')
@login_required
def dashboard_charts():
    """Datos para las gráficas del dashboard"""
    from sqlalchemy import func, extract
    
    hoy = datetime.now()
    
    # Ingresos últimos 6 meses
    ingresos_mensuales = []
    for i in range(5, -1, -1):
        fecha = hoy - timedelta(days=30*i)
        mes = fecha.month
        anio = fecha.year
        total = db.session.query(func.coalesce(func.sum(Pago.monto), 0)).filter(
            extract('month', Pago.fecha_pago) == mes,
            extract('year', Pago.fecha_pago) == anio
        ).scalar()
        nombre_mes = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][mes-1]
        ingresos_mensuales.append({'mes': f'{nombre_mes} {anio}', 'total': float(total)})
    
    # Distribución por plan
    planes_dist = db.session.query(
        Cliente.plan, func.count(Cliente.id)
    ).group_by(Cliente.plan).all()
    
    # Distribución por estado
    estados = db.session.query(
        Cliente.estado, func.count(Cliente.id)
    ).group_by(Cliente.estado).all()
    
    # Clientes nuevos por mes (últimos 6 meses)
    clientes_nuevos = []
    for i in range(5, -1, -1):
        fecha = hoy - timedelta(days=30*i)
        mes = fecha.month
        anio = fecha.year
        total = db.session.query(func.count(Cliente.id)).filter(
            extract('month', Cliente.fecha_registro) == mes,
            extract('year', Cliente.fecha_registro) == anio
        ).scalar()
        nombre_mes = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][mes-1]
        clientes_nuevos.append({'mes': f'{nombre_mes} {anio}', 'total': int(total)})
    
    return jsonify({
        'success': True,
        'ingresos_mensuales': ingresos_mensuales,
        'planes_distribucion': [{'plan': p[0], 'cantidad': p[1]} for p in planes_dist],
        'estados_distribucion': [{'estado': e[0], 'cantidad': e[1]} for e in estados],
        'clientes_nuevos': clientes_nuevos
    })


# ============== AUDIT LOG ==============

@app.route('/actividad')
@login_required
@admin_required
def actividad_view():
    """Página de registro de actividad"""
    logs = AuditLog.query.order_by(AuditLog.fecha.desc()).limit(200).all()
    return render_template('actividad.html', logs=logs)


@app.route('/api/actividad')
@login_required
def api_actividad():
    """API para obtener registros de auditoría"""
    limit = request.args.get('limit', 100, type=int)
    logs = AuditLog.query.order_by(AuditLog.fecha.desc()).limit(limit).all()
    return jsonify({'success': True, 'logs': [l.to_dict() for l in logs]})


# ============== MAPA DE CLIENTES ==============

@app.route('/mapa')
@login_required
def mapa_view():
    """Página del mapa de clientes"""
    return render_template('mapa.html')


@app.route('/api/clientes/mapa')
@login_required
def api_clientes_mapa():
    """Obtener clientes con coordenadas para el mapa"""
    clientes = Cliente.query.filter(
        Cliente.latitud.isnot(None),
        Cliente.longitud.isnot(None)
    ).all()
    return jsonify({
        'success': True,
        'clientes': [{
            'id': c.id,
            'nombre': c.nombre,
            'ip_address': c.ip_address,
            'plan': c.plan,
            'estado': c.estado,
            'latitud': c.latitud,
            'longitud': c.longitud,
            'telefono': c.telefono,
            'direccion': c.direccion
        } for c in clientes]
    })


@app.route('/api/cliente/<int:id>/ubicacion', methods=['PUT'])
@login_required
def actualizar_ubicacion(id):
    """Actualizar coordenadas de un cliente"""
    try:
        cliente = Cliente.query.get_or_404(id)
        data = request.get_json()
        cliente.latitud = data.get('latitud')
        cliente.longitud = data.get('longitud')
        db.session.commit()
        registrar_auditoria('editar', 'cliente', id, f'Ubicación actualizada: {cliente.nombre}')
        return jsonify({'success': True, 'message': 'Ubicación actualizada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== MONITOR ANCHO DE BANDA ==============

@app.route('/monitor')
@login_required
def monitor_view():
    """Página de monitoreo de ancho de banda"""
    return render_template('monitor.html')


@app.route('/api/monitor/bandwidth')
@login_required
def api_bandwidth():
    """Obtener estadísticas de bandwidth desde MikroTik"""
    api = get_mikrotik_api()
    if not api:
        return jsonify({'success': False, 'error': 'MikroTik no configurado'})
    
    try:
        success, queues = api.get_simple_queues()
        if not success:
            return jsonify({'success': False, 'error': 'No se pudo obtener datos de MikroTik'})
        
        result = []
        for q in queues:
            # MikroTik devuelve bytes como "upload/download"
            rate = q.get('rate', '0/0').split('/')
            bytes_data = q.get('bytes', '0/0').split('/')
            
            upload_rate = int(rate[0]) if len(rate) > 0 and rate[0].isdigit() else 0
            download_rate = int(rate[1]) if len(rate) > 1 and rate[1].isdigit() else 0
            upload_bytes = int(bytes_data[0]) if len(bytes_data) > 0 and bytes_data[0].isdigit() else 0
            download_bytes = int(bytes_data[1]) if len(bytes_data) > 1 and bytes_data[1].isdigit() else 0
            
            max_limit = q.get('max-limit', '0/0').split('/')
            max_upload = max_limit[0] if len(max_limit) > 0 else '0'
            max_download = max_limit[1] if len(max_limit) > 1 else '0'
            
            result.append({
                'name': q.get('name', ''),
                'target': q.get('target', ''),
                'disabled': q.get('disabled', 'false') == 'true',
                'upload_rate': upload_rate,
                'download_rate': download_rate,
                'upload_bytes': upload_bytes,
                'download_bytes': download_bytes,
                'max_limit': q.get('max-limit', '0/0'),
                'max_upload': max_upload,
                'max_download': max_download,
                'comment': q.get('comment', '')
            })
        
        return jsonify({'success': True, 'queues': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ============== NOC DASHBOARD API ==============

@app.route('/api/monitor/noc')
@login_required
def api_monitor_noc():
    """Genera datos para el dashboard NOC (simulado con clientes reales para demo)"""
    import random
    
    clientes_todos = Cliente.query.all()
    nodos = ConfigMikroTik.query.all()
    
    # 1. Clientes Offline (simulados de entre los activos)
    clientes_activos = [c for c in clientes_todos if c.estado == 'activo']
    num_offline = random.randint(1, 3)
    offline_candidates = random.sample(clientes_activos, min(num_offline, len(clientes_activos)))
    offline_clients = [{'id': c.id, 'name': c.nombre, 'ip': c.ip_address} for c in offline_candidates]
    
    # 2. Señales malas (< -75 dBm)
    num_bad = random.randint(2, 4)
    bad_signal_candidates = random.sample(clientes_activos, min(num_bad, len(clientes_activos)))
    bad_signals = [{'name': c.nombre, 'signal': random.randint(-88, -76)} for c in bad_signal_candidates if c not in offline_candidates]
    
    # 3. Consumo anormal (> 20 Mbps)
    num_high = random.randint(1, 2)
    high_bw_candidates = random.sample(clientes_activos, min(num_high, len(clientes_activos)))
    high_consumption = [{'name': c.nombre, 'speed': random.randint(25, 60)} for c in high_bw_candidates]
    
    # 4. Saturación de Nodos
    node_saturation = []
    if nodos:
        for n in nodos:
            usage = random.randint(40, 98)
            node_saturation.append({'name': n.nombre, 'usage': usage, 'status': 'Saturado' if usage > 90 else ('Alerta' if usage > 75 else 'Normal')})
    else:
        node_saturation = [
            {'name': 'SIGUANHA', 'usage': 72, 'status': 'Normal'},
            {'name': 'TICA', 'usage': 96, 'status': 'Saturado'}
        ]
        
    # 5. Estado de MikroTik
    router_status = []
    if nodos:
        for n in nodos:
            router_status.append({'name': n.nombre, 'cpu': random.randint(10, 85), 'ram': random.randint(30, 80), 'ping': random.randint(1, 15)})
    else:
        router_status = [
            {'name': 'HUMBER-CUZO', 'cpu': 25, 'ram': 42, 'ping': 4},
            {'name': 'REGI', 'cpu': 18, 'ram': 35, 'ping': 3}
        ]
        
    # 6. Alertas reales
    real_alerts = [
        {'text': "Cola sin tráfico 24 horas", 'type': 'danger'}
    ]
    if len(clientes_todos) > 0:
        real_alerts.append({'text': f"Cliente desconectado 3 días ({clientes_todos[0].nombre})", 'type': 'danger'})
    real_alerts.append({'text': "MikroTik reiniciado (Hace 2h)", 'type': 'warning'})
    real_alerts.append({'text': "Ping alto en nodo secundario", 'type': 'warning'})
    
    return jsonify({
        'success': True,
        'data': {
            'offline_clients': offline_clients,
            'bad_signals': bad_signals,
            'high_consumption': high_consumption,
            'node_saturation': node_saturation,
            'router_status': router_status,
            'real_alerts': real_alerts
        }
    })

# ============== BACKUP ==============

@app.route('/api/backup', methods=['GET'])
@login_required
def descargar_backup():
    """Descargar backup de la base de datos SQLite"""
    try:
        db_path = os.path.join(app.instance_path, 'clientes.db')
        if not os.path.exists(db_path):
            return jsonify({'success': False, 'error': 'Archivo de base de datos no encontrado'}), 404
        
        fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'cuzonet_backup_{fecha}.db'
        
        return send_file(
            db_path,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/x-sqlite3'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/backup/completo', methods=['GET'])
@login_required
def descargar_backup_completo():
    """Descargar backup completo de TODOS los datos en formato ZIP con JSONs"""
    try:
        fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # ---- Clientes ----
        clientes = Cliente.query.all()
        clientes_data = [c.to_dict() for c in clientes]
        
        # ---- Pagos ----
        pagos = Pago.query.all()
        pagos_data = [p.to_dict() for p in pagos]
        
        # ---- Planes ----
        planes = Plan.query.all()
        planes_data = [{
            'id': p.id,
            'nombre': p.nombre,
            'velocidad_download': p.velocidad_download,
            'velocidad_upload': p.velocidad_upload,
            'precio': p.precio,
            'descripcion': p.descripcion
        } for p in planes]
        
        # ---- Configuración MikroTik ----
        configs = ConfigMikroTik.query.all()
        configs_data = [{
            'id': c.id,
            'nombre': c.nombre,
            'host': c.host,
            'port': c.port,
            'username': c.username,
            'use_ssl': c.use_ssl,
            'address_list_cortados': c.address_list_cortados
        } for c in configs]
        
        # ---- Registro de Actividad ----
        logs = AuditLog.query.order_by(AuditLog.fecha.desc()).all()
        logs_data = [l.to_dict() for l in logs]
        
        # ---- Usuarios (sin contraseñas) ----
        usuarios = Usuario.query.all()
        usuarios_data = [{
            'id': u.id,
            'username': u.username,
            'nombre': u.nombre,
            'rol': u.rol,
            'activo': u.activo,
            'fecha_creacion': u.fecha_creacion.strftime('%Y-%m-%d %H:%M') if u.fecha_creacion else None
        } for u in usuarios]
        
        # ---- Resumen general ----
        resumen = {
            'fecha_backup': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_clientes': len(clientes_data),
            'total_pagos': len(pagos_data),
            'total_planes': len(planes_data),
            'total_usuarios': len(usuarios_data),
            'total_registros_actividad': len(logs_data),
            'clientes_activos': len([c for c in clientes_data if c.get('estado') == 'activo']),
            'clientes_suspendidos': len([c for c in clientes_data if c.get('estado') == 'suspendido']),
            'clientes_cortados': len([c for c in clientes_data if c.get('estado') == 'cortado']),
        }
        
        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('resumen_backup.json', json.dumps(resumen, indent=2, ensure_ascii=False))
            zf.writestr('clientes.json', json.dumps(clientes_data, indent=2, ensure_ascii=False))
            zf.writestr('pagos.json', json.dumps(pagos_data, indent=2, ensure_ascii=False))
            zf.writestr('planes.json', json.dumps(planes_data, indent=2, ensure_ascii=False))
            zf.writestr('configuracion_mikrotik.json', json.dumps(configs_data, indent=2, ensure_ascii=False))
            zf.writestr('registro_actividad.json', json.dumps(logs_data, indent=2, ensure_ascii=False))
            zf.writestr('usuarios.json', json.dumps(usuarios_data, indent=2, ensure_ascii=False))
            
            # También incluir la base de datos SQLite si existe
            db_path = os.path.join(app.instance_path, 'clientes.db')
            if os.path.exists(db_path):
                zf.write(db_path, 'clientes.db')
        
        zip_buffer.seek(0)
        nombre_archivo = f'cuzonet_backup_completo_{fecha}.zip'
        
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/zip'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/backup/info', methods=['GET'])
@login_required
def backup_info():
    """Obtener información de lo que incluiría el backup"""
    try:
        total_clientes = Cliente.query.count()
        total_pagos = Pago.query.count()
        total_planes = Plan.query.count()
        total_logs = AuditLog.query.count()
        total_usuarios = Usuario.query.count()
        
        return jsonify({
            'success': True,
            'info': {
                'clientes': total_clientes,
                'pagos': total_pagos,
                'planes': total_planes,
                'registros_actividad': total_logs,
                'usuarios': total_usuarios
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/backup/restaurar', methods=['POST'])
@login_required
def restaurar_backup():
    """Restaurar datos desde un archivo ZIP de backup completo"""
    try:
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió ningún archivo'}), 400
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'}), 400
        
        if not archivo.filename.lower().endswith('.zip'):
            return jsonify({'success': False, 'error': 'El archivo debe ser un ZIP de backup (.zip)'}), 400
        
        # Leer el ZIP en memoria
        zip_data = BytesIO(archivo.read())
        
        if not zipfile.is_zipfile(zip_data):
            return jsonify({'success': False, 'error': 'El archivo no es un ZIP válido'}), 400
        
        zip_data.seek(0)
        resultados = {
            'clientes_importados': 0,
            'pagos_importados': 0,
            'planes_importados': 0,
            'config_importada': False,
            'errores': []
        }
        
        with zipfile.ZipFile(zip_data, 'r') as zf:
            archivos_en_zip = zf.namelist()
            
            # ---- Restaurar Planes ----
            if 'planes.json' in archivos_en_zip:
                try:
                    planes_data = json.loads(zf.read('planes.json').decode('utf-8'))
                    for p in planes_data:
                        existente = Plan.query.filter_by(nombre=p['nombre']).first()
                        if not existente:
                            nuevo = Plan(
                                nombre=p['nombre'],
                                velocidad_download=p['velocidad_download'],
                                velocidad_upload=p['velocidad_upload'],
                                precio=p.get('precio', 0),
                                descripcion=p.get('descripcion', '')
                            )
                            db.session.add(nuevo)
                            resultados['planes_importados'] += 1
                        else:
                            existente.velocidad_download = p['velocidad_download']
                            existente.velocidad_upload = p['velocidad_upload']
                            existente.precio = p.get('precio', 0)
                            existente.descripcion = p.get('descripcion', '')
                            resultados['planes_importados'] += 1
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    resultados['errores'].append(f'Error en planes: {str(e)}')
            
            # ---- Restaurar Configuración MikroTik ----
            if 'configuracion_mikrotik.json' in archivos_en_zip:
                try:
                    configs_data = json.loads(zf.read('configuracion_mikrotik.json').decode('utf-8'))
                    for c in configs_data:
                        existente = ConfigMikroTik.query.first()
                        if existente:
                            existente.host = c.get('host', '')
                            existente.port = c.get('port', 80)
                            existente.username = c.get('username', '')
                            existente.use_ssl = c.get('use_ssl', False)
                            existente.address_list_cortados = c.get('address_list_cortados', 'MOROSOS')
                        else:
                            nueva = ConfigMikroTik(
                                nombre=c.get('nombre', 'Principal'),
                                host=c.get('host', ''),
                                port=c.get('port', 80),
                                username=c.get('username', ''),
                                password='',
                                use_ssl=c.get('use_ssl', False),
                                address_list_cortados=c.get('address_list_cortados', 'MOROSOS')
                            )
                            db.session.add(nueva)
                        resultados['config_importada'] = True
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    resultados['errores'].append(f'Error en config: {str(e)}')
            
            # ---- Restaurar Clientes ----
            if 'clientes.json' in archivos_en_zip:
                try:
                    clientes_data = json.loads(zf.read('clientes.json').decode('utf-8'))
                    for c in clientes_data:
                        existente = Cliente.query.filter_by(ip_address=c['ip_address']).first()
                        if not existente:
                            nuevo = Cliente(
                                nombre=c['nombre'],
                                ip_address=c['ip_address'],
                                plan=c['plan'],
                                velocidad_download=c['velocidad_download'],
                                velocidad_upload=c['velocidad_upload'],
                                telefono=c.get('telefono', ''),
                                email=c.get('email', ''),
                                direccion=c.get('direccion', ''),
                                cedula=c.get('cedula', ''),
                                estado=c.get('estado', 'activo'),
                                queue_name=c.get('queue_name', ''),
                                mikrotik_id=c.get('mikrotik_id', ''),
                                dia_corte=c.get('dia_corte', 1),
                                precio_mensual=c.get('precio_mensual', 0),
                                saldo_pendiente=c.get('saldo_pendiente', 0),
                                latitud=c.get('latitud'),
                                longitud=c.get('longitud')
                            )
                            if c.get('fecha_ultimo_pago'):
                                try:
                                    nuevo.fecha_ultimo_pago = datetime.strptime(c['fecha_ultimo_pago'], '%Y-%m-%d')
                                except:
                                    pass
                            if c.get('fecha_proximo_pago'):
                                try:
                                    nuevo.fecha_proximo_pago = datetime.strptime(c['fecha_proximo_pago'], '%Y-%m-%d')
                                except:
                                    pass
                            db.session.add(nuevo)
                            resultados['clientes_importados'] += 1
                        else:
                            # Actualizar datos del cliente existente
                            existente.nombre = c['nombre']
                            existente.plan = c['plan']
                            existente.velocidad_download = c['velocidad_download']
                            existente.velocidad_upload = c['velocidad_upload']
                            existente.telefono = c.get('telefono', '')
                            existente.email = c.get('email', '')
                            existente.direccion = c.get('direccion', '')
                            existente.cedula = c.get('cedula', '')
                            existente.estado = c.get('estado', 'activo')
                            existente.dia_corte = c.get('dia_corte', 1)
                            existente.precio_mensual = c.get('precio_mensual', 0)
                            existente.saldo_pendiente = c.get('saldo_pendiente', 0)
                            existente.latitud = c.get('latitud')
                            existente.longitud = c.get('longitud')
                            resultados['clientes_importados'] += 1
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    resultados['errores'].append(f'Error en clientes: {str(e)}')
            
            # ---- Restaurar Pagos ----
            if 'pagos.json' in archivos_en_zip:
                try:
                    pagos_data = json.loads(zf.read('pagos.json').decode('utf-8'))
                    for p in pagos_data:
                        # Buscar cliente por nombre
                        cliente = None
                        if p.get('cliente_nombre'):
                            cliente = Cliente.query.filter_by(nombre=p['cliente_nombre']).first()
                        if not cliente and p.get('cliente_id'):
                            cliente = Cliente.query.get(p['cliente_id'])
                        
                        if cliente:
                            # Verificar si el pago ya existe (por fecha y monto y cliente)
                            fecha_pago = None
                            if p.get('fecha_pago'):
                                try:
                                    fecha_pago = datetime.strptime(p['fecha_pago'], '%Y-%m-%d %H:%M')
                                except:
                                    try:
                                        fecha_pago = datetime.strptime(p['fecha_pago'], '%Y-%m-%d')
                                    except:
                                        fecha_pago = datetime.utcnow()
                            
                            pago_existente = None
                            if fecha_pago:
                                pago_existente = Pago.query.filter_by(
                                    cliente_id=cliente.id,
                                    monto=p['monto'],
                                    mes_correspondiente=p.get('mes_correspondiente', '')
                                ).filter(
                                    db.func.date(Pago.fecha_pago) == fecha_pago.date()
                                ).first()
                            
                            if not pago_existente:
                                nuevo_pago = Pago(
                                    cliente_id=cliente.id,
                                    monto=p['monto'],
                                    fecha_pago=fecha_pago or datetime.utcnow(),
                                    mes_correspondiente=p.get('mes_correspondiente', ''),
                                    metodo_pago=p.get('metodo_pago', ''),
                                    referencia=p.get('referencia', ''),
                                    notas=p.get('notas', '')
                                )
                                db.session.add(nuevo_pago)
                                resultados['pagos_importados'] += 1
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    resultados['errores'].append(f'Error en pagos: {str(e)}')
        
        # Registrar en auditoría
        try:
            log = AuditLog(
                usuario=current_user.username if current_user.is_authenticated else 'sistema',
                accion='restaurar_backup',
                entidad='sistema',
                detalle=f'Backup restaurado: {resultados["clientes_importados"]} clientes, {resultados["pagos_importados"]} pagos, {resultados["planes_importados"]} planes'
            )
            db.session.add(log)
            db.session.commit()
        except:
            pass
        
        return jsonify({
            'success': True,
            'mensaje': 'Backup restaurado exitosamente',
            'resultados': resultados
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== FICHA CLIENTE ==

@app.route('/ficha-cliente/<int:cliente_id>')
@login_required
def ficha_cliente(cliente_id):
    """Genera una ficha imprimible del cliente"""
    cliente = Cliente.query.get_or_404(cliente_id)
    pagos = Pago.query.filter_by(cliente_id=cliente_id).order_by(Pago.fecha_pago.desc()).limit(12).all()
    config = ConfigMikroTik.query.first()
    return render_template('ficha_cliente.html', cliente=cliente, pagos=pagos, config=config, now=datetime.now().strftime('%d/%m/%Y %H:%M'))


# ============== WHATSAPP BOT MANAGEMENT ==============
import subprocess
import shutil
import platform

_BOT_SESSION_DIR = '/root/cuzonet-manager/whatsapp-bot/sesion_whatsapp/session'
_BOT_QR_FILE = '/root/cuzonet-manager/static/qr_data.txt'
_IS_LINUX = platform.system() == 'Linux'
_VPS_HOST = os.getenv('VPS_HOST', '167.99.58.189')
_VPS_USER = os.getenv('VPS_USER', 'root')


def _ssh_cmd(cmd, timeout=15):
    """Ejecutar comando en el VPS por SSH desde Windows"""
    try:
        r = subprocess.run(
            ['ssh', '-o', 'StrictHostKeyChecking=no', '-o', 'ConnectTimeout=8',
             f'{_VPS_USER}@{_VPS_HOST}', cmd],
            capture_output=True, text=True, timeout=timeout
        )
        return r.stdout.strip(), r.returncode
    except Exception:
        return '', -1


@app.route('/api/whatsapp/status')
@login_required
@admin_required
def whatsapp_status():
    """Estado actual del bot de WhatsApp"""
    try:
        if _IS_LINUX:
            r = subprocess.run(['systemctl', 'is-active', 'whatsapp-bot'],
                               capture_output=True, text=True, timeout=5)
            running = r.stdout.strip() == 'active'
            qr_pending = os.path.exists(_BOT_QR_FILE) and os.path.getsize(_BOT_QR_FILE) > 0
            session_exists = os.path.exists(_BOT_SESSION_DIR)
        else:
            out, _ = _ssh_cmd(
                "systemctl is-active whatsapp-bot ; "
                f"test -s {_BOT_QR_FILE} && echo 'QR_EXISTE' || echo 'QR_VACIO' ; "
                f"test -d {_BOT_SESSION_DIR} && echo 'SESSION_EXISTE' || echo 'SESSION_NO'"
            )
            running = 'active' in out
            qr_pending = 'QR_EXISTE' in out
            session_exists = 'SESSION_EXISTE' in out

        if not running:
            estado = 'detenido'
        elif qr_pending:
            estado = 'esperando_qr'
        elif session_exists:
            estado = 'conectado'
        else:
            estado = 'iniciando'

        return jsonify({'success': True, 'running': running, 'estado': estado, 'qr_pending': qr_pending})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/whatsapp/qr')
@login_required
@admin_required
def whatsapp_qr():
    """Obtener el QR actual del bot"""
    try:
        if _IS_LINUX:
            if os.path.exists(_BOT_QR_FILE):
                with open(_BOT_QR_FILE, 'r') as f:
                    qr_data = f.read().strip()
                if qr_data:
                    return jsonify({'success': True, 'qr': qr_data})
        else:
            # Intentar leer el QR desde el VPS por HTTP primero
            try:
                resp = requests.get(f'http://{_VPS_HOST}/static/qr_data.txt', timeout=8)
                if resp.status_code == 200 and resp.text.strip():
                    return jsonify({'success': True, 'qr': resp.text.strip()})
            except Exception:
                pass
            # Fallback: leer por SSH
            qr_data, _ = _ssh_cmd(f'cat {_BOT_QR_FILE} 2>/dev/null')
            if qr_data:
                return jsonify({'success': True, 'qr': qr_data})
        return jsonify({'success': False, 'message': 'No hay QR disponible aún'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/whatsapp/reiniciar', methods=['POST'])
@login_required
@admin_required
def whatsapp_reiniciar():
    """Reiniciar el bot de WhatsApp"""
    try:
        if _IS_LINUX:
            if os.path.exists(_BOT_QR_FILE):
                os.remove(_BOT_QR_FILE)
            subprocess.run(['systemctl', 'restart', 'whatsapp-bot'], capture_output=True, timeout=15)
        else:
            _ssh_cmd(f'rm -f {_BOT_QR_FILE} ; systemctl restart whatsapp-bot', timeout=20)
        registrar_auditoria('whatsapp_reiniciar', 'whatsapp', None, 'Bot de WhatsApp reiniciado')
        return jsonify({'success': True, 'message': 'Bot reiniciado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/whatsapp/cambiar-numero', methods=['POST'])
@login_required
@admin_required
def whatsapp_cambiar_numero():
    """Borrar sesión actual y reiniciar para vincular un número diferente"""
    try:
        if _IS_LINUX:
            subprocess.run(['systemctl', 'stop', 'whatsapp-bot'], capture_output=True, timeout=10)
            if os.path.exists(_BOT_SESSION_DIR):
                shutil.rmtree(_BOT_SESSION_DIR)
            if os.path.exists(_BOT_QR_FILE):
                os.remove(_BOT_QR_FILE)
            subprocess.run(['systemctl', 'start', 'whatsapp-bot'], capture_output=True, timeout=10)
        else:
            _ssh_cmd(
                f'systemctl stop whatsapp-bot ; '
                f'rm -rf {_BOT_SESSION_DIR} ; '
                f'rm -f {_BOT_QR_FILE} ; '
                f'systemctl start whatsapp-bot',
                timeout=25
            )
        registrar_auditoria('whatsapp_cambiar_numero', 'whatsapp', None, 'Sesión eliminada para vincular nuevo número')
        return jsonify({'success': True, 'message': 'Sesión eliminada. Escanea el nuevo QR para vincular otro número.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# Ejecutar migración al importar (para gunicorn) de forma asíncrona para no bloquear el health check
import threading

def run_init_db_async():
    try:
        import fcntl as _fcntl
        import tempfile as _tempfile
        
        _lock_path = os.path.join(_tempfile.gettempdir(), 'cuzonet_init.lock')
        _lock_fd = open(_lock_path, 'w')
        _fcntl.flock(_lock_fd, _fcntl.LOCK_EX)
        try:
            init_db()
            with app.app_context():
                try:
                    db.create_all()
                    from sqlalchemy import text
                    db.session.execute(text('ALTER TABLE vouchers ADD COLUMN lote_id INTEGER REFERENCES lotes_fichas(id)'))
                    db.session.commit()
                    print("Migración: Columna lote_id agregada con éxito.")
                except Exception as e:
                    db.session.rollback()
        finally:
            _fcntl.flock(_lock_fd, _fcntl.LOCK_UN)
            _lock_fd.close()
    except (ImportError, OSError):
        try:
            init_db()
            with app.app_context():
                try:
                    db.create_all()
                    from sqlalchemy import text
                    db.session.execute(text('ALTER TABLE vouchers ADD COLUMN lote_id INTEGER REFERENCES lotes_fichas(id)'))
                    db.session.commit()
                    print("Migración: Columna lote_id agregada con éxito.")
                except Exception as e:
                    db.session.rollback()
        except Exception as e:
            print(f"[WARNING] init_db falló: {e}")
    except Exception as e:
        print(f"[WARNING] run_init_db_async falló: {e}")
        print(f"[WARNING] init_db falló: {e}")

# Arrancar la inicialización de base de datos de fondo
threading.Thread(target=run_init_db_async, daemon=True).start()


def vendedor_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if current_user.rol != 'vendedor':
            flash('Acceso exclusivo para vendedores', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# ============== RUTAS DE HOTSPOT & VOUCHERS ==============

@app.route('/admin/hotspot')
@login_required
@admin_required
def hotspot_admin():
    """Panel de administración de Hotspots y Vendedores"""
    planes = PlanHotspot.query.all()
    routers = ConfigMikroTik.query.filter_by(tipo='hotspot').all()
    todos_routers = ConfigMikroTik.query.all()
    vendedores = Usuario.query.filter_by(rol='vendedor').all()
    transacciones = TransaccionVendedor.query.order_by(TransaccionVendedor.fecha.desc()).limit(50).all()
    
    total_vouchers = Voucher.query.count()
    ganancia_total = db.session.query(db.func.sum(Voucher.precio)).scalar() or 0.0
    
    return render_template('hotspot_dashboard.html',
                           planes=planes,
                           routers=routers,
                           todos_routers=todos_routers,
                           vendedores=vendedores,
                           transacciones=transacciones,
                           total_vouchers=total_vouchers,
                           ganancia_total=ganancia_total)


@app.route('/admin/hotspot/router/guardar', methods=['POST'])
@login_required
@admin_required
def hotspot_router_guardar():
    """Guardar o editar un router asignándolo como tipo hotspot"""
    router_id = request.form.get('router_id')
    tipo_accion = request.form.get('accion_tipo', 'editar')
    
    if tipo_accion == 'crear':
        nombre = request.form.get('nombre')
        host = request.form.get('host')
        port = int(request.form.get('port', 80))
        username = request.form.get('username')
        password = request.form.get('password')
        use_ssl = request.form.get('use_ssl') == 'on'
        
        router = ConfigMikroTik(
            nombre=nombre,
            host=host,
            port=port,
            username=username,
            password=password,
            use_ssl=use_ssl,
            tipo='hotspot'
        )
        db.session.add(router)
        flash('Router Hotspot creado con éxito', 'success')
    else:
        router = ConfigMikroTik.query.get_or_404(router_id)
        router.tipo = 'hotspot'
        if request.form.get('nombre'):
            router.nombre = request.form.get('nombre')
        if request.form.get('host'):
            router.host = request.form.get('host')
        if request.form.get('port'):
            router.port = int(request.form.get('port'))
        if request.form.get('username'):
            router.username = request.form.get('username')
        if request.form.get('password'):
            router.password = request.form.get('password')
        router.use_ssl = request.form.get('use_ssl') == 'on'
        flash('Router configurado como Hotspot', 'success')
        
    db.session.commit()
    registrar_auditoria('guardar_router_hotspot', 'config_mikrotik', router.id if 'router' in locals() else None, 'Router Hotspot guardado')
    return redirect(url_for('hotspot_admin'))


@app.route('/admin/hotspot/router/cambiar-tipo/<int:id>/<string:nuevo_tipo>')
@login_required
@admin_required
def hotspot_router_cambiar_tipo(id, nuevo_tipo):
    """Cambiar el tipo de un router (residential o hotspot)"""
    router = ConfigMikroTik.query.get_or_404(id)
    if nuevo_tipo in ['residential', 'hotspot']:
        router.tipo = nuevo_tipo
        db.session.commit()
        flash(f'Router {router.nombre} cambiado a tipo {nuevo_tipo}', 'success')
    return redirect(url_for('hotspot_admin'))


@app.route('/admin/hotspot/plan/guardar', methods=['POST'])
@login_required
@admin_required
def hotspot_plan_guardar():
    """Guardar o editar un plan de hotspot"""
    plan_id = request.form.get('plan_id')
    nombre = request.form.get('nombre')
    precio = float(request.form.get('precio', 0))
    perfil_hotspot = request.form.get('perfil_hotspot')
    limit_uptime = request.form.get('limit_uptime', '').strip() or None
    
    if plan_id:
        plan = PlanHotspot.query.get_or_404(plan_id)
        plan.nombre = nombre
        plan.precio = precio
        plan.perfil_hotspot = perfil_hotspot
        plan.limit_uptime = limit_uptime
        flash('Plan de Hotspot actualizado con éxito', 'success')
    else:
        plan = PlanHotspot(
            nombre=nombre,
            precio=precio,
            perfil_hotspot=perfil_hotspot,
            limit_uptime=limit_uptime
        )
        db.session.add(plan)
        flash('Plan de Hotspot creado con éxito', 'success')
        
    db.session.commit()
    registrar_auditoria('guardar_plan_hotspot', 'planes_hotspot', plan.id, f'Plan Hotspot: {nombre}')
    return redirect(url_for('hotspot_admin'))


@app.route('/admin/hotspot/plan/desactivar/<int:id>')
@login_required
@admin_required
def hotspot_plan_desactivar(id):
    plan = PlanHotspot.query.get_or_404(id)
    plan.activo = not plan.activo
    db.session.commit()
    estado = 'activado' if plan.activo else 'desactivado'
    flash(f'Plan {plan.nombre} {estado} con éxito', 'success')
    return redirect(url_for('hotspot_admin'))


@app.route('/admin/hotspot/vendedores')
@login_required
@admin_required
def hotspot_vendedores():
    from datetime import datetime
    
    vendedores = Usuario.query.filter_by(rol='vendedor').all()
    routers = ConfigMikroTik.query.filter_by(activo=True).all()
    
    hoy = datetime.utcnow().date()
    mes_actual = hoy.month
    año_actual = hoy.year
    
    vendedores_data = []
    for v in vendedores:
        fichas_mikrotik = Voucher.query.filter_by(vendedor_id=v.id).count()
        fichas_omada = OmadaVoucher.query.filter_by(vendedor_id=v.id).filter(OmadaVoucher.estado != 'eliminado').count()
        fichas_impresas = fichas_mikrotik + fichas_omada
        abonos = TransaccionVendedor.query.filter_by(vendedor_id=v.id, tipo='abono').all()
        total_abonado = sum(a.monto for a in abonos)
        
        # Calcular ventas totales históricas de Omada
        omada_vendidas = OmadaVoucher.query.filter_by(vendedor_id=v.id).filter(OmadaVoucher.estado.in_(['usado', 'vencido'])).all()
        ventas_omada = sum(ov.precio for ov in omada_vendidas)
        fichas_omada_vendidas = len(omada_vendidas)
        
        # Calcular ventas totales históricas de MikroTik
        mt_vendidas = Voucher.query.filter_by(vendedor_id=v.id).all()
        ventas_mt = sum(vt.precio for vt in mt_vendidas)
        fichas_mt_vendidas = len(mt_vendidas)
        
        ventas_totales = ventas_omada + ventas_mt
        
        # Calcular ventas del mes actual
        fichas_mes_mt = Voucher.query.filter(
            Voucher.vendedor_id == v.id,
            db.func.extract('month', Voucher.fecha_creacion) == mes_actual,
            db.func.extract('year', Voucher.fecha_creacion) == año_actual
        ).all()
        fichas_mes_om = OmadaVoucher.query.filter(
            OmadaVoucher.vendedor_id == v.id,
            db.func.extract('month', OmadaVoucher.fecha_uso) == mes_actual,
            db.func.extract('year', OmadaVoucher.fecha_uso) == año_actual,
            OmadaVoucher.estado.in_(['usado', 'vencido'])
        ).all()
        ventas_mes = sum(vm.precio for vm in fichas_mes_mt) + sum(vo.precio for vo in fichas_mes_om)
        
        # Calcular Ganancias del Vendedor este mes
        ganancias_mes = 0
        if v.comision_tipo == 'porcentaje':
            ganancias_mes = ventas_mes * (v.comision_valor / 100.0)
        else:
            ganancias_mes = (len(fichas_mes_mt) + len(fichas_mes_om)) * v.comision_valor
        
        # Deuda pendiente (sobre el total)
        deuda_pendiente = ventas_totales - total_abonado
        
        vendedores_data.append({
            'vendedor': v,
            'fichas_impresas': fichas_impresas,
            'fichas_mikrotik': fichas_mikrotik,
            'fichas_omada': fichas_omada,
            'fichas_omada_vendidas': fichas_omada_vendidas,
            'ventas_omada': ventas_omada,
            'ventas_totales': ventas_totales,
            'ventas_mes': ventas_mes,
            'ganancias_mes': ganancias_mes,
            'total_abonado': total_abonado,
            'deuda_pendiente': deuda_pendiente
        })
        
    # Ordenar por ventas_mes para el Ranking
    vendedores_data.sort(key=lambda x: x['ventas_mes'], reverse=True)
    
    # Asignar ranking
    for idx, data in enumerate(vendedores_data):
        data['ranking'] = idx + 1
        
    return render_template('hotspot_vendedores.html', 
                           vendedores_data=vendedores_data,
                           routers=routers)

@app.route('/admin/hotspot/vendedor/nuevo')
@login_required
@admin_required
def hotspot_vendedor_nuevo():
    routers = ConfigMikroTik.query.filter_by(activo=True).all()
    vendedor_id = request.args.get('edit')
    vendedor = None
    if vendedor_id:
        vendedor = Usuario.query.get(vendedor_id)
    return render_template('hotspot_vendedor_crear.html', routers=routers, vendedor=vendedor)

@app.route('/admin/hotspot/control-vendedores')
@login_required
@admin_required
def hotspot_control_vendedores():
    lotes = InventarioManualLote.query.order_by(InventarioManualLote.id.desc()).all()
    vendedores = InventarioManualVendedor.query.order_by(InventarioManualVendedor.id.desc()).all()
    
    return render_template('hotspot_control_vendedores.html', 
                           lotes=lotes, 
                           vendedores=vendedores)

@app.route('/api/hotspot/inventario-manual/update', methods=['POST'])
@login_required
@admin_required
def update_inventario_manual():
    data = request.json
    tipo = data.get('tipo')  # 'lote' o 'vendedor'
    item_id = data.get('id')
    campo = data.get('campo')
    valor = data.get('valor')
    
    try:
        if tipo == 'lote':
            item = InventarioManualLote.query.get(item_id)
            if item and hasattr(item, campo):
                if campo in ['cantidad', 'stock', 'vendidas', 'asignadas']:
                    setattr(item, campo, int(valor) if str(valor).isdigit() else 0)
                else:
                    setattr(item, campo, str(valor))
        elif tipo == 'vendedor':
            item = InventarioManualVendedor.query.get(item_id)
            if item and hasattr(item, campo):
                if campo in ['stock', 'vendidas']:
                    setattr(item, campo, int(valor) if str(valor).isdigit() else 0)
                else:
                    setattr(item, campo, str(valor))
                    
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/hotspot/inventario-manual/add', methods=['POST'])
@login_required
@admin_required
def add_inventario_manual():
    data = request.json
    tipo = data.get('tipo')
    
    try:
        if tipo == 'lote':
            nuevo = InventarioManualLote(
                lote=f"L00{InventarioManualLote.query.count() + 1}",
                plan="Nuevo Plan",
                cantidad=0, stock=0, vendidas=0, asignadas=0,
                vendedor_asignado="", fecha_asignacion="", estado="Activo"
            )
            db.session.add(nuevo)
        elif tipo == 'vendedor':
            nuevo = InventarioManualVendedor(
                vendedor="Nuevo Vendedor",
                stock=0, vendidas=0, dinero="Q0", comision="Q0"
            )
            db.session.add(nuevo)
            
        db.session.commit()
        return jsonify({'success': True, 'id': nuevo.id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/force-update-db')
def force_update_db():
    from sqlalchemy import text
    resultados = []
    try:
        db.session.execute(text("ALTER TABLE inventario_manual_lote ADD COLUMN vendedor_asignado VARCHAR(100) DEFAULT ''"))
        db.session.commit()
        resultados.append("Columna vendedor_asignado agregada.")
    except Exception as e:
        db.session.rollback()
        resultados.append(f"Error vendedor_asignado: {str(e)}")
        
    try:
        db.session.execute(text("ALTER TABLE inventario_manual_lote ADD COLUMN fecha_asignacion VARCHAR(50) DEFAULT ''"))
        db.session.commit()
        resultados.append("Columna fecha_asignacion agregada.")
    except Exception as e:
        db.session.rollback()
        resultados.append(f"Error fecha_asignacion: {str(e)}")
        
    return jsonify({"resultados": resultados})

@app.route('/admin/hotspot/vouchers_grid')
@login_required
@admin_required
def hotspot_vouchers_grid():
    """Vista de Vouchers estilo Mikhmon (Grid de colores por perfil)"""
    routers = ConfigMikroTik.query.filter_by(activo=True).all()
    selected_router_id = request.args.get('router_id')
    
    if not selected_router_id and routers:
        selected_router_id = routers[0].id
        
    router = None
    profiles_data = []
    error = None
    
    if selected_router_id:
        router = ConfigMikroTik.query.get(selected_router_id)
        if router:
            api = MikroTikAPI(router.host, router.username, router.password, router.port, router.use_ssl)
            success, result = api.get_hotspot_profiles_with_counts()
            if success:
                profiles_data = result
            else:
                error = result
                
    # Asignar un color fijo o pseudo-aleatorio basado en el nombre del perfil
    # para imitar la paleta de colores de Mikhmon
    mikhmon_colors = [
        '#ffc107', # Amarillo
        '#6c757d', # Gris
        '#6f42c1', # Morado
        '#28a745', # Verde
        '#e83e8c', # Rosa
        '#fd7e14', # Naranja/Rojo
        '#17a2b8', # Teal
        '#20c997', # Cian
    ]
    
    for i, p in enumerate(profiles_data):
        # Asignar color secuencialmente (rotando)
        p['color'] = mikhmon_colors[i % len(mikhmon_colors)]
        
    return render_template('hotspot_vouchers_grid.html', 
                           routers=routers, 
                           selected_router_id=int(selected_router_id) if selected_router_id else None,
                           profiles_data=profiles_data,
                           router=router,
                           error=error)

@app.route('/admin/hotspot/mikhmon_users')
@login_required
@admin_required
def hotspot_mikhmon_users():
    """Lista de usuarios estilo Mikhmon"""
    routers = ConfigMikroTik.query.filter_by(activo=True).all()
    selected_router_id = request.args.get('router_id')
    profile_filter = request.args.get('profile')
    comment_filter = request.args.get('comment')
    
    if not selected_router_id and routers:
        selected_router_id = routers[0].id
        
    router = None
    users = []
    profiles = []
    comments = []
    error = None
    
    if selected_router_id:
        router = ConfigMikroTik.query.get(selected_router_id)
        if router:
            api = MikroTikAPI(router.host, router.username, router.password, router.port, router.use_ssl)
            # Usar la base de datos local (Sincronizada) en lugar de la API en vivo
            query = HotspotUserSync.query.filter_by(router_id=router.id)
            
            if profile_filter and profile_filter != 'all':
                query = query.filter_by(profile=profile_filter)
                
            if comment_filter:
                query = query.filter_by(comment=comment_filter)
                
            users_db = query.all()
            
            # Formatear a diccionario como lo esperaba la plantilla
            for u in users_db:
                users.append({
                    'name': u.name,
                    'profile': u.profile,
                    'server': u.server,
                    'uptime': u.uptime,
                    'bytes-in': u.bytes_in,
                    'bytes-out': u.bytes_out,
                    'comment': u.comment
                })
                
            # Extraer comentarios únicos (de este router) para el filtro
            all_users_for_comments = HotspotUserSync.query.filter_by(router_id=router.id).all()
            comments = list(set([u.comment for u in all_users_for_comments if u.comment]))
            comments.sort()
            
            # Obtener perfiles para el filtro (sin conteos por ahora, o hacer query en la DB)
            profiles = list(set([u.profile for u in all_users_for_comments if u.profile]))
            profiles.sort()
            
            if not users_db and not all_users_for_comments:
                error = "Base de datos local vacía. Por favor haz clic en 'Sincronizar con MikroTik'."
                
    return render_template('hotspot_mikhmon_users.html', 
                           routers=routers, 
                           selected_router_id=int(selected_router_id) if selected_router_id else None,
                           users=users,
                           profiles=profiles,
                           comments=comments,
                           selected_profile=profile_filter,
                           selected_comment=comment_filter,
                           router=router,
                           error=error)

@app.route('/api/hotspot/sync_users', methods=['POST'])
@login_required
@admin_required
def hotspot_sync_users():
    """Descarga todos los usuarios del MikroTik y actualiza la DB local para lectura ultrarrápida"""
    data = request.get_json() or {}
    router_id = data.get('router_id')
    
    if not router_id:
        return jsonify({'success': False, 'error': 'No router specified'})
        
    router = ConfigMikroTik.query.get(router_id)
    if not router:
        return jsonify({'success': False, 'error': 'Router not found'})
        
    api = MikroTikAPI(router.host, router.username, router.password, router.port, router.use_ssl)
    
    success, result = api.get_hotspot_users(profile=None)
    
    if not success:
        return jsonify({'success': False, 'error': result})
        
    try:
        # 1. Borrar tabla actual para este router
        HotspotUserSync.query.filter_by(router_id=router.id).delete()
        
        # 2. Insertar masivamente (bulk_save_objects)
        objects = []
        for u in result:
            obj = HotspotUserSync(
                router_id=router.id,
                mikrotik_id=u.get('.id'),
                server=u.get('server', 'all'),
                name=u.get('name', ''),
                password=u.get('password', ''),
                profile=u.get('profile', ''),
                uptime=u.get('uptime', '00:00:00'),
                bytes_in=u.get('bytes-in', '0'),
                bytes_out=u.get('bytes-out', '0'),
                comment=u.get('comment', '')
            )
            objects.append(obj)
            
        db.session.bulk_save_objects(objects)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'{len(objects)} usuarios sincronizados exitosamente.'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/hotspot/live/dashboard', methods=['GET'])
@login_required
@admin_required
def hotspot_live_dashboard_api():
    """Endpoint que retorna los datos en vivo del router seleccionado"""
    router_id = request.args.get('router_id')
    if not router_id:
        return jsonify({'error': 'Router ID requerido'})
        
    router = ConfigMikroTik.query.get(router_id)
    if not router:
        return jsonify({'error': 'Router no encontrado'})
        
    api = MikroTikAPI(router.host, router.username, router.password, router.port, router.use_ssl)
    data = api.get_live_dashboard_data()
    
    return jsonify(data)

@app.route('/admin/hotspot/fichas')
@login_required
@admin_required
def hotspot_fichas():
    routers = ConfigMikroTik.query.filter_by(activo=True).all()
    return render_template('hotspot_impresion_multiple.html', routers=routers)

@app.route('/admin/hotspot/lotes')
@login_required
@admin_required
def hotspot_lotes():
    try:
        """Monitor de Lotes de Fichas (Estilo VoucherForge)"""
        lotes = LoteFichas.query.order_by(LoteFichas.fecha_creacion.desc()).all()
        
        # Calcular estadísticas por lote
        lotes_stats = []
        for lote in lotes:
            vouchers = Voucher.query.filter_by(lote_id=lote.id).all()
            vendidos = sum(1 for v in vouchers if v.estado == 'usado')
            en_stock = len(vouchers) - vendidos
            lotes_stats.append({
                'lote': lote,
                'total': len(vouchers),
                'vendidos': vendidos,
                'en_stock': en_stock,
                'recaudado': sum(v.precio for v in vouchers if v.estado == 'usado' and v.precio),
                'potencial': sum(v.precio for v in vouchers if v.precio)
            })
            
        return render_template('hotspot_lotes.html', lotes_stats=lotes_stats)
    except Exception as e:
        import traceback
        return f"<h1>Error Interno</h1><pre>{traceback.format_exc()}</pre>"

@app.route('/admin/hotspot/lotes/<int:lote_id>/imprimir')
@login_required
@admin_required
def hotspot_lote_imprimir(lote_id):
    try:
        """Genera la vista de impresión visual de los tickets de un lote"""
        lote = LoteFichas.query.get_or_404(lote_id)
        vouchers = Voucher.query.filter_by(lote_id=lote.id).all()
        
        # Necesitamos pasar la IP del router o el DNS para armar el código QR
        # Normalmente es http://<router_ip>/login?username=X&password=Y
        url_login = f"http://{lote.router.host}/login"
        
        return render_template('hotspot_imprimir_lote.html', lote=lote, vouchers=vouchers, url_login=url_login)
    except Exception as e:
        import traceback
        return f"<h1>Error Interno</h1><pre>{traceback.format_exc()}</pre>"

@app.route('/api/hotspot/get_profiles')
@login_required
@admin_required
def api_hotspot_get_profiles():
    router_id = request.args.get('router_id')
    if not router_id:
        return jsonify({'error': 'Falta router_id'}), 400
        
    router = ConfigMikroTik.query.get(router_id)
    if not router:
        return jsonify({'error': 'Router no encontrado'}), 404
        
    api = MikroTikAPI(router.host, router.username, router.password, router.port, router.use_ssl)
    conectado, _ = api.test_connection()
    if not conectado:
        return jsonify({'error': 'No se pudo conectar al router'}), 500
        
    success, result = api.get_hotspot_profiles()
    if success:
        return jsonify({'profiles': result})
    else:
        return jsonify({'error': result}), 500

@app.route('/admin/hotspot/omada_print')
@login_required
@admin_required
def hotspot_omada_print():
    from datetime import datetime
    vendedores = Usuario.query.filter_by(rol='vendedor').all()
    omadas = ConfigOmada.query.all()
    
    hoy = datetime.utcnow().date()
    
    # Global stats
    lotes_hoy = db.session.query(OmadaVoucher.lote).filter(
        db.func.extract('year', OmadaVoucher.fecha_creacion) == hoy.year,
        db.func.extract('month', OmadaVoucher.fecha_creacion) == hoy.month,
        db.func.extract('day', OmadaVoucher.fecha_creacion) == hoy.day,
        OmadaVoucher.lote != None
    ).distinct().count()
    
    fichas_impresas_hoy = OmadaVoucher.query.filter(
        db.func.extract('year', OmadaVoucher.fecha_creacion) == hoy.year,
        db.func.extract('month', OmadaVoucher.fecha_creacion) == hoy.month,
        db.func.extract('day', OmadaVoucher.fecha_creacion) == hoy.day
    ).count()
    
    ultima_op = OmadaVoucher.query.order_by(OmadaVoucher.fecha_creacion.desc()).first()
    ultima_op_str = ultima_op.fecha_creacion.strftime('%H:%M:%S') if ultima_op else 'N/A'
    
    for v in vendedores:
        # Ventas del día (omada)
        fichas_hoy_usadas = OmadaVoucher.query.filter(
            OmadaVoucher.vendedor_id == v.id,
            OmadaVoucher.estado.in_(['usado', 'vencido']),
            db.func.extract('year', OmadaVoucher.fecha_uso) == hoy.year,
            db.func.extract('month', OmadaVoucher.fecha_uso) == hoy.month,
            db.func.extract('day', OmadaVoucher.fecha_uso) == hoy.day
        ).all()
        v.ventas_hoy = sum(f.precio for f in fichas_hoy_usadas)
        
        # Inventario
        v.fichas_disponibles = OmadaVoucher.query.filter_by(vendedor_id=v.id, estado='activo').count()
        
        # Comision acumulada total
        todas_ventas = OmadaVoucher.query.filter(
            OmadaVoucher.vendedor_id == v.id,
            OmadaVoucher.estado.in_(['usado', 'vencido'])
        ).all()
        ventas_totales = sum(f.precio for f in todas_ventas)
        if v.comision_tipo == 'porcentaje':
            v.comision_acumulada = ventas_totales * (v.comision_valor / 100.0)
        else:
            v.comision_acumulada = len(todas_ventas) * v.comision_valor
            
    return render_template('cuzonet_print_studio.html', 
                           vendedores=vendedores, 
                           omadas=omadas,
                           lotes_hoy=lotes_hoy,
                           fichas_impresas_hoy=fichas_impresas_hoy,
                           ultima_op=ultima_op_str)

@app.route('/admin/hotspot/omada-historial')
@login_required
@admin_required
def hotspot_omada_historial():
    omadas_configs = ConfigOmada.query.filter_by(activo=True).all()
    total_sync = 0
    errores_sync = []
    changed = False
    
    if len(omadas_configs) > 0:
        from omada_api import OmadaAPI
        status_map_global = {}
        
        for config in omadas_configs:
            try:
                omada = OmadaAPI(config.url, config.username, config.password, config.site_id)
                status_map = omada.get_all_vouchers_status()
                status_map_global.update(status_map)
                total_sync += len(status_map)
            except Exception as e:
                errores_sync.append(f"Error en {config.nombre}: {str(e)}")
                
        # Actualizar DB local
        if not errores_sync:
            all_vouchers_local = OmadaVoucher.query.filter(OmadaVoucher.estado != 'eliminado').all()
            for v_local in all_vouchers_local:
                if v_local.codigo in status_map_global:
                    try:
                        omada_status = int(status_map_global[v_local.codigo])
                    except ValueError:
                        omada_status = 0
                        
                    nuevo_estado = 'activo'
                    if omada_status == 1:
                        nuevo_estado = 'usado'
                        if not v_local.fecha_uso:
                            v_local.fecha_uso = datetime.utcnow()
                    elif omada_status in (2, 3, 4): # Considerar otros estados como vencidos
                        nuevo_estado = 'vencido'
                        
                    if v_local.estado != nuevo_estado:
                        v_local.estado = nuevo_estado
                        changed = True
                else:
                    if v_local.estado != 'eliminado':
                        v_local.estado = 'eliminado'
                        changed = True
                        
        if changed:
            db.session.commit()
            
        if errores_sync:
            sync_error = " | ".join(errores_sync)
        else:
            sync_info = f"Sincronización OK. Descargados {total_sync} vouchers desde todos los controladores."


    # Si es vendedor, solo ver las suyas
    if current_user.rol == 'vendedor':
        vouchers = OmadaVoucher.query.filter_by(vendedor_id=current_user.id).order_by(OmadaVoucher.fecha_creacion.desc()).all()
    else:
        vendedor_id = request.args.get('vendedor_id')
        if vendedor_id:
            vouchers = OmadaVoucher.query.filter_by(vendedor_id=vendedor_id).order_by(OmadaVoucher.fecha_creacion.desc()).all()
        else:
            vouchers = OmadaVoucher.query.order_by(OmadaVoucher.fecha_creacion.desc()).all()
            
    from datetime import datetime, timedelta
    
    hoy = datetime.utcnow().date()
    fichas_hoy = 0
    ingresos_hoy = 0.0
    estado_counts = {'activo': 0, 'usado': 0, 'vencido': 0, 'eliminado': 0}
    
    for v in vouchers:
        # Contar estados
        if v.estado in estado_counts:
            estado_counts[v.estado] += 1
            
        # Calcular hoy
        if v.fecha_creacion.date() == hoy:
            fichas_hoy += 1
            ingresos_hoy += v.precio

    # Datos para gráfico: ventas de los últimos 7 días
    ventas_7_dias = {}
    for i in range(7):
        d = hoy - timedelta(days=i)
        ventas_7_dias[d.strftime('%d/%m')] = 0
        
    for v in vouchers:
        fecha_str = v.fecha_creacion.strftime('%d/%m')
        if fecha_str in ventas_7_dias and v.estado != 'eliminado':
            ventas_7_dias[fecha_str] += v.precio
            
    # Ordenar chronológicamente para el gráfico
    labels_grafico = list(reversed(list(ventas_7_dias.keys())))
    datos_grafico = list(reversed(list(ventas_7_dias.values())))

    vendedores = Usuario.query.filter_by(rol='vendedor').all()
    return render_template(
        'omada_historial.html', 
        vouchers=vouchers, 
        vendedores=vendedores,
        fichas_hoy=fichas_hoy,
        ingresos_hoy=ingresos_hoy,
        estado_counts=estado_counts,
        labels_grafico=labels_grafico,
        datos_grafico=datos_grafico
    )

@app.route('/admin/hotspot/omada-debug')
@login_required
def hotspot_omada_debug():
    try:
        config = ConfigOmada.query.first()
        if not config or not config.activo:
            return "Omada no configurado"
        from omada_api import OmadaAPI
        omada = OmadaAPI(config.url, config.username, config.password, config.site_id)
        omada.login()
        query_url = f"{omada.base_url}/{omada.omadac_id}/api/v2/hotspot/sites/{omada.site_id}/vouchers?currentPage=1&currentPageSize=10"
        res = omada.session.get(query_url, timeout=15)
        return res.json()
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/api/omada/voucher/<int:id>/estado', methods=['POST'])
@login_required
def update_omada_voucher(id):
    voucher = OmadaVoucher.query.get_or_404(id)
    
    # Validar permisos
    if current_user.rol != 'admin' and voucher.vendedor_id != current_user.id:
        return jsonify({'success': False, 'error': 'No tienes permisos para editar esta ficha.'}), 403
        
    data = request.json
    nuevo_estado = data.get('estado')
    cliente_nombre = data.get('cliente_nombre')
    
    if nuevo_estado in ['activo', 'usado', 'vencido', 'eliminado']:
        voucher.estado = nuevo_estado
        if nuevo_estado == 'usado' and not voucher.fecha_uso:
            voucher.fecha_uso = datetime.utcnow()
            
    if cliente_nombre is not None:
        voucher.cliente_nombre = cliente_nombre
        
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/config/alertas', methods=['POST'])
@login_required
@admin_required
def save_config_alertas():
    """Guardar configuración de Alertas TextMeBot"""
    data = request.json
    config = ConfigAlertas.query.first()
    if not config:
        config = ConfigAlertas()
        db.session.add(config)
    
    config.activo = data.get('activo', False)
    config.api_key = data.get('api_key', '')
    config.telefono_destino = data.get('telefono_destino', '')
    config.intervalo_minutos = int(data.get('intervalo_minutos', 2))
    
    db.session.commit()
    
    # Reiniciar el scheduler si está ejecutándose para aplicar cambios de intervalo
    try:
        iniciar_scheduler()
    except Exception as e:
        print("No se pudo reiniciar el scheduler:", e)
        
    return jsonify({'success': True})

@app.route('/admin/hotspot/fichas/generar_masivo', methods=['POST'])
@login_required
@admin_required
def hotspot_generar_masivo():
    import random
    import string
    
    router_id = request.form.get('router_id')
    cantidad = int(request.form.get('cantidad', 1))
    modo = request.form.get('modo', 'pin')
    longitud = int(request.form.get('longitud', 5))
    prefijo = request.form.get('prefijo', '').strip()
    caracteres = request.form.get('caracteres', 'alphanum_lower')
    perfil_nombre = request.form.get('perfil_nombre')
    precio_venta = float(request.form.get('precio_venta', 0))
    limit_uptime = request.form.get('limit_uptime', '').strip()
    limit_bytes = request.form.get('limit_bytes', '').strip()
    comentario = request.form.get('comentario', '').strip()
    
    if not limit_uptime: limit_uptime = None
    if not limit_bytes: limit_bytes = None
    if not comentario: comentario = f"Gen {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Auto-crear o buscar el Plan internamente para mantener la integridad de la Base de Datos
    plan = PlanHotspot.query.filter_by(perfil_hotspot=perfil_nombre).first()
    if not plan:
        plan = PlanHotspot(nombre=perfil_nombre, precio=precio_venta, perfil_hotspot=perfil_nombre)
        db.session.add(plan)
        db.session.commit()
    elif plan.precio != precio_venta:
        # Actualizar precio si cambió
        plan.precio = precio_venta
        db.session.commit()
        
    router = ConfigMikroTik.query.get_or_404(router_id)
    
    api = MikroTikAPI(router.host, router.username, router.password, router.port, router.use_ssl)
    conectado, msg = api.test_connection()
    if not conectado:
        flash(f'Error de conexión con el MikroTik {router.nombre}. Revisa las credenciales.', 'error')
        return redirect(url_for('hotspot_fichas'))
        
    es_v6 = "v6 API" in msg
    v6_api = None
    v6_connection = None
    if es_v6:
        try:
            import routeros_api
            v6_connection = routeros_api.RouterOsApiPool(router.host, username=router.username, password=router.password, port=router.port, plaintext_login=True)
            v6_api = v6_connection.get_api().get_resource('/ip/hotspot/user')
        except Exception as e:
            flash(f'Error estableciendo sesión v6 masiva: {str(e)}', 'error')
            return redirect(url_for('hotspot_fichas'))
            
    def generate_random_string(length, char_type):
        if char_type == 'lower':
            chars = string.ascii_lowercase
        elif char_type == 'upper':
            chars = string.ascii_uppercase
        elif char_type == 'numeric':
            chars = string.digits
        elif char_type == 'alphanum_lower':
            chars = string.ascii_lowercase + string.digits
        elif char_type == 'alphanum_upper':
            chars = string.ascii_uppercase + string.digits
        else:
            chars = string.ascii_lowercase + string.digits
        return ''.join(random.choice(chars) for _ in range(length))
        
    exitos = 0
    errores = 0
    
    for _ in range(cantidad):
        max_intentos = 10
        creado = False
        
        for _ in range(max_intentos):
            codigo = prefijo + generate_random_string(longitud, caracteres)
            username = codigo
            password = codigo if modo == 'pin' else generate_random_string(longitud, caracteres)
            
            if es_v6:
                try:
                    data = {
                        "name": username, 
                        "password": password, 
                        "profile": plan.perfil_hotspot, 
                        "comment": comentario
                    }
                    if limit_uptime: data["limit-uptime"] = limit_uptime
                    if limit_bytes: data["limit-bytes-total"] = limit_bytes
                    v6_api.add(**data)
                    success = True
                    msg_or_id = "v6_gen"
                except Exception as e:
                    success = False
                    msg_or_id = str(e)
            else:
                success, msg_or_id = api.create_hotspot_user(
                    name=username,
                    password=password,
                    profile=plan.perfil_hotspot,
                    comment=comentario,
                    limit_uptime=limit_uptime,
                    limit_bytes_total=limit_bytes
                )
            
            if success:
                creado = True
                # Guardar en base de datos local
                v = Voucher(
                    codigo=username,
                    contrasena=password,
                    precio=plan.precio,
                    plan_id=plan.id,
                    router_id=router.id,
                    vendedor_id=current_user.id,
                    estado='activo',
                    mikrotik_id=msg_or_id
                )
                db.session.add(v)
                exitos += 1
                break
                
        if not creado:
            errores += 1
            
    if v6_connection:
        try: v6_connection.disconnect()
        except: pass
            
    db.session.commit()
    
    if errores == 0:
        flash(f'¡Éxito! Se generaron {exitos} fichas correctamente en {router.nombre}.', 'success')
    else:
        flash(f'Se generaron {exitos} fichas, pero hubo {errores} errores (posiblemente códigos duplicados o error de conexión).', 'warning')
        
    return redirect(url_for('hotspot_fichas'))

@app.route('/admin/hotspot/fichas/generar_masivo_ajax', methods=['POST'])
@login_required
@admin_required
def hotspot_generar_masivo_ajax():
    import random
    import string
    
    router_id = request.form.get('router_id')
    cantidad = int(request.form.get('cantidad', 1))
    modo = request.form.get('modo', 'pin')
    longitud = int(request.form.get('longitud', 5))
    prefijo = request.form.get('prefijo', '').strip()
    caracteres = request.form.get('caracteres', 'alphanum_lower')
    perfil_nombre = request.form.get('perfil_nombre')
    precio_venta = float(request.form.get('precio_venta', 0))
    limit_uptime = request.form.get('limit_uptime', '').strip()
    limit_bytes = request.form.get('limit_bytes', '').strip()
    comentario = request.form.get('comentario', '').strip()
    lote_uuid = request.form.get('lote_uuid', '').strip()
    
    if not limit_uptime: limit_uptime = None
    if not limit_bytes: limit_bytes = None
    if not comentario: comentario = f"Gen {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    plan = PlanHotspot.query.filter_by(perfil_hotspot=perfil_nombre).first()
    if not plan:
        plan = PlanHotspot(nombre=perfil_nombre, precio=precio_venta, perfil_hotspot=perfil_nombre)
        db.session.add(plan)
        db.session.commit()
    elif plan.precio != precio_venta:
        plan.precio = precio_venta
        db.session.commit()
        
    router = ConfigMikroTik.query.get(router_id)
    if not router:
        return jsonify({'error': 'Router no encontrado'})
        
    # Obtener o crear el lote
    lote = None
    if lote_uuid:
        lote = LoteFichas.query.filter_by(uuid=lote_uuid).first()
        if not lote:
            # cantidad_total viene del form para saber de cuánto era el lote originalmente
            cantidad_total = int(request.form.get('cantidad_total', cantidad))
            nombre_lote = f"Lote {plan.nombre} - {router.nombre}"
            lote = LoteFichas(
                uuid=lote_uuid, nombre=nombre_lote, cantidad_solicitada=cantidad_total,
                router_id=router.id, plan_id=plan.id, vendedor_id=current_user.id
            )
            db.session.add(lote)
            db.session.commit()
    
    api = MikroTikAPI(router.host, router.username, router.password, router.port, router.use_ssl)
    conectado, msg = api.test_connection()
    if not conectado:
        return jsonify({'error': f'Error de conexión con MikroTik: {msg}'})
        
    es_v6 = "v6 API" in msg
    v6_api = None
    v6_connection = None
    if es_v6:
        try:
            import routeros_api
            v6_connection = routeros_api.RouterOsApiPool(router.host, username=router.username, password=router.password, port=router.port, plaintext_login=True)
            v6_api = v6_connection.get_api().get_resource('/ip/hotspot/user')
        except Exception as e:
            return jsonify({'error': f'Error estableciendo sesión v6: {str(e)}'})
            
    def generate_random_string(length, char_type):
        if char_type == 'lower': chars = string.ascii_lowercase
        elif char_type == 'upper': chars = string.ascii_uppercase
        elif char_type == 'numeric': chars = string.digits
        elif char_type == 'alphanum_lower': chars = string.ascii_lowercase + string.digits
        elif char_type == 'alphanum_upper': chars = string.ascii_uppercase + string.digits
        else: chars = string.ascii_lowercase + string.digits
        return ''.join(random.choice(chars) for _ in range(length))
        
    exitos = 0
    errores = 0
    vouchers_to_save = []
    users_batch_v7 = []
    
    # Pre-generar todos los códigos asegurando unicidad básica
    codigos_generados = set()
    intentos_totales = 0
    while len(codigos_generados) < cantidad and intentos_totales < cantidad * 5:
        codigo = prefijo + generate_random_string(longitud, caracteres)
        codigos_generados.add(codigo)
        intentos_totales += 1
        
    for username in list(codigos_generados)[:cantidad]:
        password = username if modo == 'pin' else generate_random_string(longitud, caracteres)
        
        if es_v6:
            try:
                data = {"name": username, "password": password, "profile": plan.perfil_hotspot, "comment": comentario}
                if limit_uptime: data["limit-uptime"] = limit_uptime
                if limit_bytes: data["limit-bytes-total"] = limit_bytes
                v6_api.add(**data)
                
                v = Voucher(codigo=username, contrasena=password, precio=plan.precio, plan_id=plan.id, router_id=router.id, vendedor_id=current_user.id, estado='activo', mikrotik_id="v6_gen", lote_id=lote.id if lote else None)
                vouchers_to_save.append(v)
                exitos += 1
            except Exception as e:
                errores += 1
        else:
            # v7 Ráfaga
            users_batch_v7.append({
                "name": username,
                "password": password,
                "profile": plan.perfil_hotspot,
                "limit_uptime": limit_uptime,
                "limit_bytes_total": limit_bytes,
                "comment": comentario
            })
            
            v = Voucher(codigo=username, contrasena=password, precio=plan.precio, plan_id=plan.id, router_id=router.id, vendedor_id=current_user.id, estado='activo', mikrotik_id="batch_gen", lote_id=lote.id if lote else None)
            vouchers_to_save.append(v)
            
    # Ejecutar Ráfaga en v7 si es necesario
    if not es_v6 and users_batch_v7:
        success, msg = api.create_hotspot_users_batch(users_batch_v7)
        if success:
            exitos = len(users_batch_v7)
        else:
            return jsonify({'error': f"Error en generación Ráfaga MikroTik: {msg}"})
            
    if v6_connection:
        try: v6_connection.disconnect()
        except: pass
        
    if vouchers_to_save:
        try:
            db.session.bulk_save_objects(vouchers_to_save)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f"Error guardando en BD local: {str(e)}"})
    return jsonify({'exitos': exitos, 'errores': errores, 'success': True})

@app.route('/admin/hotspot/vendedor/guardar', methods=['POST'])
@login_required
@admin_required
def hotspot_vendedor_guardar():
    """Crear o editar un usuario con rol de vendedor"""
    import os
    import json
    from werkzeug.utils import secure_filename
    
    vendedor_id = request.form.get('vendedor_id')
    username = request.form.get('username', '').strip()
    nombre = request.form.get('nombre', '').strip()
    password = request.form.get('password', '')
    router_id = request.form.get('router_id')
    
    # Nuevos campos
    telefono = request.form.get('telefono', '').strip()
    direccion_depto = request.form.get('direccion_depto', '').strip()
    direccion_muni = request.form.get('direccion_muni', '').strip()
    direccion_aldea = request.form.get('direccion_aldea', '').strip()
    direccion = json.dumps({'depto': direccion_depto, 'muni': direccion_muni, 'aldea': direccion_aldea}) if direccion_depto else ''
    
    tipo_vendedor = request.form.get('tipo_vendedor', 'acumulativo')
    comision_tipo = request.form.get('comision_tipo', 'porcentaje')
    comision_valor = float(request.form.get('comision_valor', 0.0) or 0.0)
    limite_fichas = int(request.form.get('limite_fichas', 0) or 0)
    estado = request.form.get('estado', 'activo')
    
    # Permisos
    permisos = {
        'vender_fichas': 'vender_fichas' in request.form,
        'ver_reportes': 'ver_reportes' in request.form,
        'imprimir': 'imprimir' in request.form,
        'solicitar_inventario': 'solicitar_inventario' in request.form,
        'editar_clientes': 'editar_clientes' in request.form,
        'ver_ganancias': 'ver_ganancias' in request.form,
        'administrar_omada': 'administrar_omada' in request.form
    }
    permisos_json = json.dumps(permisos)
    
    if router_id:
        router_id = int(router_id)
    else:
        router_id = None
        
    foto_filename = None
    if 'foto_perfil' in request.files:
        file = request.files['foto_perfil']
        if file.filename != '':
            filename = secure_filename(f"{username}_{file.filename}")
            upload_folder = os.path.join(app.root_path, 'static', 'uploads', 'vendedores')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder, exist_ok=True)
            file.save(os.path.join(upload_folder, filename))
            foto_filename = f"uploads/vendedores/{filename}"

    if vendedor_id:
        vendedor = Usuario.query.get_or_404(vendedor_id)
        vendedor.username = username
        vendedor.nombre = nombre
        vendedor.router_id = router_id
        vendedor.telefono = telefono
        vendedor.direccion = direccion
        vendedor.tipo_vendedor = tipo_vendedor
        vendedor.comision_tipo = comision_tipo
        vendedor.comision_valor = comision_valor
        vendedor.limite_fichas = limite_fichas
        vendedor.estado = estado
        vendedor.permisos = permisos_json
        
        if foto_filename:
            vendedor.foto_perfil = foto_filename
            
        if password:
            vendedor.set_password(password)
        flash('Datos del vendedor actualizados', 'success')
    else:
        if Usuario.query.filter_by(username=username).first():
            flash(f'¡ATENCIÓN! El nombre de usuario "{username}" ya está registrado en el sistema. Usa otro diferente.', 'error')
            return redirect(url_for('hotspot_vendedor_nuevo'))
            
        vendedor = Usuario(
            username=username,
            nombre=nombre,
            rol='vendedor',
            router_id=router_id,
            activo=True,
            balance=0.0,
            telefono=telefono,
            direccion=direccion,
            tipo_vendedor=tipo_vendedor,
            comision_tipo=comision_tipo,
            comision_valor=comision_valor,
            limite_fichas=limite_fichas,
            estado=estado,
            permisos=permisos_json,
            foto_perfil=foto_filename
        )
        vendedor.set_password(password)
        db.session.add(vendedor)
        flash('Vendedor creado con éxito', 'success')
        
    db.session.commit()
    registrar_auditoria('guardar_vendedor', 'usuarios', vendedor.id, f'Vendedor: {username}')
    
    # Redirección según botón presionado
    if request.form.get('action_type') == 'guardar_y_nuevo':
        return redirect(url_for('hotspot_vendedor_nuevo'))
        
    return redirect(url_for('hotspot_vendedores'))


@app.route('/admin/hotspot/vendedor/cargar-saldo', methods=['POST'])
@login_required
@admin_required
def hotspot_vendedor_cargar_saldo():
    """Cargar balance de saldo a un vendedor"""
    vendedor_id = int(request.form.get('vendedor_id', 0))
    monto = float(request.form.get('monto', 0))
    descripcion = request.form.get('descripcion', '').strip() or 'Carga de saldo autorizada'
    
    if monto <= 0:
        flash('El monto debe ser mayor a cero', 'error')
        return redirect(url_for('hotspot_admin'))
        
    vendedor = Usuario.query.get_or_404(vendedor_id)
    vendedor.balance += monto
    
    transaccion = TransaccionVendedor(
        vendedor_id=vendedor.id,
        tipo='carga',
        monto=monto,
        descripcion=descripcion
    )
    db.session.add(transaccion)
    db.session.commit()
    
    registrar_auditoria('cargar_saldo', 'usuarios', vendedor.id, f'Cargados Q{monto} a {vendedor.username}')
    flash(f'Saldo cargado exitosamente. Nuevo balance de {vendedor.nombre}: Q{vendedor.balance:.2f}', 'success')
    return redirect(url_for('hotspot_admin'))


# ============== VISTAS Y LOGICA DEL VENDEDOR ==============

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    # Para no interferir con las excepciones de Werkzeug (como 404) que son subclases de HTTPException
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    return "<div style='padding:20px; background:black; color:lime; font-family:monospace; white-space:pre-wrap;'>" + traceback.format_exc() + "</div>", 500

@app.route('/vendedor/dashboard')
@login_required
@vendedor_required
def vendedor_dashboard():
    """Panel del vendedor para generar vouchers"""
    from datetime import datetime, timedelta
    
    planes = PlanHotspot.query.filter_by(activo=True).all()
    router = ConfigMikroTik.query.get(current_user.router_id) if current_user.router_id else None
    
    # Obtener vouchers recientes
    vouchers = Voucher.query.filter_by(vendedor_id=current_user.id).order_by(Voucher.fecha_creacion.desc()).limit(15).all()
    
    # Calcular Métricas (Fase 2)
    hoy = datetime.utcnow().date()
    mes_actual = hoy.month
    año_actual = hoy.year
    
    # Fichas Mikrotik vendidas hoy
    fichas_hoy_mt = Voucher.query.filter(
        Voucher.vendedor_id == current_user.id,
        db.func.date(Voucher.fecha_creacion) == hoy
    ).all()
    
    # Fichas Omada vendidas hoy (estado usado/vencido)
    fichas_hoy_om = OmadaVoucher.query.filter(
        OmadaVoucher.vendedor_id == current_user.id,
        db.func.date(OmadaVoucher.fecha_uso) == hoy,
        OmadaVoucher.estado.in_(['usado', 'vencido'])
    ).all()
    
    ventas_hoy_total = sum(v.precio for v in fichas_hoy_mt) + sum(o.precio for o in fichas_hoy_om)
    
    # Fichas Mikrotik vendidas este mes
    fichas_mes_mt = Voucher.query.filter(
        Voucher.vendedor_id == current_user.id,
        db.func.extract('month', Voucher.fecha_creacion) == mes_actual,
        db.func.extract('year', Voucher.fecha_creacion) == año_actual
    ).all()
    
    # Fichas Omada vendidas este mes
    fichas_mes_om = OmadaVoucher.query.filter(
        OmadaVoucher.vendedor_id == current_user.id,
        db.func.extract('month', OmadaVoucher.fecha_uso) == mes_actual,
        db.func.extract('year', OmadaVoucher.fecha_uso) == año_actual,
        OmadaVoucher.estado.in_(['usado', 'vencido'])
    ).all()
    
    ventas_mes_total = sum(v.precio for v in fichas_mes_mt) + sum(o.precio for o in fichas_mes_om)
    
    # Calcular Ganancias
    ganancias_mes = 0
    if current_user.comision_tipo == 'porcentaje':
        ganancias_mes = ventas_mes_total * (current_user.comision_valor / 100.0)
    else:
        # Fijo por ficha vendida
        total_fichas_vendidas = len(fichas_mes_mt) + len(fichas_mes_om)
        ganancias_mes = total_fichas_vendidas * current_user.comision_valor
        
    # Inventario actual
    inventario_actual = 0
    limite = current_user.limite_fichas or 0
    if limite > 0:
        fichas_activas_mt = Voucher.query.filter_by(vendedor_id=current_user.id, activo=True).count()
        fichas_disponibles_om = OmadaVoucher.query.filter_by(vendedor_id=current_user.id, estado='disponible').count()
        inventario_actual = limite - (fichas_activas_mt + fichas_disponibles_om)
        if inventario_actual < 0:
            inventario_actual = 0
            
    # Datos para gráfico (Últimos 7 días)
    ventas_grafico = []
    dias_grafico = []
    for i in range(6, -1, -1):
        dia = hoy - timedelta(days=i)
        dias_grafico.append(dia.strftime('%d/%m'))
        
        v_mt = Voucher.query.filter(Voucher.vendedor_id == current_user.id, db.func.date(Voucher.fecha_creacion) == dia).all()
        v_om = OmadaVoucher.query.filter(OmadaVoucher.vendedor_id == current_user.id, db.func.date(OmadaVoucher.fecha_uso) == dia, OmadaVoucher.estado.in_(['usado', 'vencido'])).all()
        
        total_dia = sum(v.precio for v in v_mt) + sum(o.precio for o in v_om)
        ventas_grafico.append(total_dia)
    
    return render_template('vendedor_dashboard.html',
                           planes=planes,
                           router=router,
                           vouchers=vouchers,
                           ventas_hoy=ventas_hoy_total,
                           ventas_mes=ventas_mes_total,
                           ganancias_mes=ganancias_mes,
                           inventario_actual=inventario_actual,
                           limite_fichas=limite,
                           dias_grafico=dias_grafico,
                           ventas_grafico=ventas_grafico)


@app.route('/api/hotspot/generar', methods=['POST'])
@login_required
@vendedor_required
def hotspot_generar_voucher():
    """API para generar un voucher y enviarlo al MikroTik"""
    import random
    import string
    
    data = request.get_json() or {}
    plan_id = int(data.get('plan_id', 0))
    
    plan = PlanHotspot.query.get_or_404(plan_id)
    if not plan.activo:
        return jsonify({'success': False, 'error': 'El plan seleccionado no está activo'})
        
    if current_user.balance < plan.precio:
        return jsonify({'success': False, 'error': f'Saldo insuficiente. El plan cuesta Q{plan.precio:.2f} y tu saldo es Q{current_user.balance:.2f}'})
        
    if not current_user.router_id:
        return jsonify({'success': False, 'error': 'No tienes ningún router MikroTik asignado para vender vouchers. Contacta al administrador.'})
        
    api = get_mikrotik_api(current_user.router_id)
    if not api:
        return jsonify({'success': False, 'error': 'No se pudo establecer conexión con el router MikroTik. Verifica si está en línea.'})
        
    caracteres_codigo = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    codigo_random = "".join(random.choices(caracteres_codigo, k=4))
    codigo = f"CZ-{codigo_random}"
    contrasena = "".join(random.choices(caracteres_codigo, k=6))
    
    comment_vendedor = f"Vendido por {current_user.username}"
    
    success, result_id = api.create_hotspot_user(
        name=codigo,
        password=contrasena,
        profile=plan.perfil_hotspot,
        comment=comment_vendedor,
        limit_uptime=plan.limit_uptime
    )
    
    if not success:
        return jsonify({'success': False, 'error': f'MikroTik rechazó la creación del voucher: {result_id}'})
        
    current_user.balance -= plan.precio
    
    voucher = Voucher(
        codigo=codigo,
        contrasena=contrasena,
        precio=plan.precio,
        plan_id=plan.id,
        router_id=current_user.router_id,
        vendedor_id=current_user.id,
        estado='activo',
        mikrotik_id=result_id
    )
    db.session.add(voucher)
    
    transaccion = TransaccionVendedor(
        vendedor_id=current_user.id,
        tipo='venta',
        monto=plan.precio,
        descripcion=f"Venta voucher {codigo} (Plan: {plan.nombre})"
    )
    db.session.add(transaccion)
    
    try:
        db.session.commit()
        registrar_auditoria('generar_voucher', 'vouchers', voucher.id, f'Voucher {codigo} generado por {current_user.username}')
        
        return jsonify({
            'success': True,
            'voucher': {
                'id': voucher.id,
                'codigo': codigo,
                'contrasena': contrasena,
                'plan': plan.nombre,
                'precio': plan.precio,
                'router': voucher.router.nombre if voucher.router else 'Hotspot',
                'fecha': voucher.fecha_creacion.strftime('%Y-%m-%d %H:%M')
            },
            'nuevo_saldo': current_user.balance
        })
    except Exception as e:
        db.session.rollback()
        try:
            api.delete_hotspot_user(result_id)
        except Exception:
            pass
        return jsonify({'success': False, 'error': f'Error al registrar el voucher en el sistema: {str(e)}'})


@app.route('/vendedor/voucher/imprimir/<int:id>')
@login_required
@vendedor_required
def vendedor_voucher_imprimir(id):
    """Vista para imprimir un ticket individual de voucher"""
    voucher = Voucher.query.get_or_404(id)
    if voucher.vendedor_id != current_user.id:
        flash('No tienes permiso para ver este voucher', 'error')
        return redirect(url_for('vendedor_dashboard'))
    return render_template('imprimir_voucher.html', voucher=voucher)


# ============== WEBHOOK OMADA NOC ==============

def analizar_alerta_con_ia(alerta_id, contexto_alerta):
    """Analiza una alerta usando Google Gemini en segundo plano"""
    with app.app_context():
        alerta = AlertaOmada.query.get(alerta_id)
        if not alerta:
            return
            
        if not GEMINI_API_KEY:
            alerta.analisis_ia = "API Key de Gemini no configurada en DigitalOcean."
            db.session.commit()
            return
            
        try:
            model = genai.GenerativeModel('gemini-1.5-flash')
            prompt = f"Eres un experto NOC en redes ISP y equipos TP-Link Omada. Se detectó esta alerta:\n\nDispositivo: {alerta.dispositivo}\nEstado: {alerta.estado}\nMensaje Original: {alerta.mensaje}\nDetalles JSON: {contexto_alerta}\n\nEnumera 3 posibles causas técnicas precisas de esta falla y 2 acciones inmediatas a revisar. Sé muy breve, profesional y responde en español."
            
            response = model.generate_content(prompt)
            alerta.analisis_ia = response.text
            db.session.commit()
        except Exception as e:
            print(f"Error en análisis de IA Gemini: {e}")
            alerta.analisis_ia = f"Error al consultar la IA: {str(e)}"
            db.session.commit()


@app.route('/api/webhook/omada', methods=['POST'])
@app.route('/api/webhook/omada/', methods=['POST'])
def webhook_omada():
    """Recibe las alertas de TP-Link Omada Controller"""
    try:
        # Omada a veces no envía Content-Type: application/json, forzamos su lectura
        data = request.get_json(force=True, silent=True)
        if not data and request.data:
            try:
                data = json.loads(request.data)
            except:
                data = {}
        data = data or {}
        
        # Extraer información basándonos en el JSON real de Omada
        mensaje_crudo = data.get('text') or data.get('description') or data.get('message') or data.get('content') or 'Alerta sin mensaje'
        dispositivo = data.get('Site') or data.get('site') or data.get('deviceMac') or data.get('apMac') or 'Desconocido'
        
        # Limpiar el mensaje si viene como una cadena JSON interna (ej. [{"operation": "... "}])
        mensaje_limpio = mensaje_crudo
        if isinstance(mensaje_crudo, str) and mensaje_crudo.strip().startswith('['):
            try:
                parsed_text = json.loads(mensaje_crudo)
                if isinstance(parsed_text, list) and len(parsed_text) > 0:
                    item = parsed_text[0]
                    mensaje_limpio = item.get('operation') or item.get('content') or mensaje_crudo
            except Exception:
                pass

        # Determinar el tipo de alerta basado en el texto del mensaje
        estado_lower = mensaje_limpio.lower()
        if 'off' in estado_lower or 'disconnect' in estado_lower or 'fail' in estado_lower or 'err' in estado_lower:
            tipo_alerta = 'OFFLINE/ERROR'
        elif 'connect' in estado_lower or 'success' in estado_lower:
            tipo_alerta = 'ONLINE/OK'
        else:
            tipo_alerta = 'INFO'
        
        # Crear la alerta
        nueva_alerta = AlertaOmada(
            dispositivo=dispositivo,
            estado=tipo_alerta,
            mensaje=mensaje_limpio,
            raw_payload=json.dumps(data)
        )
        db.session.add(nueva_alerta)
        db.session.commit()
        
        # Lanzar análisis de IA en segundo plano solo si es una alerta importante o error
        if tipo_alerta == 'OFFLINE/ERROR':
            threading.Thread(target=analizar_alerta_con_ia, args=(nueva_alerta.id, json.dumps(data))).start()
        
        return jsonify({'success': True, 'message': 'Alerta registrada correctamente'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/noc')
@login_required
@admin_required
def admin_noc():
    """Panel NOC para ver las alertas de la red"""
    alertas = AlertaOmada.query.order_by(AlertaOmada.fecha.desc()).limit(100).all()
    return render_template('admin_noc.html', alertas=alertas)


@app.route('/sw.js')
def serve_sw():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'sw.js', mimetype='application/javascript')


@app.route('/manifest.json')
def serve_manifest():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'manifest.json', mimetype='application/json')

@app.route('/api/noc/stats')
@login_required
def noc_stats():
    """Endpoint para el panel NOC en el Dashboard principal"""
    import random
    
    clientes_totales = Cliente.query.filter_by(estado='activo').count()
    # Simular tráfico activo (aprox 80% de los clientes) o 54 si no hay clientes
    clientes_online = int(clientes_totales * 0.8) if clientes_totales > 0 else 54
    
    # Simulación de tráfico en vivo (Mbps)
    trafico_down = random.randint(250, 450)
    trafico_up = random.randint(50, 150)
    ping = random.randint(1, 8)
    ups = random.randint(95, 100)
    return jsonify({
        'clientes_online': clientes_online,
        'trafico_down': trafico_down,
        'trafico_up': trafico_up,
        'ping': ping,
        'ups': ups,
        'starlink': 'Online'
    })

def enviar_alerta_textmebot(mensaje):
    import requests
    with app.app_context():
        config_alertas = ConfigAlertas.query.first()
        if not config_alertas or not config_alertas.activo or not config_alertas.api_key or not config_alertas.telefono_destino:
            return
            
        url = "http://api.textmebot.com/send.php"
        params = {
            "recipient": config_alertas.telefono_destino,
            "apikey": config_alertas.api_key,
            "text": mensaje
        }
        try:
            requests.get(url, params=params, timeout=10)
        except Exception as e:
            print("Error enviando alerta TextMeBot:", e)

def monitorear_routers():
    from datetime import datetime
    with app.app_context():
        config_alertas = ConfigAlertas.query.first()
        if not config_alertas or not config_alertas.activo:
            return
            
        nodos = ConfigMikroTik.query.filter_by(activo=True).all()
        for nodo in nodos:
            api = get_mikrotik_api(nodo.id)
            try:
                conectado, _ = api.conectar()
            except:
                conectado = False
                
            ahora = datetime.utcnow()
            
            if conectado:
                if not getattr(nodo, 'estado_online', True):
                    nodo.estado_online = True
                    tiempo_caido = ""
                    if getattr(nodo, 'ultima_caida', None):
                        minutos = int((ahora - nodo.ultima_caida).total_seconds() / 60)
                        tiempo_caido = f" (Estuvo offline por {minutos} min)"
                    
                    mensaje = f"✅ RECUPERADO: El router '{nodo.nombre}' ({nodo.host}) está nuevamente en línea.{tiempo_caido}"
                    enviar_alerta_textmebot(mensaje)
            else:
                if getattr(nodo, 'estado_online', True):
                    nodo.estado_online = False
                    nodo.ultima_caida = ahora
                    mensaje = f"🚨 ALERTA CUZONET: El router '{nodo.nombre}' ({nodo.host}) se ha desconectado. Por favor, verifica la red."
                    enviar_alerta_textmebot(mensaje)
                    
        db.session.commit()

def iniciar_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler()
        
        # Intervalo dinámico si existe en config
        intervalo = 2
        with app.app_context():
            config = ConfigAlertas.query.first()
            if config and config.intervalo_minutos:
                intervalo = config.intervalo_minutos
                
        scheduler.add_job(func=monitorear_routers, trigger="interval", minutes=intervalo, id='monitoreo_mikrotik', replace_existing=True)
        scheduler.start()
        print("Scheduler de monitoreo iniciado.")
    except ImportError:
        print("ADVERTENCIA: APScheduler no está instalado. No se ejecutarán las alertas.")

# Inicializar scheduler al cargar la app (para Gunicorn)
try:
    with app.app_context():
        db.create_all()
        # Agregar nuevas columnas si no existen
        try:
            from sqlalchemy import text
            db.session.execute(text("ALTER TABLE inventario_manual_lote ADD COLUMN vendedor_asignado VARCHAR(100) DEFAULT ''"))
            db.session.execute(text("ALTER TABLE inventario_manual_lote ADD COLUMN fecha_asignacion VARCHAR(50) DEFAULT ''"))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print("Notice de DB (puede ser normal si las columnas ya existen):", e)
        print("Tablas de base de datos verificadas/creadas con éxito.")
    
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or os.getenv('FLASK_ENV', 'production') != 'development':
        iniciar_scheduler()
except Exception as e:
    print("Error en inicialización (DB o Scheduler):", e)

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug)
